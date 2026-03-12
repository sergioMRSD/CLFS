"""
CLFS_Manager.py

Orchestrator that loads CLFS_newformat.csv and answer.json, groups household members,
invokes validators from CLFS_Brain.py, applies corrections, writes a _validated.xlsx with
highlights and a _report.xlsx summary.

Usage:
    python3 CLFS_Manager.py   # runs main pipeline (will look for CLFS_newformat.csv and answer.json)

"""
from pathlib import Path
import json
from typing import List, Dict, Any
import pandas as pd  # type: ignore
from openpyxl.styles import PatternFill  # type: ignore
from openpyxl import load_workbook  # type: ignore

from CLFS_Brain import (
    validate_identification_type,
    validate_residential_st,
    validate_h_sep_y,
    validate_activity_status,
    validate_i_l,
    validate_none_of_the_above_exclusive,
    validate_travel_time_format,
    validate_years_in_employment_consistency,
    validate_num_children,
    validate_oaw_income_threshold,
    validate_occupation_details,
    validate_employment_consistency,
    validate_seeking_work_logic,
    validate_duration_numeric,
    ValidationResult,
)

ROOT = Path(__file__).parent
CSV_PATH = ROOT / "CLFS_newformat.csv"
ANSWER_JSON = ROOT / "answer.json"

YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def _find_household_id_column(df: pd.DataFrame) -> str:
    # In this CSV layout the first two columns are metadata; start search from column index 2
    cols = list(df.columns)[2:]
    candidates = [c for c in cols if c.lower() in ('household_id', 'hhid', 'household', 'hh_id', 'hhid_str', 'household id')]
    if candidates:
        return str(candidates[0])
    # fallback: pick the first non-metadata column if available, otherwise first column
    return str(cols[0]) if cols else str(df.columns[0])


def _extract_identification_options(answer_json: Dict[str, Any]) -> List[str]:
    # Heuristic: find any fieldOptions lists that contain 'Singapore Citizen'
    def walk(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k == 'fieldOptions' and isinstance(v, list):
                    yield v
                else:
                    yield from walk(v)
        elif isinstance(obj, list):
            for item in obj:
                yield from walk(item)
    for options in walk(answer_json):
        try:
            if any('Singapore Citizen' in str(opt) for opt in options):
                return [str(o) for o in options]
        except Exception:
            continue
    # fallback empty
    return []


def run_pipeline(csv_path: Path = CSV_PATH, answer_json_path: Path = ANSWER_JSON):
    if not csv_path.exists():
        print(f"Input CSV not found: {csv_path}")
        return
    # Read CSV using the real header row (row 6) and skip the top metadata rows
    # Try tab-separated first (common for exported surveys), then fallback to python engine
    try:
        df = pd.read_csv(csv_path, sep='\t', skiprows=5)
    except Exception as e:
        print(f"TSV parse failed: {e}; attempting fallback parsing with python engine and sep='\t' (skiprows=5)")
        try:
            df = pd.read_csv(csv_path, engine='python', sep='\t', skiprows=5)
        except Exception as e2:
            print(f"Fallback TSV parse failed: {e2}; attempting generic sep=None parse")
            try:
                df = pd.read_csv(csv_path, engine='python', sep=None, skiprows=5)
            except Exception as e3:
                print(f"Fallback generic parse failed: {e3}")
                return
    print(f"Loaded {len(df)} rows from {csv_path.name}")

    hh_col = _find_household_id_column(df)
    print(f"Using household id column: {hh_col}")

    # Load answer.json for identification options
    options = []
    if answer_json_path.exists():
        with open(answer_json_path, 'r', encoding='utf-8') as f:
            aj = json.load(f)
            options = _extract_identification_options(aj)
    else:
        print("Warning: answer.json not found; identification validation will be less strict")

    # Prepare book-keeping
    errors = []  # list of dicts: row, col, rule, message
    corrections = []

    # We'll operate on a copy to apply corrections
    mod_df = df.copy()

    # Group by household and iterate
    for hh, grp in df.groupby(hh_col):
        # collect household ethnicities for language rule — heuristic: look for column names containing 'ethnic'
        eth_cols = [c for c in df.columns if 'ethnic' in c.lower()]
        household_ethnicities = []
        if eth_cols:
            household_ethnicities = [str(x) for x in grp[eth_cols].stack().dropna().astype(str).tolist()]

        for idx, row in grp.iterrows():
            # Example candidate columns (best-effort): Identification Type, Where are you currently staying?, Marital Status,
            # Labour Force Status, Are you actively looking for a new job?, PostalCode, TravelTime, TotalYearsEmployed,
            # YearsCurrentJob, Age, AgeStartedEmployment, NumChildren, PrimaryLanguage, EmploymentStatus, MonthlyIncome

            # we try to map a few common column names (case-insensitive)
            def _get(col_candidates):
                # Ignore first two metadata columns when searching for fields
                columns_to_search = list(df.columns)[2:]
                lower_candidates = [x.lower() for x in col_candidates]
                for c in columns_to_search:
                    if c.lower() in lower_candidates:
                        return row[c]
                return None

            identification = _get(['Identification Type', 'identification_type', 'id_type'])
            res_st = _get(['Where are you currently staying?', 'where_currently_staying', 'residential_status'])
            marital = _get(['Marital Status', 'marital_status'])
            labour = _get(['Labour Force Status', 'labour_force_status', 'activity_status'])
            looking = _get(['Are you actively looking for a new job?', 'looking_for_job', 'i_l'])
            # postal code check removed per final cleanup
            travel = _get(['TravelTime', 'travel_time', 'travel_minutes'])
            total_years = _get(['TotalYearsEmployed', 'total_years'])
            years_current = _get(['YearsCurrentJob', 'years_current'])
            age = _get(['Age', 'age'])
            age_started = _get(['At what age did you start employment'])
            num_children = _get(['Number of children given birth to'])
            employment_status = _get(['Employment Status', 'employment_status'])
            monthly_income = _get(['Last drawn GMI'])
            # Next-15 candidate fields (best-effort mapping)
            e_occ = _get(['What kind of occupation were you looking for?'])
            w_desc = _get(['Main tasks / duties'])
            emp_flag = _get(['_EMP_', 'is_employed', 'employed_flag'])
            e_empst = _get(['E_EMPST', 'employment_status_detail'])
            empst = _get(['_EMPST', 'emp_status'])
            u_l = _get(['U_L', 'actively_looking'])
            u_w = _get(['U_W', 'available_to_start'])
            i_w = _get(['I_W', 'inability_to_work'])
            e_w = _get(['E_W', 'other_availability_flag'])
            duration = _get(['_DUR', 'duration', 'duration_years'])

            # Run the validators from CLFS_Brain
            # Survey rules
            vr = validate_identification_type(identification, allowed_options=options)
            if not vr.is_valid:
                errors.append({'row': idx, 'col': 'Identification Type', 'rule': vr.rule_applied, 'message': vr.message})
            elif vr.corrected_value is not None and vr.corrected_value != identification:
                mod_df.at[idx, 'Identification Type'] = vr.corrected_value
                corrections.append({'row': idx, 'col': 'Identification Type', 'from': identification, 'to': vr.corrected_value, 'rule': vr.rule_applied})

            vr = validate_residential_st(res_st)
            if not vr.is_valid:
                errors.append({'row': idx, 'col': 'Where are you currently staying?', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_h_sep_y(marital)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'Marital Status', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_activity_status(labour)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'Labour Force Status', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_i_l(looking)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'Looking for job', 'rule': vr.rule_applied, 'message': vr.message})

            # Next-15 validators
            vr = validate_occupation_details(employment_status, e_occ, w_desc)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'Occupation', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_employment_consistency(emp_flag, e_empst, empst, labour)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'EmploymentConsistency', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_seeking_work_logic(u_l, u_w, i_w, e_w)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'SeekingWork', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_duration_numeric(duration)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'Duration', 'rule': vr.rule_applied, 'message': vr.message})

            # Logic rules
            vr = validate_none_of_the_above_exclusive(_get(['SomeMultiSelectQuestion','multi_select']))
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'MultiSelect', 'rule': vr.rule_applied, 'message': vr.message})

            # postal code validation removed

            vr = validate_travel_time_format(travel)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'TravelTime', 'rule': vr.rule_applied, 'message': vr.message})

            # Cross-field rules
            vr = validate_years_in_employment_consistency(total_years, years_current, age, age_started)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'EmploymentYears', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_num_children(num_children, age)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'NumChildren', 'rule': vr.rule_applied, 'message': vr.message})

            # language-vs-ethnicity validation removed

            # Financial rules
            vr = validate_oaw_income_threshold(employment_status, monthly_income)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'MonthlyIncome', 'rule': vr.rule_applied, 'message': vr.message})

    # Write validated dataframe to Excel
    out_validated = ROOT / (csv_path.stem + "_validated.xlsx")
    mod_df.to_excel(out_validated, index=False)

    # Apply highlights for errors and corrections
    wb = load_workbook(out_validated)
    ws = wb.active
    if ws is None:
        print(f"Warning: workbook has no active sheet; skipping highlights for {out_validated.name}")
    else:
        # Map dataframe column name to excel col index
        col_to_idx = {c: i + 1 for i, c in enumerate(mod_df.columns)}

        for e in errors:
            r = int(e['row']) + 2  # header row + 1 (pandas -> excel)
            c = col_to_idx.get(e['col'], None)
            if c is None:
                continue
            cell = ws.cell(row=r, column=c)
            cell.fill = YELLOW_FILL
        for c in corrections:
            r = int(c['row']) + 2
            col = c['col']
            ci = col_to_idx.get(col, None)
            if ci is None:
                continue
            ws.cell(row=r, column=ci).fill = YELLOW_FILL

        wb.save(out_validated)
    print(f"Wrote validated workbook: {out_validated.name}")

    # Build a report summary: counts by rule and by household
    import collections
    rule_counts = collections.Counter([e['rule'] for e in errors if e.get('rule')])
    df_rules = pd.DataFrame([{'rule': r, 'count': c} for r, c in rule_counts.items()])
    df_corrections = pd.DataFrame(corrections)

    out_report = ROOT / (csv_path.stem + "_report.xlsx")
    with pd.ExcelWriter(out_report) as writer:
        df_rules.to_excel(writer, sheet_name='rule_counts', index=False)
        pd.DataFrame(errors).to_excel(writer, sheet_name='errors', index=False)
        if not df_corrections.empty:
            df_corrections.to_excel(writer, sheet_name='corrections', index=False)
    print(f"Wrote report workbook: {out_report.name}")


if __name__ == '__main__':
    run_pipeline()
