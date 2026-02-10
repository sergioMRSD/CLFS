import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import CLFS_validation_rules as rules
import SSOC_assigner_V3 as ssoc


RELIGION_RECLASS_MAP = {
    "mahayana": "Buddhism",
    "theravada": "Buddhism",
    "vajrayana": "Buddhism",
    "nichiren": "Buddhism",
    "soka gakkai": "Buddhism",
    "catholicism": "Christianity",
    "methodism": "Christianity",
    "evangelicalism": "Christianity",
    "anglicanism": "Christianity",
    "presbyterianism": "Christianity",
    "pentecostalism": "Christianity",
    "lutheranism": "Christianity",
    "sunni islam": "Islam",
    "shia islam": "Islam",
    "ahmadiyya islam": "Islam",
    "shaivism": "Hinduism",
    "vaishnavism": "Hinduism",
    "shaktism": "Hinduism",
    "atheist": "No religion",
    "agnostic": "No religion",
    "freethinker": "No religion",
}

COUNTRY_LIST = {
    "afghanistan",
    "albania",
    "algeria",
    "andorra",
    "angola",
    "antigua & barbuda",
    "argentina",
    "armenia",
    "australia",
    "austria",
    "azerbaijan",
    "bahamas",
    "bahrain",
    "bangladesh",
    "barbados",
    "belarus",
    "belgium",
    "belize",
    "benin",
    "bhutan",
    "bolivia",
    "bosnia & herzegovina",
    "botswana",
    "brazil",
    "brunei",
    "bulgaria",
    "burkina faso",
    "burundi",
    "cabo verde",
    "cambodia",
    "cameroon",
    "canada",
    "central african republic",
    "chad",
    "chile",
    "china",
    "colombia",
    "comoros",
    "congo",
    "costa rica",
    "côte d'ivoire",
    "croatia",
    "cuba",
    "cyprus",
    "czech republic",
    "denmark",
    "djibouti",
    "dominica",
    "dominican republic",
    "dr congo",
    "ecuador",
    "egypt",
    "el salvador",
    "equatorial guinea",
    "eritrea",
    "estonia",
    "eswatini",
    "ethiopia",
    "fiji",
    "finland",
    "france",
    "gabon",
    "gambia",
    "georgia",
    "germany",
    "ghana",
    "greece",
    "grenada",
    "guatemala",
    "guinea",
    "guinea-bissau",
    "guyana",
    "haiti",
    "holy see",
    "honduras",
    "hungary",
    "iceland",
    "india",
    "indonesia",
    "iran",
    "iraq",
    "ireland",
    "israel",
    "italy",
    "jamaica",
    "japan",
    "jordan",
    "kazakhstan",
    "kenya",
    "kiribati",
    "kuwait",
    "kyrgyzstan",
    "laos",
    "latvia",
    "lebanon",
    "lesotho",
    "liberia",
    "libya",
    "liechtenstein",
    "lithuania",
    "luxembourg",
    "madagascar",
    "malawi",
    "malaysia",
    "maldives",
    "mali",
    "malta",
    "marshall islands",
    "mauritania",
    "mauritius",
    "mexico",
    "micronesia",
    "moldova",
    "monaco",
    "mongolia",
    "montenegro",
    "morocco",
    "mozambique",
    "myanmar",
    "namibia",
    "nauru",
    "nepal",
    "netherlands",
    "new zealand",
    "nicaragua",
    "niger",
    "nigeria",
    "north korea",
    "north macedonia",
    "norway",
    "oman",
    "pakistan",
    "palau",
    "panama",
    "papua new guinea",
    "paraguay",
    "peru",
    "philippines",
    "poland",
    "portugal",
    "qatar",
    "romania",
    "russia",
    "rwanda",
    "saint kitts & nevis",
    "saint lucia",
    "samoa",
    "san marino",
    "sao tome & principe",
    "saudi arabia",
    "senegal",
    "serbia",
    "seychelles",
    "sierra leone",
    "singapore",
    "slovakia",
    "slovenia",
    "solomon islands",
    "somalia",
    "south africa",
    "south korea",
    "south sudan",
    "spain",
    "sri lanka",
    "st. vincent & grenadines",
    "state of palestine",
    "sudan",
    "suriname",
    "sweden",
    "switzerland",
    "syria",
    "tajikistan",
    "tanzania",
    "thailand",
    "timor-leste",
    "togo",
    "tonga",
    "trinidad & tobago",
    "tunisia",
    "turkey",
    "turkmenistan",
    "tuvalu",
    "uganda",
    "ukraine",
    "united arab emirates",
    "united kingdom",
    "united states",
    "uruguay",
    "uzbekistan",
    "vanuatu",
    "venezuela",
    "vietnam",
    "yemen",
    "zambia",
    "zimbabwe",
}

# TODO: populate with actual industry strata lookup list (lowercase establishment name -> SSIC code)
STRATA_LOOKUP: list[tuple[str, str]] = []

NO_FREELANCE_TEXT = (
    "I did not take up freelance or assignment-based work through online platforms in the last 12 months"
)

# --- GMI/HQA reference data (for SSOC gating) ---
GMI_GROUP_BANDS = [
    {"group": 1, "hqa": "Degree,Master", "gmi_range": ">$9000"},
    {"group": 2, "hqa": "Diploma and above", "gmi_range": "$7000 - $8999"},
    {"group": 3, "hqa": "Diploma and above", "gmi_range": "$3000 - $6999"},
    {"group": 4, "hqa": "Secondary/ITE", "gmi_range": "$2000 - $2999"},
    {"group": 5, "hqa": "Secondary and below", "gmi_range": "Below $2000"},
]

GMI_HQA_ENTRIES = [
    {"hqa": "Degree/Masters", "ssoc": "11201", "title": "Managing director/Chief executive officer", "gmi_range": ">$7000"},
    {"hqa": "Diploma and above", "ssoc": "33491", "title": "Managing director/Chief executive officer", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33491", "title": "Management executive", "gmi_range": "$3000 - $6999"},
    {"hqa": "Degree/Masters", "ssoc": "11203", "title": "Chief operating officer/General manager", "gmi_range": ">$7000"},
    {"hqa": "Diploma and above", "ssoc": "33491", "title": "Chief operating officer/General manager", "gmi_range": "$3000 - $6999"},
    {"hqa": "Degree/Masters", "ssoc": "12112", "title": "Administration manager", "gmi_range": ">$7000"},
    {"hqa": "Diploma and above", "ssoc": "24299", "title": "Administration manager / Other administration professional n.e.c.", "gmi_range": "$6000 - 6999"},
    {"hqa": "Diploma and above", "ssoc": "33492", "title": "Administration manager", "gmi_range": "$2800 - $5999"},
    {"hqa": "Diploma and above", "ssoc": "33492", "title": "Operations officer (administrative)", "gmi_range": "$2800 - $5999"},
    {"hqa": "Secondary/ITE", "ssoc": "41101", "title": "Administration manager", "gmi_range": "Below $2799"},
    {"hqa": "Secondary/ITE", "ssoc": "41101", "title": "Office Clerk", "gmi_range": "Below $2799"},
    {"hqa": "Degree/Masters", "ssoc": "12212", "title": "Business development manager", "gmi_range": ">$9000"},
    {"hqa": "Diploma and above", "ssoc": "24212", "title": "Business development executive", "gmi_range": "$7000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "24212", "title": "Business consultant", "gmi_range": "$7000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "33221", "title": "Business development manager", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33221", "title": "Business development executive", "gmi_range": "$3000 - $6999"},
    {"hqa": "Degree/Masters", "ssoc": "12222", "title": "Marketing manager", "gmi_range": ">$9000"},
    {"hqa": "Diploma and above", "ssoc": "24314", "title": "Marketing manager", "gmi_range": "$7000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "24314", "title": "Digital marketing professional", "gmi_range": "$7000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "33224", "title": "Marketing manager", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33224", "title": "Online sales channel executive", "gmi_range": "$3000 - $6999"},
    {"hqa": "Degree/Masters", "ssoc": "12121", "title": "Personnel/Human resource manager", "gmi_range": ">$9000"},
    {"hqa": "Diploma and above", "ssoc": "24233", "title": "Personnel/Human resource manager", "gmi_range": "$3000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "24233", "title": "Personnel/Human resource officer", "gmi_range": "$3000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "41102", "title": "Personnel/Human resource clerk", "gmi_range": "Below $3000"},
    {"hqa": "Secondary/ITE", "ssoc": "41102", "title": "Personnel/Human resource clerk", "gmi_range": "Below $3000"},
    {"hqa": "Degree/Masters", "ssoc": "12211", "title": "Sales manager", "gmi_range": ">$7000"},
    {"hqa": "Diploma and below", "ssoc": "52201", "title": "Sales manager", "gmi_range": "Below $3000"},
    {"hqa": "Diploma and below", "ssoc": "52201", "title": "Sales supervisor", "gmi_range": "Below $3000"},
    {"hqa": "Degree/Masters", "ssoc": "13461", "title": "Financial services manager", "gmi_range": ">$9000"},
    {"hqa": "Degree/Masters", "ssoc": "24131", "title": "Financial services manager", "gmi_range": "$7000 - $8999"},
    {"hqa": "Degree/Masters", "ssoc": "24131", "title": "Financial analyst", "gmi_range": "$7000 - $8999"},
    {"hqa": "Diploma and above", "ssoc": "33160", "title": "Financial services manager", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33160", "title": "Financial services back office administrator", "gmi_range": "$3000 - $6999"},
    {"hqa": "Secondary/ITE", "ssoc": "43111", "title": "Financial services manager", "gmi_range": "Below $2999"},
    {"hqa": "Secondary/ITE", "ssoc": "43111", "title": "Bookkeeper", "gmi_range": "Below $2999"},
    {"hqa": "Degree/Masters", "ssoc": "24111", "title": "Accountant (excluding tax accountant)", "gmi_range": ">$7000"},
    {"hqa": "Diploma and above", "ssoc": "33131", "title": "Accountant (excluding tax accountant)", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33131", "title": "Assistant accountant", "gmi_range": "$3000 - $6999"},
    {"hqa": "Secondary/ITE", "ssoc": "43112", "title": "Accountant (excluding tax accountant)", "gmi_range": "Below $2999"},
    {"hqa": "Secondary/ITE", "ssoc": "43112", "title": "Ledger/Accounts clerk", "gmi_range": "Below $2999"},
    {"hqa": "Diploma and above", "ssoc": "22200", "title": "Registered nurse and related nursing professional (excluding enrolled nurse)", "gmi_range": "Any"},
    {"hqa": "Secondary/ITE", "ssoc": "32200", "title": "Enrolled/Assistant nurse (excluding registered nurse)", "gmi_range": "Any"},
    {"hqa": "Degree/Masters", "ssoc": "21511", "title": "Electrical engineer", "gmi_range": ">$7000"},
    {"hqa": "Diploma and above", "ssoc": "31002", "title": "Electrical engineer", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "31002", "title": "Assistant electrical engineer", "gmi_range": "$3000 - $6999"},
    {"hqa": "ITE and below", "ssoc": "74110", "title": "Assistant electrical engineer", "gmi_range": "Below $2999"},
    {"hqa": "ITE and below", "ssoc": "74110", "title": "Electrical engineer", "gmi_range": "Below $2999"},
    {"hqa": "ITE and below", "ssoc": "74110", "title": "Electrician", "gmi_range": "Below $2999"},
    {"hqa": "Degree/Masters", "ssoc": "13241", "title": "Supply and distribution/Logistics/Warehousing manager", "gmi_range": ">$7000"},
    {"hqa": "Diploma and above", "ssoc": "33461", "title": "Supply and distribution/Logistics/Warehousing manager", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33461", "title": "Logistics/Production planner", "gmi_range": "$3000 - $6999"},
    {"hqa": "Diploma and above", "ssoc": "33461", "title": "Production planning clerk", "gmi_range": "$3000 - $6999"},
    {"hqa": "ITE and below", "ssoc": "43222", "title": "Logistics/Production planner", "gmi_range": "$2000 - $2999"},
    {"hqa": "ITE and below", "ssoc": "43222", "title": "Production planning clerk", "gmi_range": "$2000 - $2999"},
    {"hqa": "Secondary and below", "ssoc": "93334", "title": "Supply and distribution/Logistics/Warehousing manager", "gmi_range": "Below $1999"},
    {"hqa": "Secondary and below", "ssoc": "93334", "title": "Logistics/Production planner", "gmi_range": "Below $1999"},
    {"hqa": "Secondary and below", "ssoc": "93334", "title": "Warehouse worker", "gmi_range": "Below $1999"},
    {"hqa": "Degree/Masters", "ssoc": "33551", "title": "Police inspector", "gmi_range": "Above $3000"},
    {"hqa": "Degree/Masters", "ssoc": "33551", "title": "Police officer", "gmi_range": "Above $3000"},
    {"hqa": "Secondary/ITE/Diploma", "ssoc": "54121", "title": "Police inspector", "gmi_range": "Above $3000"},
    {"hqa": "Secondary/ITE/Diploma", "ssoc": "54121", "title": "Police officer", "gmi_range": "Above $3000"},
    {"hqa": "Diploma and above", "ssoc": "32530", "title": "Community health worker", "gmi_range": "Above $2500"},
    {"hqa": "Diploma and above", "ssoc": "32530", "title": "Healthcare assistant", "gmi_range": "Above $2500"},
    {"hqa": "ITE and below", "ssoc": "53201", "title": "Community health worker", "gmi_range": "Below $2499"},
    {"hqa": "ITE and below", "ssoc": "53201", "title": "Healthcare assistant", "gmi_range": "Below $2499"},
    {"hqa": "Diploma and above", "ssoc": "36991", "title": "Relief teacher", "gmi_range": "Above $2500"},
    {"hqa": "Secondary/ITE", "ssoc": "53120", "title": "Relief teacher", "gmi_range": "Below $2499"},
    {"hqa": "Secondary/ITE", "ssoc": "53120", "title": "Teacher aide", "gmi_range": "Below $2499"},
    {"hqa": "Diploma and above", "ssoc": "36100", "title": "Preschool education teacher", "gmi_range": "Above $2500"},
    {"hqa": "Secondary/ITE", "ssoc": "53113", "title": "Preschool education teacher", "gmi_range": "Below $2499"},
    {"hqa": "Secondary/ITE", "ssoc": "53113", "title": "Child/After school care centre worker", "gmi_range": "Below $2499"},
]


# Column name to HouseholdMember attribute mapping
COLUMN_MAPPING = {
    "full_name": "Full Name",
    "date_of_birth": "Date of Birth (DD/MM/YYYY)",
    "age": "Age",
    "contact_number": "Contact Number",
    "tenancy_of_household": "Tenancy of Household",
    "hire_foreign_domestic_workers": "Do you hire any Foreign Domestic Workers in this household?",
    "num_foreign_domestic_workers": "How many Foreign Domestic Workers do you have?",
    "foreign_domestic_workers_received_bonus": "Did your Foreign Domestic Worker(s) receive any bonus during the last 12 months (including the 13th month Annual Wage Supplement)?",
    "sex": "Sex",
    "place_of_birth": "Place of Birth",
    "identification_type": "Identification Type",
    "race": "Race",
    "where_currently_staying": "Where are you currently staying?",
    "main_reason_living_abroad": "What is your main reason for living abroad?",
    "religion": "What is your religion?",
    "marital_status": "Marital Status",
    "number_of_children": "Number of children given birth to",
    "highest_academic_qualification": "Highest Academic Qualification",
    "field_of_study_highest_academic": "Field of study of your highest academic qualification attained?",
    "place_of_study_highest_academic": "Place of study for your Highest Academic Attained in?",
    "has_vocational_skills_certificates": "Have you ever obtained any Vocational or Skills certificates/qualifications, (e.g. (WSQ) and (ESS) certificates, or formal certifications that validate knowledge and skills in a particular field)?",
    "highest_vocational_certificate": "What is the highest vocational or skills certificate/qualification obtained?",
    "field_of_study_vocational": "What is the field of study of your highest vocational or skills certificate/qualification?",
    "place_of_study_vocational": "Where is the place of study for your highest vocational or skills certificate/qualification?",
    "care_economy": "Care economy",
    "artificial_intelligence": "Artificial Intelligence",
    "digital_skills": "Digital skills (excluding Artificial Intelligence)",
    "green_economy": "Green economy",
    "industry_4_0": "Industry 4.0",
    "programmes_used_to_upgrade_skills": "Have you utilised any of the following programmes/initiatives to upgrade your skills or switch jobs?",
    "ever_retired": "Have you ever retired from any job?",
    "retirement_age": "What age retire?",
    "labour_force_status": "Labour Force Status",
    "employment_status_last_week": "Employment Status as of last week",
    "organisation_type_last_week": "The organisation that employed you last week was a/an:",
    "paid_internship_traineeship": "Was your main job last week a paid internship, traineeship or apprenticeship?",
    "reason_for_internship": "What was the main reason you were in a paid internship, traineeship, or apprenticeship?",
    "salary_paid_by_contracting_agency": "Is your salary paid by an employment/labour contracting agency (e.g., BGC Group, PERSOLKELLY)?",
    "deployed_to_another_organisation": "Are you deployed to work in another organisation that supervises your work?",
    "business_trade_type": "The business or trade you operated/helping in last week was a/an:",
    "acra_registered_business_owner": "Are you an owner/partner of an ACRA-registered business in this job?",
    "business_incorporated": "Is this business incorporated (e.g., name of business ends with private limited or its equivalent)?",
    "sets_price_for_goods_services": "Do you usually set the price for the goods or services you provide in this job?",
    "job_title": "Job Title",
    "main_tasks_duties": "Main tasks / duties",
    "skills_description": "Which of the following statements best describe your skills in relation to what is needed for your job/business?",
    "qualification_needed_for_job": "In your view, what level of qualification, if any, is needed to carry out the tasks and duties of your job/business?",
    "field_of_study_needed_for_job": "In your view, which field of study is needed to carry out the tasks and duties of your job/business?",
    "name_of_establishment_last_week": "Name of Establishment you were working last week?",
    "reasons_self_employed": "What were your reason(s) for being self-employed?",
    "prefer_to_be": "Would you prefer to be a/an",
    "reasons_for_taking_job": "What were your reason(s) for taking up this job?",
    "freelance_platforms": "Did you perform any freelance or assignment-based work via any of the following online platform(s) in the last 12 months?",
    "job_accommodations": "Does your current job accommodate the working arrangements you need (e.g. shorter working hours, provision of flexible work arrangements)?",
    "keen_reasons": "I was keen in this job and took it up because:",
    "not_keen_reasons": "I was not keen in this job, but still took it up because:",
    "usual_hours_of_work": "Usual hours of work",
    "reason_working_part_time": "Reason that working part time rather than full time?",
    "person_taking_care_of": "The person you are mainly taking care of is your",
    "care_recipient_age": "What is his/her age?",
    "care_recipient_disabled_ill": "Is he/she disabled/ill?",
    "main_reason_part_time": "Main reason for working part-time rather than full-time:",
    "work_full_time_if_care_services_available": "Would you work full-time if suitable care services were available and/or affordable?",
    "willing_work_additional_hours": "Willing to work additional hours?",
    "available_additional_work": "Available for additional work?",
    "actively_sought_additional_hours": "Actively sought additional hours of work in the past four weeks?",
    "gmi": "GMI",
    "bonus_received_last_12_months": "Bonus received from your job(s) during the last 12 months",
    "employed_at_least_10_months": "Employed for at least 10 months during the last 12 months?",
    "num_jobs_held_last_week": "How many jobs did you hold last week?",
    "when_began_current_employer": "When did you begin working for your current employer?",
    "type_of_employment": "Type of Employment?",
    "contract_duration": "Contract duration",
    "began_as_fixed_term": "Begin as a fixed-term contract employee in your current job?",
    "employer_gave_paid_leave": "Did your employer give you paid leave or compensation instead?",
    "employer_gave_paid_sick_leave": "Did your employer give you paid sick leave when you were on MC",
    "employer_gave_rest_day_weekly": "Did your employer give you at least one rest day each week?",
    "satisfied_current_job": "Are you satisfied with your current job?",
    "actively_looking_new_job": "Are you actively looking for a new job?",
    "how_looked_for_job_last_4_weeks": "How did you look for a job or employment during the last 4 weeks?",
    "looking_for_permanent_job": "Are you looking for a permanent job?",
    "available_start_new_job": "Are you available to start work on the new job upon quitting the current job?",
    "looking_to_better_utilise_skills": "Is the main reason for looking for a new job to better utilise your skills?",
    "num_job_changes_last_2_years": "Number of Job changes in the last 2 years",
    "when_left_last_job": "When did you leave your last job?",
    "age_started_employment": "At what age did you start employment",
    "establishment_name_last_worked": "Name of Establishment you were working last worked",
    "interest_from_savings_last_12_months": "How much interest did you receive from savings (e.g., current and saving accounts, fixed deposits) in the last 12 months?",
    "dividends_interests_investments_last_12_months": "How much dividends and interests did you receive from other investment sources (e.g., bonds, shares, unit trust, personal loans to persons outside your households) in the last 12 months?",
    "freelance_online_platforms_last_12_months": "Did you perform any freelance or assignment-based work via any of the following online platform(s) in the last 12 months?",
    "self_employed_last_12_months": "At any point in the last 12 months, were you self-employed?",
    "worked_own_business_last_12_months": "At any point in the last 12 months, did you work on your own (i.e., without paid employees) while running your own business or trade?",
    "ns_industry": "NS Industry",
    "remarks": "Remarks",
}


@dataclass
class HouseholdMember:
    # Basic Information
    full_name: str
    date_of_birth: Optional[str] = None
    age: Optional[int] = None
    contact_number: Optional[str] = None
    tenancy_of_household: Optional[str] = None
    hire_foreign_domestic_workers: Optional[str] = None
    num_foreign_domestic_workers: Optional[int] = None
    foreign_domestic_workers_received_bonus: Optional[str] = None
    sex: Optional[str] = None
    place_of_birth: Optional[str] = None
    identification_type: Optional[str] = None
    race: Optional[str] = None
    where_currently_staying: Optional[str] = None
    main_reason_living_abroad: Optional[str] = None
    religion: Optional[str] = None
    marital_status: Optional[str] = None
    number_of_children: Optional[int] = None
    
    # Education
    highest_academic_qualification: Optional[str] = None
    field_of_study_highest_academic: Optional[str] = None
    place_of_study_highest_academic: Optional[str] = None
    has_vocational_skills_certificates: Optional[str] = None
    highest_vocational_certificate: Optional[str] = None
    field_of_study_vocational: Optional[str] = None
    place_of_study_vocational: Optional[str] = None
    
    # Skills & Training
    care_economy: Optional[str] = None
    artificial_intelligence: Optional[str] = None
    digital_skills: Optional[str] = None
    green_economy: Optional[str] = None
    industry_4_0: Optional[str] = None
    programmes_used_to_upgrade_skills: Optional[str] = None
    
    # Retirement
    ever_retired: Optional[str] = None
    retirement_age: Optional[int] = None
    
    # Employment Status
    labour_force_status: Optional[str] = None
    employment_status_last_week: Optional[str] = None
    organisation_type_last_week: Optional[str] = None
    paid_internship_traineeship: Optional[str] = None
    reason_for_internship: Optional[str] = None
    salary_paid_by_contracting_agency: Optional[str] = None
    deployed_to_another_organisation: Optional[str] = None
    business_trade_type: Optional[str] = None
    acra_registered_business_owner: Optional[str] = None
    business_incorporated: Optional[str] = None
    sets_price_for_goods_services: Optional[str] = None
    
    # Current Job Details
    job_title: Optional[str] = None
    main_tasks_duties: Optional[str] = None
    skills_description: Optional[str] = None
    
    qualification_needed_for_job: Optional[str] = None
    field_of_study_needed_for_job: Optional[str] = None
    name_of_establishment_last_week: Optional[str] = None
    reasons_self_employed: Optional[str] = None
    prefer_to_be: Optional[str] = None
    reasons_for_taking_job: Optional[str] = None
    keen_reasons: Optional[str] = None
    not_keen_reasons: Optional[str] = None
    
    # Work Hours & Arrangements
    usual_hours_of_work: Optional[float] = None
    reason_working_part_time: Optional[str] = None
    person_taking_care_of: Optional[str] = None
    care_recipient_age: Optional[int] = None
    care_recipient_disabled_ill: Optional[str] = None
    main_reason_part_time: Optional[str] = None
    work_full_time_if_care_services_available: Optional[str] = None
    willing_work_additional_hours: Optional[str] = None
    available_additional_work: Optional[str] = None
    actively_sought_additional_hours: Optional[str] = None
    
    # Compensation & Benefits
    gmi: Optional[float] = None
    bonus_received_last_12_months: Optional[float] = None
    employed_at_least_10_months: Optional[str] = None
    num_jobs_held_last_week: Optional[int] = None
    when_began_current_employer: Optional[str] = None
    type_of_employment: Optional[str] = None
    contract_duration: Optional[str] = None
    began_as_fixed_term: Optional[str] = None
    employer_gave_paid_leave: Optional[str] = None
    employer_gave_paid_sick_leave: Optional[str] = None
    employer_gave_rest_day_weekly: Optional[str] = None
    
    # Job Satisfaction & Search
    satisfied_current_job: Optional[str] = None
    actively_looking_new_job: Optional[str] = None
    how_looked_for_job_last_4_weeks: Optional[str] = None
    looking_for_permanent_job: Optional[str] = None
    available_start_new_job: Optional[str] = None
    looking_to_better_utilise_skills: Optional[str] = None
    num_job_changes_last_2_years: Optional[int] = None
    
    # Previous Employment
    when_left_last_job: Optional[str] = None
    usual_hours_work_previous: Optional[float] = None
    employment_status_previous: Optional[str] = None
    type_employment_previous: Optional[str] = None
    contract_duration_previous: Optional[str] = None
    job_title_previous: Optional[str] = None
    main_tasks_duties_previous: Optional[str] = None
    establishment_name_previous: Optional[str] = None
    age_started_employment: Optional[int] = None
    breaks_in_employment: Optional[int] = None
    
    # Work Relocation
    relocated_from_singapore: Optional[str] = None
    first_relocation_experience: Optional[str] = None
    relocation_total_duration: Optional[str] = None
    how_work_stint_arose: Optional[str] = None
    job_title_relocated: Optional[str] = None
    job_industry_sector_relocated: Optional[str] = None
    last_drawn_gmi_relocated: Optional[float] = None
    company_type_relocated: Optional[str] = None
    location_of_work_relocated: Optional[str] = None
    
    # Job Search Status
    actively_looking_jobs_past_4_weeks: Optional[str] = None
    looked_for_job_last_12_months: Optional[str] = None
    want_to_work_at_present: Optional[str] = None
    already_secured_job: Optional[str] = None
    how_soon_expect_start_new_job: Optional[str] = None
    available_work_next_2_weeks: Optional[str] = None
    when_available_to_work: Optional[str] = None
    how_long_looking_for_job_weeks: Optional[int] = None
    what_doing_while_looking: Optional[str] = None
    occupation_looking_for: Optional[str] = None
    main_step_to_look_employment: Optional[str] = None
    other_steps_look_employment: Optional[str] = None
    experienced_difficulties_securing_job: Optional[str] = None
    main_difficulty_encountered: Optional[str] = None
    other_difficulties_encountered: Optional[str] = None
    
    # Work History
    ever_worked_before: Optional[str] = None
    employment_status_last_worked: Optional[str] = None
    job_title_last_worked: Optional[str] = None
    main_tasks_duties_last_worked: Optional[str] = None
    establishment_name_last_worked: Optional[str] = None
    usual_hours_work_last_worked: Optional[float] = None
    last_drawn_gmi_last_worked: Optional[float] = None
    main_reason_left_last_job: Optional[str] = None
    reason_left_elaboration: Optional[str] = None
    reason_left_temporary_nature: Optional[str] = None
    reason_left_illness_injury: Optional[str] = None
    
    # Care Responsibilities (leaving job)
    person_taking_care_of_leaving: Optional[str] = None
    care_recipient_age_leaving: Optional[int] = None
    care_recipient_disabled_leaving: Optional[str] = None
    main_reason_leaving_due_care: Optional[str] = None
    work_full_time_if_care_services_leaving: Optional[str] = None
    when_left_last_job_months: Optional[int] = None
    
    # Second Relocation Info
    relocated_from_singapore_2: Optional[str] = None
    first_relocation_experience_2: Optional[str] = None
    relocation_total_duration_2: Optional[str] = None
    how_work_stint_arose_2: Optional[str] = None
    job_title_relocated_2: Optional[str] = None
    job_industry_sector_relocated_2: Optional[str] = None
    last_drawn_gmi_relocated_2: Optional[float] = None
    company_type_relocated_2: Optional[str] = None
    location_work_relocated_2: Optional[str] = None
    
    # Not Working/Not Looking
    main_reason_not_working_not_looking: Optional[str] = None
    ever_retired_2: Optional[str] = None
    retirement_age_2: Optional[int] = None
    person_taking_care_of_2: Optional[str] = None
    care_recipient_age_2: Optional[int] = None
    care_recipient_disabled_2: Optional[str] = None
    main_reason_not_working_not_looking_detail: Optional[str] = None
    work_if_care_services_available: Optional[str] = None
    ever_worked_before_2: Optional[str] = None
    when_left_last_job_months_2: Optional[int] = None
    employment_status_last_worked_2: Optional[str] = None
    job_title_last_worked_2: Optional[str] = None
    main_tasks_duties_last_worked_2: Optional[str] = None
    establishment_name_last_worked_2: Optional[str] = None
    usual_hours_work_last_worked_2: Optional[float] = None
    
    # Future Work Plans
    intend_look_job_future: Optional[str] = None
    when_intend_look_job: Optional[str] = None
    prefer_full_time_or_part_time: Optional[str] = None
    
    # Self-Employment & Gig Work
    self_employed_last_12_months: Optional[str] = None
    self_employed_last_12_months_2: Optional[str] = None
    worked_own_business_last_12_months: Optional[str] = None
    freelance_online_platforms_last_12_months: Optional[str] = None
    held_licences_permits_last_12_months: Optional[str] = None
    did_work_related_to_licences: Optional[str] = None
    reason_holding_licence_not_working: Optional[str] = None
    
    # Income from Non-Employment Sources
    interest_from_savings_last_12_months: Optional[float] = None
    revise_interest_earned_answer: Optional[str] = None
    interest_from_savings_revised: Optional[float] = None
    dividends_interests_investments_last_12_months: Optional[float] = None
    other_income_non_employment: Optional[str] = None
    income_from_rents_last_12_months: Optional[float] = None
    allowances_contributions_last_12_months: Optional[float] = None
    other_sources_income_last_12_months: Optional[float] = None
    
    # Care Provision
    provide_care_to_individuals: Optional[str] = None
    provide_care_to_individuals_2: Optional[str] = None
    individuals_have_long_term_care_needs: Optional[str] = None
    individuals_with_long_term_care_relationship: Optional[str] = None
    how_long_providing_caregiving_support: Optional[str] = None
    expect_provide_support_6_months: Optional[str] = None
    
    # Disabilities & Difficulties
    difficulty_seeing: Optional[str] = None
    difficulty_hearing: Optional[str] = None
    difficulty_body_movement: Optional[str] = None
    difficulty_self_care: Optional[str] = None
    long_lasting_difficulties: Optional[str] = None
    
    # Work Accommodations
    job_accommodates_working_arrangements: Optional[str] = None
    job_accommodates_working_arrangements_2: Optional[str] = None
    
    # Additional Info
    ns_industry: Optional[str] = None
    remarks: Optional[str] = None


def _normalize_value(value: object) -> Optional[str]:
    if pd.isna(value):
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_header(text: object) -> str:
    if text is None:
        return ""
    return str(text).strip().lower()


def _column_matches(col_name: object, target: str) -> bool:
    col_norm = _normalize_header(col_name)
    target_norm = _normalize_header(target)
    return target_norm == col_norm or target_norm in col_norm


def _find_column_name(columns: list, target: str) -> Optional[str]:
    if not columns:
        return None
    target_norm = _normalize_header(target)

    exact_matches = [c for c in columns if _normalize_header(c) == target_norm]
    if exact_matches:
        return exact_matches[0]

    partial_matches = [c for c in columns if target_norm and target_norm in _normalize_header(c)]
    if not partial_matches:
        return None

    return sorted(partial_matches, key=lambda c: len(_normalize_header(c)))[0]


def _find_column_indices(columns: list, target: str) -> list[int]:
    return [i for i, col in enumerate(columns) if _column_matches(col, target)]


def _get_cell_value(df: pd.DataFrame, row_idx: int, target: str) -> Optional[object]:
    col_name = _find_column_name(list(df.columns), target)
    if not col_name:
        return None
    return df.at[row_idx, col_name]


def _get_column_index(df: pd.DataFrame, target: str) -> tuple[Optional[str], Optional[int]]:
    col_name = _find_column_name(list(df.columns), target)
    if not col_name:
        return None, None
    return col_name, df.columns.get_loc(col_name)


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _ensure_ssic_column(df: pd.DataFrame) -> tuple[pd.DataFrame, Optional[str]]:
    est_col = _find_column_name(list(df.columns), "Name of Establishment you were working last week?")
    if not est_col:
        return df, None

    ssic_col = _find_column_name(list(df.columns), "SSIC Code")
    if ssic_col:
        return df, ssic_col

    cols = list(df.columns)
    insert_at = cols.index(est_col) + 1
    cols.insert(insert_at, "SSIC Code")
    df = df.reindex(columns=cols)
    return df, "SSIC Code"


def _gmi_range_to_group(gmi_range: str) -> Optional[int]:
    if not gmi_range:
        return None
    text = _normalize_text(gmi_range)
    if "any" in text:
        return None
    if ">" in text and "9000" in text:
        return 1
    if "7000" in text and "8999" in text:
        return 2
    if "3000" in text and ("6999" in text or "5999" in text):
        return 3
    if "2000" in text and "2999" in text:
        return 4
    if "below" in text:
        return 5
    if "above" in text and ("3000" in text or "2500" in text):
        return 3
    return None


def _build_ssoc_min_group_map(entries: list[dict]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for entry in entries:
        ssoc_code = str(entry.get("ssoc", "")).strip()
        gmi_range = str(entry.get("gmi_range", "")).strip()
        if not ssoc_code:
            continue
        group = _gmi_range_to_group(gmi_range)
        if group is None:
            continue
        current = mapping.get(ssoc_code)
        mapping[ssoc_code] = group if current is None else min(current, group)
    return mapping


SSOC_MIN_GROUP_BY_CODE = _build_ssoc_min_group_map(GMI_HQA_ENTRIES)


def _is_degree_hqa(hqa_value: Optional[str]) -> bool:
    text = _normalize_text(hqa_value)
    return "degree" in text or "master" in text


def _parse_gmi_value(gmi_value: Optional[object]) -> Optional[float]:
    if gmi_value is None or gmi_value == "":
        return None
    try:
        return float(gmi_value)
    except (ValueError, TypeError):
        return None


def _gmi_in_range(gmi_value: float, gmi_range: str) -> bool:
    if gmi_value is None or gmi_range is None:
        return False
    text = _normalize_text(gmi_range)
    if "any" in text:
        return True

    numbers = [int(n) for n in re.findall(r"\d+", text)]
    if not numbers:
        return False

    if "below" in text:
        return gmi_value <= numbers[0]
    if "above" in text or text.startswith(">"):
        return gmi_value >= numbers[0]
    if len(numbers) >= 2:
        return numbers[0] <= gmi_value <= numbers[1]
    return False


def _hqa_matches_entry(hqa_value: Optional[str], entry_hqa: str) -> bool:
    hqa = _normalize_text(hqa_value)
    entry = _normalize_text(entry_hqa)
    if not hqa or not entry:
        return False

    if "degree" in entry or "master" in entry:
        return "degree" in hqa or "master" in hqa
    if "diploma and above" in entry:
        return "diploma" in hqa or "degree" in hqa or "master" in hqa
    if "secondary/ite/diploma" in entry:
        return "secondary" in hqa or "ite" in hqa or "diploma" in hqa
    if "secondary/ite" in entry:
        return "secondary" in hqa or "ite" in hqa
    if "diploma and below" in entry:
        return "degree" not in hqa and "master" not in hqa
    if "ite and below" in entry:
        return "ite" in hqa or "secondary" in hqa or "below" in hqa
    if "secondary and below" in entry:
        return "degree" not in hqa and "master" not in hqa and "diploma" not in hqa

    return entry in hqa


def _select_candidate_by_examples(top_5: list[dict], hqa_value: Optional[str], gmi_value: Optional[float]) -> Optional[str]:
    if not top_5 or gmi_value is None:
        return None

    for candidate in top_5:
        code = str(candidate.get("code", "")).strip()
        if not code:
            continue
        for entry in GMI_HQA_ENTRIES:
            if code != str(entry.get("ssoc", "")).strip():
                continue
            if not _hqa_matches_entry(hqa_value, str(entry.get("hqa", ""))):
                continue
            if _gmi_in_range(gmi_value, str(entry.get("gmi_range", ""))):
                return code
    return None


def _hqa_matches_band(hqa_value: Optional[str], band_hqa: str) -> bool:
    return _hqa_matches_entry(hqa_value, band_hqa)


def _required_group_from_band(hqa_value: Optional[str], gmi_value: Optional[float]) -> Optional[int]:
    if gmi_value is None:
        return None
    for band in GMI_GROUP_BANDS:
        if not _hqa_matches_band(hqa_value, band.get("hqa", "")):
            continue
        if _gmi_in_range(gmi_value, band.get("gmi_range", "")):
            return band.get("group")
    return None


def _select_candidate_by_band(top_5: list[dict], required_group: Optional[int]) -> Optional[str]:
    if required_group is None or not top_5:
        return None
    for candidate in top_5:
        code = str(candidate.get("code", "")).strip()
        if not code:
            continue
        min_group = SSOC_MIN_GROUP_BY_CODE.get(code)
        if min_group is None or min_group >= required_group:
            return code
    return None


def _get_member_column_groups(columns: list[str]) -> list[dict[str, Optional[int]]]:
    full_name_indices = _find_column_indices(columns, "Full Name")
    dob_indices = _find_column_indices(columns, "Date of Birth (DD/MM/YYYY)")

    groups: list[dict[str, Optional[int]]] = []
    for idx, full_name_idx in enumerate(full_name_indices):
        next_full_name_idx = (
            full_name_indices[idx + 1] if idx + 1 < len(full_name_indices) else len(columns)
        )
        dob_idx = next(
            (i for i in dob_indices if full_name_idx < i < next_full_name_idx),
            None,
        )
        groups.append({"full_name_idx": full_name_idx, "dob_idx": dob_idx})

    return groups


def extract_household_members(df: pd.DataFrame) -> list[list[HouseholdMember]]:
    columns = list(df.columns)
    groups = _get_member_column_groups(columns)
    households: list[list[HouseholdMember]] = []

    for _, row in df.iterrows():
        members: list[HouseholdMember] = []
        for group in groups:
            name = _normalize_value(row.iloc[group["full_name_idx"]])
            if not name:
                continue
            
            # Create member with name (required)
            member = HouseholdMember(full_name=name)
            
            # Populate all mapped attributes from the row
            for attr_name, col_name in COLUMN_MAPPING.items():
                if attr_name == "full_name":
                    # Already set
                    continue
                matched_col = _find_column_name(columns, col_name)
                if matched_col:
                    col_idx = columns.index(matched_col)
                    value = _normalize_value(row.iloc[col_idx])
                    # Try to convert to appropriate type
                    if value:
                        if attr_name in ["age", "num_foreign_domestic_workers", "number_of_children", 
                                        "retirement_age", "care_recipient_age", "num_jobs_held_last_week",
                                        "num_job_changes_last_2_years", "care_recipient_age_leaving", 
                                        "when_left_last_job_months", "care_recipient_age_2", 
                                        "when_left_last_job_months_2", "how_long_looking_for_job_weeks",
                                        "age_started_employment", "breaks_in_employment"]:
                            try:
                                value = int(float(value))
                            except (ValueError, TypeError):
                                pass
                        elif attr_name in ["gmi", "bonus_received_last_12_months", "usual_hours_of_work",
                                          "last_drawn_gmi_relocated", "usual_hours_work_previous",
                                          "usual_hours_work_last_worked", "usual_hours_work_last_worked_2",
                                          "gmi", "bonus_received_last_12_months", "usual_hours_of_work",
                                          "interest_from_savings_last_12_months", "interest_from_savings_revised",
                                          "dividends_interests_investments_last_12_months", 
                                          "income_from_rents_last_12_months", "allowances_contributions_last_12_months",
                                          "other_sources_income_last_12_months", "last_drawn_gmi_relocated_2"]:
                            try:
                                value = float(value)
                            except (ValueError, TypeError):
                                pass
                        setattr(member, attr_name, value)
            
            members.append(member)
        households.append(members)

    return households

def load_input_files(folder_path="Operating_Table"):
    """
    Load all .xlsx and .csv files from the specified folder.
    
    Args:
        folder_path (str): Path to the folder containing input files
        
    Returns:
        dict: Dictionary with filenames as keys and DataFrames as values
    """
    input_files = {}
    
    # Check if folder exists
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return input_files
    
    # Find all .xlsx and .csv files in the folder
    for file in Path(folder_path).glob("*.xlsx"):
        try:
            print(f"Loading {file.name}...")
            df = pd.read_excel(file)
            df = _clean_dataframe(df)
            input_files[file.name] = df
            print(f"Successfully loaded {file.name} with {len(df)} rows and {len(df.columns)} columns")
        except Exception as e:
            print(f"Error loading {file.name}: {e}")
    
    for file in Path(folder_path).glob("*.csv"):
        try:
            print(f"Loading {file.name}...")
            header_row_idx = 5  # Row 6 (0-based index)
            df = pd.read_csv(
                file,
                header=0,
                skiprows=range(header_row_idx),
                encoding="utf-8-sig"
            )
            df = _clean_dataframe(df)

            input_files[file.name] = df
            print(f"Successfully loaded {file.name} with {len(df)} rows and {len(df.columns)} columns")
        except Exception as e:
            print(f"Error loading {file.name}: {e}")
    
    if not input_files:
        print(f"No .xlsx or .csv files found in '{folder_path}'")
    
    return input_files


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all")
    df.columns = [str(col).strip() for col in df.columns]
    response_col = _find_column_name(list(df.columns), "Response ID")
    if response_col:
        df = df[df[response_col].notna()]
    return df.reset_index(drop=True)


def _ensure_ssec_column(df: pd.DataFrame) -> pd.DataFrame:
    ssec_col = _find_column_name(list(df.columns), "SSEC Code")
    if ssec_col:
        df[ssec_col] = df[ssec_col].astype("object")
        return df
    hqa_col = _find_column_name(list(df.columns), "Highest Academic Qualification")
    if not hqa_col:
        return df

    cols = list(df.columns)
    insert_at = cols.index(hqa_col) + 1
    cols.insert(insert_at, "SSEC Code")
    df = df.reindex(columns=cols)
    df["SSEC Code"] = df["SSEC Code"].astype("object")
    return df


SSOC_DEFINITIONS_FILE = os.environ.get(
    "SSOC_DEFINITIONS_FILE",
    str(Path("references") / "ssoc2024-detailed-definitions.xlsx")
)
SSOC_EXPERT_MAP_FILE = os.environ.get(
    "SSOC_EXPERT_MAP_FILE",
    str(Path("references") / "Library_of_SSOC_eng_manager.xlsx")
)
SSOC_MIN_SCORE = float(os.environ.get("SSOC_MIN_SCORE", "0.05"))

_SSOC_RESOURCES_CACHE: Optional[dict] = None


def _load_ssoc_resources() -> Optional[dict]:
    global _SSOC_RESOURCES_CACHE
    if _SSOC_RESOURCES_CACHE is not None:
        return _SSOC_RESOURCES_CACHE

    defs_path = SSOC_DEFINITIONS_FILE
    if not defs_path or not os.path.exists(defs_path):
        return None

    defs, title_map = ssoc.load_definitions(
        defs_path,
        ssoc.DEFAULT_DEF_SHEET,
        ssoc.DEFAULT_DEF_SKIP_ROWS,
        debug=False
    )

    expert_map = {}
    expert_path = SSOC_EXPERT_MAP_FILE
    if expert_path and os.path.exists(expert_path):
        expert_map = ssoc.load_expert_map(expert_path, debug=False)

    _SSOC_RESOURCES_CACHE = {
        "defs": defs,
        "title_map": title_map,
        "expert_map": expert_map,
    }
    return _SSOC_RESOURCES_CACHE


def _ensure_ssoc_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    """
    Insert "SSOC Code" columns after each "Main tasks / duties" column (paired with a Job Title).
    Returns updated DataFrame and list of column group metadata.
    """
    columns = list(df.columns)
    duty_indices = _find_column_indices(columns, "Main tasks / duties")
    title_indices = _find_column_indices(columns, "Job Title")

    if not duty_indices:
        return df, []

    groups: list[dict] = []
    for idx, duty_idx in enumerate(duty_indices):
        next_duty_idx = duty_indices[idx + 1] if idx + 1 < len(duty_indices) else len(columns)
        title_idx = None
        if title_indices:
            titles_after = [i for i in title_indices if duty_idx < i < next_duty_idx]
            if titles_after:
                title_idx = titles_after[0]
            else:
                titles_before = [i for i in title_indices if i < duty_idx]
                title_idx = titles_before[-1] if titles_before else None
        groups.append({"title_idx": title_idx, "duties_idx": duty_idx, "ssoc_idx": None})

    offset = 0
    for group in sorted(groups, key=lambda g: g["duties_idx"]):
        duties_idx = group["duties_idx"] + offset
        title_idx = group["title_idx"] + offset if group["title_idx"] is not None else None
        insert_at = duties_idx + 1

        if insert_at < len(columns) and _column_matches(columns[insert_at], "SSOC Code"):
            group["ssoc_idx"] = insert_at
        else:
            df.insert(insert_at, "SSOC Code", "", allow_duplicates=True)
            columns.insert(insert_at, "SSOC Code")
            group["ssoc_idx"] = insert_at
            offset += 1

        group["duties_idx"] = duties_idx
        group["title_idx"] = title_idx

    df = df.copy()
    return df, groups


def _add_ft_pt_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, list[tuple[int, int, str]]]:
    """
    Rule 12: Add FT/PT columns next to each "Usual hours of work" column.
    Returns updated DataFrame and a list of changes (row_idx, col_idx, value).
    """
    changes: list[tuple[int, int, str]] = []
    col_idx = 0

    while col_idx < len(df.columns):
        columns = list(df.columns)
        col_name = columns[col_idx]

        if _column_matches(col_name, "Usual hours of work"):
            series = df.iloc[:, col_idx]

            insert_at = col_idx + 1
            if insert_at < len(columns) and _column_matches(columns[insert_at], "FT/PT"):
                ftpt_col_idx = insert_at
            else:
                df.insert(insert_at, "FT/PT", "", allow_duplicates=True)
                ftpt_col_idx = insert_at

            for row_idx, value in series.items():
                if pd.isna(value) or str(value).strip() == "":
                    if df.iat[row_idx, ftpt_col_idx] not in (None, ""):
                        df.iat[row_idx, ftpt_col_idx] = ""
                        changes.append((row_idx, ftpt_col_idx, ""))
                    continue
                try:
                    hours = float(value)
                except (ValueError, TypeError):
                    if df.iat[row_idx, ftpt_col_idx] not in (None, ""):
                        df.iat[row_idx, ftpt_col_idx] = ""
                        changes.append((row_idx, ftpt_col_idx, ""))
                    continue
                ft_pt = "FT" if hours >= 35 else "PT"
                df.iat[row_idx, ftpt_col_idx] = ft_pt
                changes.append((row_idx, ftpt_col_idx, ft_pt))

            col_idx = ftpt_col_idx + 1
            continue

        col_idx += 1

    df = df.copy()
    return df, changes


def create_output_directory():
    """Create output folder if it doesn't exist"""
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    return output_dir


def create_validation_report(rule_errors: list[dict], source_filename: str) -> Optional[Path]:
    """
    Create a validation report Excel file with summary and details sheets.

    Sheet 1: Summary of errors with frequency counts
    Sheet 2: Detailed errors with Response ID and Full Name

    Args:
        rule_errors: List of error dicts
        source_filename: Input filename

    Returns:
        Path to the report file if created
    """
    if not rule_errors:
        return None

    output_dir = create_output_directory()
    filename = Path(source_filename).stem
    report_path = output_dir / f"{filename}_validation_report.xlsx"

    details_df = pd.DataFrame(rule_errors)

    summary_df = (
        details_df
        .groupby(["rule", "column", "message"], dropna=False)
        .size()
        .reset_index(name="count")
        .sort_values("count", ascending=False)
    )

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        details_df.to_excel(writer, sheet_name="Details", index=False)

    print(f"\n✓ Validation report saved to: {report_path}")
    return report_path


def save_with_highlights(
    df: pd.DataFrame,
    original_file_path: str,
    changes: dict,
    error_cells: set[tuple[int, int]]
):
    """
    Save modified Excel file with cells highlighted in orange for changes
    and yellow for detected errors.
    
    Args:
        df: Modified DataFrame
        original_file_path: Path to original file
        changes: Dictionary with format {(row, col): (old_value, new_value)}
        error_cells: Set of (row, col) positions for error highlights
    """
    output_dir = create_output_directory()
    
    # Create output filename
    original_path = Path(original_file_path)
    filename = original_path.stem
    
    output_path = output_dir / f"{filename}_validated.xlsx"
    
    # Save the dataframe
    df.to_excel(output_path, index=False, engine="openpyxl")
    
    # Now apply highlights to changed cells
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Blue highlight for changed cells
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for (row_idx, col_idx), (old_val, new_val) in changes.items():
        # Excel rows are 1-indexed and we need to account for header row
        excel_row = row_idx + 2  # +1 for 1-indexing, +1 for header
        excel_col = col_idx + 1  # +1 for 1-indexing
        
        cell = ws.cell(row=excel_row, column=excel_col)
        cell.fill = orange_fill
        cell.value = new_val

    # Apply yellow highlights for errors (no value changes)
    for (row_idx, col_idx) in error_cells:
        excel_row = row_idx + 2
        excel_col = col_idx + 1
        cell = ws.cell(row=excel_row, column=excel_col)
        cell.fill = yellow_fill
    
    wb.save(output_path)
    print(f"\n✓ Validated file saved to: {output_path}")
    return output_path


def main():
    """Main function to run the validator."""
    print("CLFS Data Validator")
    print("=" * 50)

    print("\nModule diagnostics")
    print("-" * 50)
    print(f"rules.__file__: {getattr(rules, '__file__', 'unknown')}")
    ssec_count = len(getattr(rules, "SSEC_CANDIDATES", []) or [])
    print(f"SSEC_CANDIDATES count: {ssec_count}")
    print(f"has validate_qualification_place: {hasattr(rules, 'validate_qualification_place')}")
    
    # Load all .xlsx and .csv files from Operating_Table folder
    files = load_input_files()
    
    print(f"\nTotal files loaded: {len(files)}")
    
    # Display summary of loaded files
    for filename, df in files.items():
        print(f"\n{filename}:")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")

        households = extract_household_members(df)
        total_members = sum(len(members) for members in households)
        print(f"  Households parsed: {len(households)}")
        print(f"  Household members parsed: {total_members}")
        
        # Display household member details
        print(f"\n  Household Member Details:")
        for household_idx, members in enumerate(households, 1):
            print(f"\n  Household {household_idx}:")
            for member_idx, member in enumerate(members, 1):
                print(f"    Member {member_idx}:")
                print(f"      Name: {member.full_name}")
                print(f"      DOB: {member.date_of_birth}")
                print(f"      Age: {member.age}")
                print(f"      Labour Force Status: {member.labour_force_status}")
                print(f"      Employment Status: {member.employment_status_last_week}")
                print(f"      Job Title: {member.job_title}")
        
        # Apply validation rules
        print(f"\n{'=' * 50}")
        print(f"Applying Validation Rules...")
        print(f"{'=' * 50}")
        
        df = _ensure_ssec_column(df)
        df, ssic_col = _ensure_ssic_column(df)
        df, ssoc_groups = _ensure_ssoc_columns(df)
        df, ftpt_changes = _add_ft_pt_columns(df)

        # Track changes and errors for output
        changes = {}
        error_cells = set()
        modified_df = df.copy()
        for row_idx, col_idx, value in ftpt_changes:
            changes[(row_idx, col_idx)] = ("", value)

        rule_errors = []

        # RULE 16: Religion reclass for Others
        religion_col = _find_column_name(list(df.columns), "What is your religion?")
        if religion_col:
            col_idx = df.columns.get_loc(religion_col)
            for row_idx, value in df[religion_col].items():
                if pd.isna(value):
                    continue
                raw = str(value).strip()
                raw_lower = raw.lower()
                if raw_lower.startswith("others:"):
                    text = raw_lower.split(":", 1)[1].strip()
                    for denom, reclass in RELIGION_RECLASS_MAP.items():
                        if denom in text:
                            modified_df.at[row_idx, religion_col] = reclass
                            changes[(row_idx, col_idx)] = (raw, reclass)
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": _get_cell_value(df, row_idx, "Response ID"),
                                "member_index": None,
                                "member": _get_cell_value(df, row_idx, "Full Name"),
                                "rule": "RULE 16",
                                "column": religion_col,
                                "message": f"Reclassified to {reclass}",
                            })
                            break

        # RULE 17: Religion consistency for "No religion"
        if religion_col:
            col_idx = df.columns.get_loc(religion_col)
            for row_idx, value in df[religion_col].items():
                if pd.isna(value):
                    continue
                raw = str(value).strip()
                raw_lower = raw.lower()
                if "no religion" in raw_lower and raw != "No religion":
                    modified_df.at[row_idx, religion_col] = "No religion"
                    changes[(row_idx, col_idx)] = (raw, "No religion")
                    rule_errors.append({
                        "file": filename,
                        "row": row_idx + 1,
                        "response_id": _get_cell_value(df, row_idx, "Response ID"),
                        "member_index": None,
                        "member": _get_cell_value(df, row_idx, "Full Name"),
                        "rule": "RULE 17",
                        "column": religion_col,
                        "message": "Normalized to 'No religion'",
                    })

        # RULE 18: Place of Birth validation for Others
        pob_col = _find_column_name(list(df.columns), "Place of Birth")
        if pob_col:
            col_idx = df.columns.get_loc(pob_col)
            for row_idx, value in df[pob_col].items():
                if pd.isna(value):
                    continue
                raw = str(value).strip()
                raw_lower = raw.lower()
                if raw_lower.startswith("others:"):
                    text = raw_lower.split(":", 1)[1].strip()
                    if text and text not in COUNTRY_LIST:
                        error_cells.add((row_idx, col_idx))
                        rule_errors.append({
                            "file": filename,
                            "row": row_idx + 1,
                            "response_id": _get_cell_value(df, row_idx, "Response ID"),
                            "member_index": None,
                            "member": _get_cell_value(df, row_idx, "Full Name"),
                            "rule": "RULE 18",
                            "column": pob_col,
                            "message": "Invalid country in Others: Place of Birth",
                        })

        ssoc_resources = _load_ssoc_resources()
        ssoc_debug = os.environ.get("SSOC_DEBUG", "").strip().lower() in {"1", "true", "yes"}
        ssoc_use_gmi_hqa = False
        ssoc_debug_fh = None
        ssoc_debug_path = None
        if ssoc_debug:
            output_dir = create_output_directory()
            ssoc_debug_path = output_dir / "ssoc_debug.log"
            ssoc_debug_fh = open(ssoc_debug_path, "w", encoding="utf-8")
        if not ssoc_groups:
            print("  ⚠ SSOC mapping skipped (no Job Title/Main tasks columns found)")
        elif not ssoc_resources:
            print("  ⚠ SSOC mapping skipped (SSOC definitions file not found). Set SSOC_DEFINITIONS_FILE env var.")
        else:
            print("  ✓ SSOC definitions loaded; assigning SSOC codes")
            for row_idx in range(len(df)):
                for group_idx, group in enumerate(ssoc_groups):
                    title_idx = group.get("title_idx")
                    duties_idx = group.get("duties_idx")
                    ssoc_idx = group.get("ssoc_idx")
                    if ssoc_idx is None or duties_idx is None:
                        continue

                    member = None
                    if row_idx < len(households) and group_idx < len(households[row_idx]):
                        member = households[row_idx][group_idx]

                    title_val = df.iat[row_idx, title_idx] if title_idx is not None else ""
                    duties_val = df.iat[row_idx, duties_idx] if duties_idx is not None else ""
                    title_text = "" if pd.isna(title_val) else str(title_val)
                    duties_text = "" if pd.isna(duties_val) else str(duties_val)

                    if not _normalize_text(title_text) and not _normalize_text(duties_text):
                        continue

                    hqa_value = member.highest_academic_qualification if member else None
                    gmi_value = _parse_gmi_value(member.gmi if member else None)

                    ssoc_code, _, _, _, top_5, _ = ssoc.best_match_duties_priority(
                        title_text,
                        duties_text,
                        ssoc_resources["defs"],
                        ssoc_resources["title_map"],
                        ssoc_resources["expert_map"],
                        SSOC_MIN_SCORE,
                        "" if hqa_value is None else str(hqa_value),
                        occ_group_hint_raw=None,
                        company_industry=""
                    )

                    if ssoc_use_gmi_hqa:
                        example_code = _select_candidate_by_examples(top_5 or [], hqa_value, gmi_value)
                        if example_code:
                            ssoc_code = example_code
                        else:
                            required_group = _required_group_from_band(hqa_value, gmi_value)
                            band_code = _select_candidate_by_band(top_5 or [], required_group)
                            if band_code:
                                ssoc_code = band_code

                    if ssoc_debug:
                        top_5_codes = [str(c.get("code", "")).strip() for c in (top_5 or []) if str(c.get("code", "")).strip()]
                        debug_line = (
                            f"SSOC DEBUG row={row_idx + 1} group={group_idx + 1} "
                            f"title='{title_text}' duties='{duties_text}' "
                            f"hqa='{hqa_value}' gmi='{gmi_value}' "
                            f"top5={top_5_codes} selected='{ssoc_code}'"
                        )
                        print(debug_line)
                        if ssoc_debug_fh:
                            ssoc_debug_fh.write(debug_line + "\n")

                    old_val = modified_df.iat[row_idx, ssoc_idx]
                    if str(old_val).strip() != str(ssoc_code).strip():
                        modified_df.iat[row_idx, ssoc_idx] = ssoc_code
                        changes[(row_idx, ssoc_idx)] = (old_val, ssoc_code)

            if ssoc_debug_fh:
                ssoc_debug_fh.close()
                print(f"  ✓ SSOC debug log saved to: {ssoc_debug_path}")

        if ssic_col and STRATA_LOOKUP:
            print("  ✓ SSIC lookup loaded; assigning SSIC codes")
            est_col = _find_column_name(list(df.columns), "Name of Establishment you were working last week?")
            ssic_matched_col, ssic_idx = _get_column_index(df, "SSIC Code")
            if est_col and ssic_matched_col is not None and ssic_idx is not None:
                for row_idx in range(len(df)):
                    est_val = df.at[row_idx, est_col]
                    if pd.isna(est_val) or str(est_val).strip() == "":
                        continue
                    est_norm = _normalize_text(est_val)
                    match = next(
                        (code for name, code in STRATA_LOOKUP if est_norm in name),
                        None
                    )
                    if match:
                        old_val = modified_df.iat[row_idx, ssic_idx]
                        if str(old_val).strip() != str(match).strip():
                            modified_df.iat[row_idx, ssic_idx] = match
                            changes[(row_idx, ssic_idx)] = (old_val, match)
                    else:
                        error_cells.add((row_idx, ssic_idx))
                        rule_errors.append({
                            "file": filename,
                            "row": row_idx + 1,
                            "response_id": _get_cell_value(df, row_idx, "Response ID"),
                            "member_index": None,
                            "member": _get_cell_value(df, row_idx, "Full Name"),
                            "rule": "RULE 14",
                            "column": ssic_matched_col,
                            "message": "Unable to match SSIC Code from establishment name",
                        })
        elif ssic_col:
            print("  ⚠ SSIC lookup skipped (STRATA_LOOKUP is empty)")
        
        # RULE 1: Others option validation
        print(f"\nRULE 1: Others option validation")
        print("-" * 50)
        
        rule1_issues = 0
        rule1_corrected = 0
        
        # Check all columns with "Others:" options
        for attr_name, question_config in rules.QUESTIONS_WITH_OTHERS.items():
            col_name = question_config["column_name"]

            matched_col = _find_column_name(list(df.columns), col_name)
            if not matched_col:
                print(f"  ⚠ Column '{col_name}' not found in data")
                continue

            col_idx = df.columns.get_loc(matched_col)

            for row_idx, value in df[matched_col].items():
                if pd.isna(value):
                    continue
                
                result = rules.validate_others_option(str(value), attr_name)
                
                if result.corrected_value and result.corrected_value != str(value):
                    print(f"  ✓ Row {row_idx + 1} ({col_name}): {result.message}")
                    print(f"    Before: {result.original_value}")
                    print(f"    After:  {result.corrected_value}")
                    modified_df.at[row_idx, matched_col] = result.corrected_value
                    changes[(row_idx, col_idx)] = (result.original_value, result.corrected_value)
                    rule1_corrected += 1
                    response_id = _get_cell_value(df, row_idx, "Response ID")
                    member_name = _get_cell_value(df, row_idx, "Full Name")
                    rule_errors.append({
                        "file": filename,
                        "row": row_idx + 1,
                        "response_id": response_id,
                        "member_index": None,
                        "member": member_name,
                        "rule": f"RULE 1 - {col_name}",
                        "column": matched_col,
                        "message": result.message
                    })
        
        print(f"\nRULE 1 Summary: {rule1_corrected} corrected")
        
        # RULE 2-13: Additional validation rules from colleague's work
        print(f"\nRULES 2-13: Data quality validations")
        print("-" * 50)

        ssec_enabled = bool(getattr(rules, "SSEC_CANDIDATES", []))
        if not ssec_enabled:
            print("  ⚠ SSEC mapping skipped (SSEC_CANDIDATES is empty)")
        
        # Iterate through all household members for validation
        for household_idx, members in enumerate(households, 1):
            for member_idx, member in enumerate(members, 1):
                row_idx = household_idx - 1  # Adjust for 0-based indexing
                response_id = _get_cell_value(df, row_idx, "Response ID")
                
                # RULE 2: Age started employment validation
                if member.age_started_employment is not None:
                    result = rules.validate_age_started_employment(member.age_started_employment)
                    if not result.is_valid:
                        col_name = "At what age did you start employment"
                        matched_col, col_idx = _get_column_index(df, col_name)
                        if matched_col is not None and col_idx is not None:
                            error_cells.add((row_idx, col_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 2",
                                "column": matched_col,
                                "message": result.message
                            })
                
                # RULE 3: Bonus validation
                if member.bonus_received_last_12_months is not None:
                    result = rules.validate_bonus(member.bonus_received_last_12_months)
                    if not result.is_valid:
                        col_name = "Bonus received from your job(s) during the last 12 months"
                        matched_col, col_idx = _get_column_index(df, col_name)
                        if matched_col is not None and col_idx is not None:
                            error_cells.add((row_idx, col_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 3",
                                "column": matched_col,
                                "message": result.message
                            })
                
                # RULE 4: Previous company name validation
                if member.establishment_name_last_worked is not None:
                    result = rules.validate_previous_company_name(member.establishment_name_last_worked)
                    if not result.is_valid:
                        col_name = "Name of Establishment you were working last worked"
                        matched_col, col_idx = _get_column_index(df, col_name)
                        if matched_col is not None and col_idx is not None:
                            error_cells.add((row_idx, col_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 4",
                                "column": matched_col,
                                "message": result.message
                            })

                # RULE 15: Current establishment name validation
                if member.name_of_establishment_last_week is not None:
                    result = rules.validate_previous_company_name(member.name_of_establishment_last_week)
                    if not result.is_valid:
                        col_name = "Name of Establishment you were working last week?"
                        matched_col, col_idx = _get_column_index(df, col_name)
                        if matched_col is not None and col_idx is not None:
                            error_cells.add((row_idx, col_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 15",
                                "column": matched_col,
                                "message": result.message
                            })
                
                # RULE 5: Interest from savings validation
                if member.interest_from_savings_last_12_months is not None:
                    result = rules.validate_interest_from_savings(member.interest_from_savings_last_12_months)
                    if not result.is_valid:
                        col_name = "How much interest did you receive from savings (e.g., current and saving accounts, fixed deposits) in the last 12 months?"
                        matched_col, col_idx = _get_column_index(df, col_name)
                        if matched_col is not None and col_idx is not None:
                            error_cells.add((row_idx, col_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 5",
                                "column": matched_col,
                                "message": result.message
                            })
                
                # RULE 6: Dividends/investment interest validation
                if member.dividends_interests_investments_last_12_months is not None:
                    result = rules.validate_dividends_investment_interest(member.dividends_interests_investments_last_12_months)
                    if not result.is_valid:
                        col_name = "How much dividends and interests did you receive from other investment sources (e.g., bonds, shares, unit trust, personal loans to persons outside your households) in the last 12 months?"
                        matched_col, col_idx = _get_column_index(df, col_name)
                        if matched_col is not None and col_idx is not None:
                            error_cells.add((row_idx, col_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 6",
                                "column": matched_col,
                                "message": result.message
                            })
                
                # RULE 7: Freelance work vs Own Account Worker consistency
                if member.freelance_online_platforms_last_12_months is not None:
                    result = rules.validate_freelance_employment_consistency(
                        member.employment_status_last_week,
                        member.freelance_online_platforms_last_12_months
                    )
                    if not result.is_valid:
                        # Highlight both employment status and freelance columns
                        emp_col = "Employment Status as of last week"
                        free_col = "Did you perform any freelance or assignment-based work via any of the following online platform(s) in the last 12 months?"
                        emp_matched, emp_idx = _get_column_index(df, emp_col)
                        if emp_matched is not None and emp_idx is not None:
                            error_cells.add((row_idx, emp_idx))
                        free_matched, free_idx = _get_column_index(df, free_col)
                        if free_matched is not None and free_idx is not None:
                            error_cells.add((row_idx, free_idx))
                        rule_errors.append({
                            "file": filename,
                            "row": row_idx + 1,
                            "response_id": response_id,
                            "member_index": member_idx,
                            "member": member.full_name,
                            "rule": "RULE 7",
                            "column": f"{emp_matched or emp_col} & {free_matched or free_col}",
                            "message": result.message
                        })

                # RULE 19: Freelance requires self-employed and own-account
                freelance_val = _normalize_text(member.freelance_online_platforms_last_12_months)
                if freelance_val and freelance_val != _normalize_text(NO_FREELANCE_TEXT):
                    se_val = _normalize_text(member.self_employed_last_12_months)
                    oa_val = _normalize_text(member.worked_own_business_last_12_months)
                    if se_val != "yes" or oa_val != "yes":
                        self_col = "At any point in the last 12 months, were you self-employed?"
                        own_col = "At any point in the last 12 months, did you work on your own (i.e., without paid employees) while running your own business or trade?"
                        free_col = "Did you perform any freelance or assignment-based work via any of the following online platform(s) in the last 12 months?"
                        self_matched, self_idx = _get_column_index(df, self_col)
                        own_matched, own_idx = _get_column_index(df, own_col)
                        free_matched, free_idx = _get_column_index(df, free_col)
                        for idx in [self_idx, own_idx, free_idx]:
                            if idx is not None:
                                error_cells.add((row_idx, idx))
                        rule_errors.append({
                            "file": filename,
                            "row": row_idx + 1,
                            "response_id": response_id,
                            "member_index": member_idx,
                            "member": member.full_name,
                            "rule": "RULE 19",
                            "column": f"{self_matched or self_col} & {own_matched or own_col} & {free_matched or free_col}",
                            "message": "Freelance selected but self-employed/own-account not both Yes",
                        })

                # RULE 8: Validate Highest Academic Qualification vs Place of Study
                qualification = member.highest_academic_qualification
                place = member.place_of_study_highest_academic
                if qualification and place:
                    matches = rules.validate_qualification_place(str(qualification), str(place))
                    if matches:
                        qual_col = "Highest Academic Qualification"
                        place_col = "Place of study for your Highest Academic Attained in?"
                        qual_matched, qual_idx = _get_column_index(df, qual_col)
                        if qual_matched is not None and qual_idx is not None:
                            error_cells.add((row_idx, qual_idx))
                        place_matched, place_idx = _get_column_index(df, place_col)
                        if place_matched is not None and place_idx is not None:
                            error_cells.add((row_idx, place_idx))

                        for match in matches:
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": f"RULE 8 - {match['rule_id']}",
                                "column": f"{qual_matched or qual_col} & {place_matched or place_col}",
                                "message": match["reason"]
                            })

                # RULE 9: Assign SSEC Code based on Highest Academic Qualification
                
                # RULE 10: Internship/Employment type validation
                internship_value = member.paid_internship_traineeship
                employment_value = member.type_of_employment
                result = rules.validate_internship_employment_rule(internship_value, employment_value)
                if not result.is_valid:
                    internship_col = "Was your main job last week a paid internship, traineeship or apprenticeship?"
                    employment_col = "Type of Employment?"
                    employment_matched, employment_idx = _get_column_index(df, employment_col)
                    if employment_matched is not None and employment_idx is not None:
                        error_cells.add((row_idx, employment_idx))
                    internship_matched, internship_idx = _get_column_index(df, internship_col)
                    if internship_matched is not None and internship_idx is not None:
                        error_cells.add((row_idx, internship_idx))
                    rule_errors.append({
                        "file": filename,
                        "row": row_idx + 1,
                        "response_id": response_id,
                        "member_index": member_idx,
                        "member": member.full_name,
                        "rule": "RULE 10",
                        "column": f"{internship_matched or internship_col} & {employment_matched or employment_col}",
                        "message": result.message
                    })

                # RULE 11: Job title validation
                result = rules.validate_job_title_rule(member.job_title)
                if not result.is_valid:
                    job_col = "Job Title"
                    job_matched, job_idx = _get_column_index(df, job_col)
                    if job_matched is not None and job_idx is not None:
                        error_cells.add((row_idx, job_idx))
                    rule_errors.append({
                        "file": filename,
                        "row": row_idx + 1,
                        "response_id": response_id,
                        "member_index": member_idx,
                        "member": member.full_name,
                        "rule": "RULE 11",
                        "column": job_matched or job_col,
                        "message": result.message
                    })

                # RULE 13: Usual hours of work must be numeric
                result = rules.validate_usual_hours_value(member.usual_hours_of_work)
                if not result.is_valid:
                    hours_col = "Usual hours of work"
                    hours_matched, hours_idx = _get_column_index(df, hours_col)
                    if hours_matched is not None and hours_idx is not None:
                        error_cells.add((row_idx, hours_idx))
                    rule_errors.append({
                        "file": filename,
                        "row": row_idx + 1,
                        "response_id": response_id,
                        "member_index": member_idx,
                        "member": member.full_name,
                        "rule": "RULE 13",
                        "column": hours_matched or hours_col,
                        "message": result.message
                    })
                if ssec_enabled and qualification:
                    ssec_code, ssec_score = rules.best_ssec_match(str(qualification))
                    ssec_col, ssec_idx = _get_column_index(df, "SSEC Code")
                    if ssec_col is not None and ssec_idx is not None:
                        if ssec_code:
                            modified_df.at[row_idx, ssec_col] = ssec_code
                            changes[(row_idx, ssec_idx)] = ("", ssec_code)
                        else:
                            error_cells.add((row_idx, ssec_idx))
                            rule_errors.append({
                                "file": filename,
                                "row": row_idx + 1,
                                "response_id": response_id,
                                "member_index": member_idx,
                                "member": member.full_name,
                                "rule": "RULE 9",
                                "column": ssec_col,
                                "message": "Unable to map SSEC Code from Highest Academic Qualification"
                            })
        
        # Display errors found
        if rule_errors:
            print(f"\n  ✗ Found {len(rule_errors)} validation errors:")
            for error in rule_errors:
                print(f"    Row {error['row']} - {error['member']}")
                print(f"    {error['rule']}: {error['message']}")
                print(f"    Column: {error['column']}")
                print()
        else:
            print(f"  ✓ No validation errors found")
        
        print(f"\nRULES 2-13 Summary: {len(rule_errors)} errors found")

        # Create validation report (summary + details)
        create_validation_report(rule_errors, filename)
        
        # Save validated output if changes were made
        if changes or error_cells:
            original_path = Path("Operating_Table") / filename
            save_with_highlights(modified_df, str(original_path), changes, error_cells)


if __name__ == "__main__":
    main()
