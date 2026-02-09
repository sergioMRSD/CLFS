import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import traceback

# --- CONFIGURATION ---
FOLDER_PATH = r'C:\Users\momgjq\OneDrive - SG Govt M365\Desktop\CLFS_code' 
STRATA_FILE = 'SSIC_List_2Sep2025 1.xlsx'
COUNTRIES_FILE = 'Countries.xlsx'
RELIGION_FILE = 'Religion list.xlsx'
REPORT_NAME = 'validation_error_report.xlsx'

# Formatting Styles
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# Header Names
HDR_ESTABLISHMENT = 'Name of Establishment you were working last week?'
HDR_SSIC = 'SSIC Code'
HDR_AGE_START = 'At what age did you start employment'
HDR_BONUS = 'Bonus received from your job(s) during the last 12 months'
HDR_PREV_ESTAB = 'Name of Establishment you were working last worked'
HDR_RELIGION = 'What is your religion?'
HDR_BIRTH_PLACE = 'Place of Birth'
HDR_SELF_EMPLOYED = 'At any point in the last 12 months, were you self-employed?'
HDR_OWN_ACCOUNT = 'At any point in the last 12 months, did you work on your own (i.e., without paid employees) while running your own business or trade?'
HDR_FREELANCE = 'Did you perform any freelance or assignment-based work via any of the following online platform(s) in the last 12 months?'

NO_FREELANCE_TEXT = 'I did not take up freelance or assignment-based work through online platforms in the last 12 months'

def run_validation_pipeline():
    print("Loading reference files...")
    strata_lookup = load_strata_lookup()
    country_list = load_country_list()
    religion_map = load_religion_reclass_list()
    
    all_errors = [] # Collect all errors from all files first
    
    for filename in os.listdir(FOLDER_PATH):
        if (filename.endswith('.xlsx') or filename.endswith('.xls')) and \
           filename not in [STRATA_FILE, COUNTRIES_FILE, RELIGION_FILE, REPORT_NAME]:
            
            file_path = os.path.join(FOLDER_PATH, filename)
            print(f"Processing: {filename}")
            
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                file_error_log = []

                # Execute Rules
                file_error_log = execute_rule_ssic(ws, filename, strata_lookup, file_error_log)
                file_error_log = execute_rule_employment_age(ws, filename, file_error_log)
                file_error_log = execute_rule_bonus(ws, filename, file_error_log)
                file_error_log = execute_rule_prev_estab(ws, filename, file_error_log)
                file_error_log = execute_rule_current_estab(ws, filename, file_error_log)
                
                if religion_map:
                    file_error_log = execute_rule_religion_reclass(ws, filename, religion_map, file_error_log)
                
                file_error_log = execute_rule_religion_consistency(ws, filename, file_error_log)

                if country_list:
                    file_error_log = execute_rule_place_of_birth(ws, filename, country_list, file_error_log)

                file_error_log = execute_rule_freelance_consistency(ws, filename, file_error_log)

                wb.save(file_path)
                all_errors.extend(file_error_log)
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                traceback.print_exc()

    if all_errors:
        save_report_to_excel(all_errors)
    
    print("\n--- Validation Pipeline Finished ---")

# --- UTILITY FUNCTIONS ---

def is_valid_row(ws, row_idx):
    row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
    non_empty = sum(1 for val in row_values if val is not None and str(val).strip() != "")
    return non_empty >= 2

def save_report_to_excel(log_data):
    """Saves the error log to an Excel workbook, appending if it exists."""
    report_path = os.path.join(FOLDER_PATH, REPORT_NAME)
    new_df = pd.DataFrame(log_data)
    
    if os.path.exists(report_path):
        try:
            existing_df = pd.read_excel(report_path)
            final_df = pd.concat([existing_df, new_df], ignore_index=True)
        except:
            final_df = new_df
    else:
        final_df = new_df

    # Write to Excel with auto-adjusted column widths
    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='Errors')
        worksheet = writer.sheets['Errors']
        for idx, col in enumerate(final_df.columns):
            max_len = max(final_df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50) # Cap width at 50

    print(f"Error report updated: {REPORT_NAME}")

# --- DATA LOADERS ---

def load_strata_lookup():
    path = os.path.join(FOLDER_PATH, STRATA_FILE)
    if not os.path.exists(path): return None
    df = pd.read_excel(path, usecols="C,G", skiprows=1, nrows=38442, header=None)
    df = df.dropna(subset=[df.columns[0]])
    return [(str(row[df.columns[0]]).lower(), row[df.columns[1]]) for _, row in df.iterrows()]

def load_country_list():
    path = os.path.join(FOLDER_PATH, COUNTRIES_FILE)
    if not os.path.exists(path): return None
    df = pd.read_excel(path, sheet_name='Countries', header=None, usecols="A", nrows=195)
    return set(df.iloc[:, 0].astype(str).str.strip().str.lower())

def load_religion_reclass_list():
    path = os.path.join(FOLDER_PATH, RELIGION_FILE)
    if not os.path.exists(path): return None
    df = pd.read_excel(path, sheet_name='Religion', usecols="A,B", nrows=19)
    return [(str(row[df.columns[0]]).strip().lower(), str(row[df.columns[1]]).strip()) for _, row in df.dropna().iterrows()]

# --- VALIDATION RULES ---

def execute_rule_ssic(ws, filename, strata_lookup, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_ESTABLISHMENT]
    for col_idx in sorted(target_indices, reverse=True):
        ssic_col = col_idx + 1
        if ws.cell(row=header_row, column=ssic_col).value != HDR_SSIC:
            ws.insert_cols(ssic_col)
            ws.cell(row=header_row, column=ssic_col).value = HDR_SSIC
        
        if strata_lookup:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                if not is_valid_row(ws, row_idx): continue
                input_val = ws.cell(row=row_idx, column=col_idx).value
                if not input_val: continue
                match = next((code for name, code in strata_lookup if str(input_val).strip().lower() in name), None)
                if match: ws.cell(row=row_idx, column=ssic_col).value = match
                else:
                    ws.cell(row=row_idx, column=ssic_col).fill = YELLOW_FILL
                    error_log.append({'File': filename, 'Row': row_idx, 'Column': 'SSIC Code', 'Input': input_val, 'Error': 'No match found'})
    return error_log

def execute_rule_religion_reclass(ws, filename, religion_map, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_RELIGION]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value or "").strip().lower()
            if val.startswith("others:"):
                text = val.split(':', 1)[1].strip()
                for denom, reclass in religion_map:
                    if denom in text:
                        error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_RELIGION, 'Input': cell.value, 'Error': f'Reclassified to {reclass}'})
                        cell.value = reclass
                        cell.fill = GREEN_FILL
                        break
    return error_log

def execute_rule_religion_consistency(ws, filename, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_RELIGION]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None: continue
            raw_val = str(cell.value).strip()
            val_lower = raw_val.lower()
            if "no religion" in val_lower:
                if val_lower != "no religion":
                    error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_RELIGION, 'Input': raw_val, 'Error': "Conflict with 'No religion'. Corrected."})
                    cell.value = "No religion"
                    cell.fill = GREEN_FILL
    return error_log

def execute_rule_employment_age(ws, filename, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_AGE_START]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None or str(val).strip() == "": continue
            if not isinstance(val, int) or isinstance(val, bool) or not (13 <= val <= 100):
                ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL
                error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_AGE_START, 'Input': val, 'Error': 'Must be integer 13-100'})
    return error_log

def execute_rule_bonus(ws, filename, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_BONUS]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            raw = ws.cell(row=row_idx, column=col_idx).value
            if raw is None or str(raw).strip() == "": continue
            s = str(raw)
            if ',' in s or '-' in s:
                ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL
                error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_BONUS, 'Input': raw, 'Error': 'Invalid symbols detected'})
            else:
                try:
                    num = float(raw)
                    if not (0 <= num <= 99): raise ValueError
                except:
                    ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL
                    error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_BONUS, 'Input': raw, 'Error': 'Out of 0-99 range'})
    return error_log

def execute_rule_prev_estab(ws, filename, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_PREV_ESTAB]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if not val: continue
            if val.isdigit() or sum(c.isalpha() for c in val) < 3:
                ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL
                error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_PREV_ESTAB, 'Input': val, 'Error': 'Invalid establishment name'})
    return error_log

def execute_rule_current_estab(ws, filename, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_ESTABLISHMENT]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if not val: continue
            if val.isdigit() or sum(c.isalpha() for c in val) < 3:
                ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL
                error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_ESTABLISHMENT, 'Input': val, 'Error': 'Invalid establishment name'})
    return error_log

def execute_rule_place_of_birth(ws, filename, country_list, error_log):
    header_row = 6
    target_indices = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_BIRTH_PLACE]
    for col_idx in target_indices:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            cell = ws.cell(row=row_idx, column=col_idx)
            val = str(cell.value or "").strip().lower()
            if val.startswith("others:"):
                parts = val.split(':', 1)
                if len(parts) > 1 and parts[1].strip() not in country_list:
                    cell.fill = YELLOW_FILL
                    error_log.append({'File': filename, 'Row': row_idx, 'Column': HDR_BIRTH_PLACE, 'Input': cell.value, 'Error': 'Invalid Country'})
    return error_log

def execute_rule_freelance_consistency(ws, filename, error_log):
    header_row = 6
    h1 = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_SELF_EMPLOYED]
    h2 = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_OWN_ACCOUNT]
    h3 = [c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value == HDR_FREELANCE]
    for i in range(min(len(h1), len(h2), len(h3))):
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not is_valid_row(ws, row_idx): continue
            v1 = str(ws.cell(row=row_idx, column=h1[i]).value or "").strip()
            v2 = str(ws.cell(row=row_idx, column=h2[i]).value or "").strip()
            v3 = str(ws.cell(row=row_idx, column=h3[i]).value or "").strip()
            if v3 != "" and v3 != NO_FREELANCE_TEXT:
                if v1 != "Yes" or v2 != "Yes":
                    for c in [h1[i], h2[i], h3[i]]: ws.cell(row=row_idx, column=c).fill = ORANGE_FILL
                    error_log.append({'File': filename, 'Row': row_idx, 'Column': f'Member {i+1} Freelance', 'Input': v3, 'Error': 'Freelance inconsistent with status'})
    return error_log

if __name__ == "__main__":
    run_validation_pipeline()
