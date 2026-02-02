# -*- coding: utf-8 -*-
"""
Created on Mon Feb  2 14:26:31 2026

@author: momgjq
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import re

# =========================
# CONFIGURATION
# =========================
FOLDER_PATH = r"C:\Users\MOMGJQ\Desktop\CLFS"   # <-- EDIT THIS
ERROR_REPORT_PATH = "error_report.xlsx"
HEADER_ROW = 6

# Validation headers
AGE_HEADER = "At what age did you start employment"
BONUS_HEADER = "Bonus received from your job(s) during the last 12 months"
PREV_COMPANY_HEADER = "Name of Establishment you were working last worked"
INTEREST_HEADER = "How much interest did you receive from savings (e.g., current and saving accounts, fixed deposits) in the last 12 months?"
DIVIDENDS_HEADER = "How much dividends and interests did you receive from other investment sources (e.g., bonds, shares, unit trust, personal loans to persons outside your households) in the last 12 months?"

# Headers for freelance vs Own Account Worker validation
EMPLOYMENT_STATUS_HEADER = "Employment Status as of last week"
FREELANCE_HEADER = "Did you perform any freelance or assignment-based work via any of the following online platform(s) in the last 12 months?"

ERROR_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

error_log = []

# =========================
# HELPER FUNCTIONS
# =========================
def row_has_min_filled_cells(row_cells, min_count=2):
    filled = [cell for cell in row_cells if cell.value not in (None, "")]
    return len(filled) >= min_count

def is_valid_age(value):
    if not isinstance(value, (int, float)):
        return False
    if isinstance(value, float) and not value.is_integer():
        return False
    value = int(value)
    return 13 <= value <= 100

def is_valid_bonus(value):
    if isinstance(value, str):
        if "," in value or "-" in value:
            return False
        try:
            value = float(value)
        except ValueError:
            return False
    elif isinstance(value, (int, float)):
        value = float(value)
    else:
        return False
    return 0 <= value <= 99

def is_valid_prev_company(value):
    if not isinstance(value, str):
        value = str(value)
    letters = re.findall(r"[A-Za-z]", value)
    if len(letters) < 3:
        return False
    numeric_only = value.replace(" ", "").isdigit()
    if numeric_only:
        return False
    return True

def is_valid_interest(value):
    try:
        value = float(value)
    except (ValueError, TypeError):
        return False
    return 0 <= value <= 10

def is_valid_dividends(value):
    try:
        value = float(value)
    except (ValueError, TypeError):
        return False
    return 0 <= value <= 50

def run_validations(ws, row_idx, header_columns_dict):
    errors = []

    # ----- AGE VALIDATION -----
    for col_idx in header_columns_dict.get(AGE_HEADER, []):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value in (None, ""):
            continue
        if not is_valid_age(value):
            errors.append({
                "column": AGE_HEADER,
                "cell": cell,
                "message": "Invalid age. Must be a whole number between 13 and 100."
            })

    # ----- BONUS VALIDATION -----
    for col_idx in header_columns_dict.get(BONUS_HEADER, []):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value in (None, ""):
            continue
        if not is_valid_bonus(value):
            errors.append({
                "column": BONUS_HEADER,
                "cell": cell,
                "message": "Invalid bonus. Must be numeric between 0 and 99, no commas or minus signs."
            })

    # ----- PREVIOUS COMPANY VALIDATION -----
    for col_idx in header_columns_dict.get(PREV_COMPANY_HEADER, []):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value in (None, ""):
            continue
        if not is_valid_prev_company(value):
            errors.append({
                "column": PREV_COMPANY_HEADER,
                "cell": cell,
                "message": "Invalid company name. Must contain at least 3 letters and not be purely numeric."
            })

    # ----- INTEREST VALIDATION -----
    for col_idx in header_columns_dict.get(INTEREST_HEADER, []):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value in (None, ""):
            continue
        if not is_valid_interest(value):
            errors.append({
                "column": INTEREST_HEADER,
                "cell": cell,
                "message": "Invalid interest. Must be numeric between 0 and 10 (decimals allowed)."
            })

    # ----- DIVIDENDS VALIDATION -----
    for col_idx in header_columns_dict.get(DIVIDENDS_HEADER, []):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value in (None, ""):
            continue
        if not is_valid_dividends(value):
            errors.append({
                "column": DIVIDENDS_HEADER,
                "cell": cell,
                "message": "Invalid dividends/other investment interest. Must be numeric between 0 and 50 (decimals allowed)."
            })

    # ----- FREELANCE VS OWN ACCOUNT WORKER VALIDATION -----
    employment_cols = header_columns_dict.get(EMPLOYMENT_STATUS_HEADER, [])
    freelance_cols = header_columns_dict.get(FREELANCE_HEADER, [])

    # Pair them for each household member
    for emp_col, free_col in zip(employment_cols, freelance_cols):
        emp_cell = ws.cell(row=row_idx, column=emp_col)
        free_cell = ws.cell(row=row_idx, column=free_col)

        if free_cell.value in (None, ""):
            continue  # Blank is ignored

        free_text = str(free_cell.value).strip()
        # If the household member did any freelance work (i.e., not the "I did not take up..." option)
        if free_text != "I did not take up freelance or assignment-based work through online platforms in the last 12 months":
            # Must be Own Account Worker
            emp_text = str(emp_cell.value).strip() if emp_cell.value is not None else ""
            if emp_text != "Own Account Worker (Self-employed without paid employees)":
                # Highlight both cells and log error
                emp_cell.fill = ERROR_FILL
                free_cell.fill = ERROR_FILL
                errors.append({
                    "column": f"{EMPLOYMENT_STATUS_HEADER} & {FREELANCE_HEADER}",
                    "cell": f"Row {row_idx}",
                    "message": "Mismatch: Freelance work selected but Employment Status is not Own Account Worker."
                })

    return errors

# =========================
# MAIN PROCESSING
# =========================
for file_name in os.listdir(FOLDER_PATH):
    if not file_name.endswith(".xlsx"):
        continue

    file_path = os.path.join(FOLDER_PATH, file_name)
    wb = load_workbook(file_path)

    for ws in wb.worksheets:
        max_col = ws.max_column

        # Identify columns for each header
        header_columns_dict = {}
        for col_idx in range(1, max_col + 1):
            header_value = ws.cell(row=HEADER_ROW, column=col_idx).value
            if header_value in [
                AGE_HEADER,
                BONUS_HEADER,
                PREV_COMPANY_HEADER,
                INTEREST_HEADER,
                DIVIDENDS_HEADER,
                EMPLOYMENT_STATUS_HEADER,
                FREELANCE_HEADER
            ]:
                header_columns_dict.setdefault(header_value, []).append(col_idx)

        if not header_columns_dict:
            continue

        # Scan rows below header
        for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
            row_cells = ws[row_idx]
            if not row_has_min_filled_cells(row_cells):
                continue

            errors = run_validations(ws, row_idx, header_columns_dict)
            for err in errors:
                # Skip freelance mismatch errors because they are already highlighted
                if "Mismatch" not in err["message"]:
                    cell = err["cell"]
                    cell.fill = ERROR_FILL

                error_log.append({
                    "File Name": file_name,
                    "Sheet Name": ws.title,
                    "Row": row_idx,
                    "Column": err["column"],
                    "Error Message": err["message"]
                })

    wb.save(file_path)

# =========================
# WRITE ERROR REPORT
# =========================
if error_log:
    df_errors = pd.DataFrame(error_log)
    df_errors.to_excel(ERROR_REPORT_PATH, index=False)
else:
    print("No errors detected.")
