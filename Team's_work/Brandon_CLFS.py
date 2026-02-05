# -*- coding: utf-8 -*-
"""
Created on Wed Jan 28 14:14:23 2026

@author: momlyx14
"""
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =========================
# CONFIGURATION
# =========================
EXCEL_FOLDER = Path(r"C:\Users\momlyx14\OneDrive - SG Govt M365\Desktop\CLFS Test 2\Test")  # <-- change this
ERROR_REPORT_PATH = Path(r"C:\Users\momlyx14\OneDrive - SG Govt M365\Desktop\CLFS Test 2\error_report.xlsx")

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

VALIDATED_FOLDER = Path(r"C:\Users\momlyx14\OneDrive - SG Govt M365\Desktop\CLFS Test 2\Validated")
VALIDATED_FOLDER.mkdir(exist_ok=True)


# =========================
# Class Creation
# =========================

class IndividualResponse:
    def __init__(self, file_name, sheet_name, excel_row):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.excel_row = excel_row
        self.answers = {}      # header -> list of (value, column_index)

    def add_answer(self, header, value, col_idx):
        if header not in self.answers:
            self.answers[header] = []
        self.answers[header].append({
            "value": value,
            "column": col_idx
        })

    def get_answers(self, header):
        return self.answers.get(header, [])

# =========================
# HELPER FUNCTIONS
# =========================
def get_excel_files(folder: Path):
    return list(folder.glob("*.xlsx"))


def row_has_min_cells(ws, row, min_cells=2):
    filled = 0
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            filled += 1
        if filled >= min_cells:
            return True
    return False


def find_headers(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=6, column=col).value #looks from row 6 for the headers
        if value:
            header = normalise_header(value)
            headers.setdefault(header, []).append(col)
    return headers


def build_individuals(ws, headers, file_name, sheet_name,  data_start_row=7):
    individuals = []

    for row in range(data_start_row, ws.max_row + 1):
        if not row_has_min_cells(ws, row):
            continue

        individual = IndividualResponse(
            file_name=file_name,
            sheet_name=sheet_name,
            excel_row=row
        )

        for header, columns in headers.items():
            for col_idx in columns:
                value = ws.cell(row=row, column=col_idx).value
                individual.add_answer(header, value, col_idx)

        individuals.append(individual)

    return individuals

# normalise function
def normalise(value):
    if value is None:
        return ""
    return str(value).strip().lower()

#normalise header function
def normalise_header(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split()).lower()


# =========================
# VALIDATION PLACEHOLDER
# =========================
def validate_individual(individual):
    errors = []

    errors.extend(
        validate_internship_employment_rule(individual)
    )

    errors.extend(validate_job_title_rule(individual))    

    # Future rules go here
    # errors.extend(validate_some_other_rule(individual))

    return errors

def validate_internship_employment_rule(individual):
    """
    Rule:
    If 'Was your main job last week a paid internship, traineeship or apprenticeship?' == 'Yes'
    then 'Type of Employment?' MUST be 'Fixed-Term contract employee'
    """

    errors = []

    internship_header = normalise_header(
    "Was your main job last week a paid internship, traineeship or apprenticeship?"
    )
    employment_header = normalise_header("Type of Employment?")


    internship_answers = individual.get_answers(internship_header)
    employment_answers = individual.get_answers(employment_header)

    # If either header does not exist, rule is not applicable
    if not internship_answers or not employment_answers:
        return errors

    # Check if ANY internship answer is "Yes"
    internship_yes = any(
       normalise(ans["value"]) == "yes" for ans in internship_answers
    )

    if not internship_yes:
        return errors

    # Internship = Yes → employment must be Fixed-Term
    for emp in employment_answers:
        if normalise(emp["value"]) in (
            "permanent employee",
            "casual/on-call employee"
        ):
            errors.append({
                "file": individual.file_name,
                "sheet": individual.sheet_name,
                "row": individual.excel_row,
                "column": emp["column"],
                "header": employment_header,
                "error": (
                    "Internship/Traineeship/Apprenticeship must be "
                    "Fixed-Term contract employee"
                )
            })
    
    

    return errors

def validate_job_title_rule(individual):
    """
    Rule:
    'Job Title' must be at least 4 letters and cannot contain numbers.
    """

    errors = []

    header = normalise_header("Job Title")
    answers = individual.get_answers(header)

    if not answers:
        return errors  # header not present, skip

    for ans in answers:
        raw_value = ans["value"]
        if raw_value is None:
            continue  # skip empty cell        
        value = normalise(raw_value)
        if not value:
            continue  # skip empty string after normalisation

        if len(value) < 4 or any(char.isdigit() for char in value):
            errors.append({
                "file": individual.file_name,
                "sheet": individual.sheet_name,
                "row": individual.excel_row,
                "column": ans["column"],
                "header": header,
                "error": "Job Title must be at least 4 letters and contain no numbers"
            })

    return errors

def add_ft_pt_column(ws, headers):
    """
    Adds a column 'FT/PT' next to 'Usual hours of work'
    FT if hours > 35
    PT if hours < 35
    """

    usual_hours_header = normalise_header("Usual hours of work")

    if usual_hours_header not in headers:
        return  # header not found → do nothing

 # IMPORTANT:
    # Iterate in REVERSE order so column insertion doesn't shift later indices
    for hours_col in sorted(headers[usual_hours_header], reverse=True):

        # Step 1: check if there is at least ONE value below this column
        has_any_value = False
        for row in range(7, ws.max_row + 1):
            if ws.cell(row=row, column=hours_col).value not in (None, ""):
                has_any_value = True
                break

        if not has_any_value:
            continue  # skip completely empty hours column

        # Step 2: insert FT/PT column beside it
        ft_pt_col = hours_col + 1
        ws.insert_cols(ft_pt_col)


        # Step 3: Write header
        ws.cell(row=6, column=ft_pt_col).value = "FT/PT"

        # Fill values
        for row in range(7, ws.max_row + 1):
            value = ws.cell(row=row, column=hours_col).value

            if value in (None, ""):
                continue

            try:
                hours = float(value)
            except ValueError:
                continue

            if hours >= 35:
                ws.cell(row=row, column=ft_pt_col).value = "FT"
            else:
                ws.cell(row=row, column=ft_pt_col).value = "PT"


# =========================
# MAIN PROCESSOR
# =========================
def process_files():
    error_log = []

    for excel_file in get_excel_files(EXCEL_FOLDER):
        print(f"Processing {excel_file.name}")
        wb = load_workbook(excel_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = find_headers(ws)
            
            add_ft_pt_column(ws, headers)
            #Re-read after headers after column insertion to avoid misalignment bugs 
            headers = find_headers(ws)
            

            individuals = build_individuals(
                ws=ws,
                headers=headers,
                file_name=excel_file.name,
                sheet_name=sheet_name,
                data_start_row=7 #fixed rows for where code starts comiling individual data
            )

            for individual in individuals:
                errors = validate_individual(individual)

                for err in errors:
                    ws.cell(
                        row=err["row"],
                        column=err["column"]
                    ).fill = YELLOW_FILL

                    error_log.append(err)

        output_file = VALIDATED_FOLDER / f"{excel_file.stem}_validated.xlsx" #updates the validated_folder does not create new ones
        wb.save(output_file)

    return error_log


# =========================
# ERROR REPORT
# =========================
def write_error_report(errors):
    if not errors:
        print("No errors found.")
        return

    df = pd.DataFrame(errors)
    df.to_excel(ERROR_REPORT_PATH, index=False)
    print(f"Error report written to {ERROR_REPORT_PATH}")


# =========================
# ENTRY POINT
# =========================
if __name__ == "__main__":
    errors = process_files()
    write_error_report(errors)
