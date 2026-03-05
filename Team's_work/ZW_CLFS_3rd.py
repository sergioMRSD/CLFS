#!/usr/bin/env python3
"""Unified validator script that runs all checks and produces a highlighted Excel.

Usage: run from repo root. It will read the uploaded CSV named
`CLFS_sample_input_validated.xlsx - Sheet1.csv` and write outputs to ./out
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


CSV_PATH = "CLFS_sample_input_validated.xlsx - Sheet1.csv"
OUT_DIR = Path("out")


def safe_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    s = str(v).strip()
    if s == "":
        return None
    # remove common separators/currency markers
    s = s.replace(",", "").replace("$", "").replace("\u200b", "")
    # attempt to extract leading numeric portion
    try:
        return float(s)
    except Exception:
        # try to find digits in string
        import re

        m = re.search(r"-?\d+[\d\.]*", s)
        if m:
            try:
                return float(m.group(0))
            except Exception:
                return None
    return None


def safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def map_ssoc_to_group(val: Any) -> Optional[int]:
    s = safe_str(val)
    if s == "":
        return None
    # find first digit 1-9
    import re

    m = re.search(r"([1-9])", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def find_column(cols: List[str], candidates: List[str]) -> Optional[str]:
    low = [c.lower() for c in cols]
    for cand in candidates:
        candl = cand.lower()
        for i, h in enumerate(low):
            if candl in h:
                return cols[i]
    return None


def find_interest_columns(cols: List[str]) -> List[str]:
    kws = [
        "how much interest did you receive from savings",
        "please indicate the amount you received from interest from savings",
        "how much dividends and interests did you receive",
        "interest",
        "dividend",
    ]
    found = []
    low = [c.lower() for c in cols]
    for pattern in kws:
        for i, h in enumerate(low):
            if pattern in h and cols[i] not in found:
                found.append(cols[i])
                break
    # fallback: include any columns containing 'interest' or 'dividend'
    for i, h in enumerate(low):
        if ("interest" in h or "dividend" in h) and cols[i] not in found:
            found.append(cols[i])
    # limit to 3
    return found[:3]


def main():
    OUT_DIR.mkdir(exist_ok=True)
    df = pd.read_csv(CSV_PATH, dtype=object)
    cols = list(df.columns)

    # Determine columns
    hours_col = find_column(cols, ["usual hours of work", "Usual hours of work"] ) or "Usual hours of work"
    ssoc_col = find_column(cols, ["ssoc code", "ssoc"]) or "SSOC Code"
    labour_status_col = find_column(cols, ["labour force status", "labour force status"]) or "Labour Force Status"
    age_col = find_column(cols, ["age", "age "]) or "Age"

    interest_cols = find_interest_columns(cols)

    cik_col = find_column(cols, ["regular cash and in-kind allowances", "cash and in-kind allowances"]) or "How much did you receive from regular cash and in-kind allowances or contributions (including alimony) from children, relatives, friends not staying in this household in the last 12 months"
    oth_col = find_column(cols, ["sources other than employment", "sources other than employment and the above"]) or "How much did you receive from sources other than employment and the above (e.g. regular pension payments, regular annuity payouts (excluding CPF Life, CPF Retirement Sum Scheme), social welfare grants, etc.) in the last 12 months"

    # prepare outputs
    rows = []
    summary = {
        "HW_001": 0,
        "HW_002": 0,
        "HW_003": 0,
        "HW_004": 0,
        "INTR_001": 0,
        "INTR_002": 0,
        "CIK_001": 0,
        "OTH_001": 0,
    }
    skipped = {k: 0 for k in summary}

    for idx, row in df.reset_index(drop=True).iterrows():
        r = row.to_dict()
        triggered = []
        t_cells: List[str] = []

        # HW_001
        h = safe_number(r.get(hours_col))
        if h is None:
            skipped["HW_001"] += 1
        else:
            if h > 99:
                triggered.append("HW_001")
                t_cells.append(hours_col)
                summary["HW_001"] += 1

        # SSOC group rules
        ss = r.get(ssoc_col)
        grp = map_ssoc_to_group(ss)
        if grp is None:
            skipped["HW_002"] += 1
            skipped["HW_003"] += 1
        else:
            if h is not None:
                if grp in {1,2,3}:
                    if h <= 10 or h >= 50:
                        triggered.append("HW_002")
                        if hours_col not in t_cells:
                            t_cells.append(hours_col)
                        summary["HW_002"] += 1
                elif 4 <= grp <= 9:
                    if h <= 10 or h >= 25:
                        triggered.append("HW_003")
                        if hours_col not in t_cells:
                            t_cells.append(hours_col)
                        summary["HW_003"] += 1
            else:
                # hours missing
                skipped["HW_002"] += 1
                skipped["HW_003"] += 1

        # HW_004 student check
        lab = safe_str(r.get(labour_status_col)).lower()
        is_student = False
        if lab:
            if "stud" in lab or "student" in lab or "studying" in lab:
                is_student = True
        if is_student:
            if h is None:
                skipped["HW_004"] += 1
            else:
                if h > 40:
                    triggered.append("HW_004")
                    if hours_col not in t_cells:
                        t_cells.append(hours_col)
                    summary["HW_004"] += 1

        # Interest rules - per-column
        age_v = safe_number(r.get(age_col))
        if not interest_cols:
            skipped["INTR_001"] += 1
            skipped["INTR_002"] += 1
        else:
            if age_v is None:
                skipped["INTR_001"] += 1
                skipped["INTR_002"] += 1
            else:
                for ic in interest_cols:
                    v = safe_number(r.get(ic))
                    if v is None:
                        continue
                    if age_v < 18 and v >= 10000:
                        if "INTR_001" not in triggered:
                            triggered.append("INTR_001")
                        if ic not in t_cells:
                            t_cells.append(ic)
                        summary["INTR_001"] += 1
                    if age_v >= 18 and v >= 600000:
                        if "INTR_002" not in triggered:
                            triggered.append("INTR_002")
                        if ic not in t_cells:
                            t_cells.append(ic)
                        summary["INTR_002"] += 1

        # CIK
        cik_v = safe_number(r.get(cik_col))
        if cik_v is None:
            skipped["CIK_001"] += 1
        else:
            if cik_v >= 24000:
                triggered.append("CIK_001")
                if cik_col not in t_cells:
                    t_cells.append(cik_col)
                summary["CIK_001"] += 1

        # Other sources
        oth_v = safe_number(r.get(oth_col))
        if oth_v is None:
            skipped["OTH_001"] += 1
        else:
            if oth_v >= 19000:
                triggered.append("OTH_001")
                if oth_col not in t_cells:
                    t_cells.append(oth_col)
                summary["OTH_001"] += 1

        rows.append({
            "row_id": idx,
            "triggered_rules": ";".join(triggered),
            "triggered_cells": t_cells,
        })

    out_rows = pd.DataFrame(rows)
    out_rows.to_csv(OUT_DIR / "validation_results.csv", index=False)

    # write summary with flagged and skipped counts
    summary_report = {}
    for k in summary:
        summary_report[k] = {"flagged": summary[k], "skipped": skipped.get(k, 0)}
    with open(OUT_DIR / "validation_summary.json", "w", encoding="utf-8") as fh:
        json.dump(summary_report, fh, indent=2)

    # write highlighted excel
    excel_path = OUT_DIR / "validation_highlighted.xlsx"
    # write df to excel first
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="data", index=False)

    wb = load_workbook(excel_path)
    ws = wb["data"]
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    header = list(df.columns)
    col_index = {c: i + 1 for i, c in enumerate(header)}

    for _, r in out_rows.iterrows():
        rid = int(r["row_id"]) + 2
        t_cells = r["triggered_cells"]
        if not isinstance(t_cells, list):
            # sometimes pandas writes lists as strings; try to parse
            try:
                t_cells = json.loads(t_cells)
            except Exception:
                t_cells = []
        for col in t_cells:
            if col in col_index:
                cell = ws.cell(row=rid, column=col_index[col])
                cell.fill = yellow

    wb.save(excel_path)

    print("Done. Outputs written to:")
    print(" -", OUT_DIR / "validation_results.csv")
    print(" -", OUT_DIR / "validation_summary.json")
    print(" -", excel_path)


if __name__ == "__main__":
    main()
