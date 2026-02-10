#!/usr/bin/env python3
"""
CLFS pipeline: map SSEC codes, validate qualification vs place, and clean "None of the above" multi-select.

Usage:
  python clfs_pipeline.py --input CLFS_sample_input.csv --ssec "Classification...xlsx" --rules CLFS_rules_and_routing.json

Produces:
 - CLFS_sample_input_processed.csv
 - CLFS_sample_input_processed.xlsx (highlights: yellow for flags/corrections)

This script runs the three existing steps in order without changing their logic.
"""

from pathlib import Path
import argparse
import json
import re
import sys
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# --- Utilities copied/adapted from existing scripts (kept simple) ---
def detect_csv_header(csv_path: Path, needle='highest academic qualification', max_lines=200):
    header_row = None
    with open(csv_path, 'r', encoding='utf-8') as fh:
        for i, line in enumerate(fh):
            if needle in line.lower():
                header_row = i
                break
            if i >= max_lines:
                break
    return header_row


def locate_target_column(columns):
    col_map_local = {str(c).lower(): c for c in columns}
    for k in col_map_local:
        if k.strip() == 'highest academic qualification':
            return col_map_local[k]
    for k in col_map_local:
        if 'highest' in k and 'qualification' in k:
            return col_map_local[k]
    return None


def normalize_text(s: str) -> str:
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[\-_/\\(),.;:]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def split_multi_select(value):
    if pd.isna(value):
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    s = str(value)
    parts = re.split(r'[\|;\n\r,\\/]+', s)
    parts = [p.strip() for p in parts if p.strip()]
    return parts


# --- SSEC parsing and matching (adapted from map_ssec.py) ---
def load_ssec_candidates(ssec_path: Path):
    candidates = []
    xls = pd.ExcelFile(ssec_path)
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(ssec_path, sheet_name=sheet_name, header=None, dtype=str)
        except Exception:
            continue
        if df.shape[1] < 2:
            continue
        for idx in range(len(df)):
            code = df.iat[idx, 0]
            desc = df.iat[idx, 1]
            if pd.isna(code) or pd.isna(desc):
                continue
            code_s = str(code).strip()
            desc_s = str(desc).strip()
            if re.match(r'^[0-9A-Za-z]{1,3}$', code_s):
                if code_s.lower().startswith('ssec'):
                    continue
                candidates.append((code_s, desc_s))

    if not candidates:
        xls_all = pd.read_excel(ssec_path, sheet_name=None, dtype=str)
        for sheet_name, df in xls_all.items():
            if df.empty:
                continue
            cols = [str(c).lower() for c in df.columns]
            code_col = None
            desc_col = None
            for i, col in enumerate(cols):
                if 'code' in col or ('ssec' in col and 'code' in col):
                    code_col = df.columns[i]
                if any(x in col for x in ('description', 'title', 'qualification', 'label', 'class')):
                    desc_col = df.columns[i]
            if code_col is not None and desc_col is not None:
                for idx, row in df[[code_col, desc_col]].iterrows():
                    code = row[code_col]
                    desc = row[desc_col]
                    if pd.isna(code) and pd.isna(desc):
                        continue
                    candidates.append((str(code).strip(), str(desc).strip()))
                continue
            # heuristic by proximity
            possible_cols = list(df.columns)
            found = False
            for i in range(len(possible_cols)):
                for j in range(len(possible_cols)):
                    if i == j:
                        continue
                    col_a = possible_cols[i]
                    col_b = possible_cols[j]
                    sample_a = [str(x) for x in df[col_a].dropna().astype(str).head(10).tolist()]
                    sample_b = [str(x) for x in df[col_b].dropna().astype(str).head(10).tolist()]
                    if not sample_a or not sample_b:
                        continue
                    sum_len_a = sum(len(x) for x in sample_a)
                    avg_len_a = (sum_len_a / len(sample_a)) if sample_a else 0
                    sum_len_b = sum(len(x) for x in sample_b)
                    avg_len_b = (sum_len_b / len(sample_b)) if sample_b else 0
                    if avg_len_a <= 6 and avg_len_b >= 10:
                        for idx, row in df[[col_a, col_b]].iterrows():
                            code = row[col_a]
                            desc = row[col_b]
                            if pd.isna(code) and pd.isna(desc):
                                continue
                            candidates.append((str(code).strip(), str(desc).strip()))
                        found = True
                        break
                if found:
                    break

    # dedupe
    seen = set()
    normalized = []
    for code, desc in candidates:
        key = (code, desc)
        if key in seen:
            continue
        seen.add(key)
        normalized.append((code, desc))
    return normalized


def best_match(qualification: str, candidates, threshold=85):
    qn = normalize_text(qualification)
    if not qn:
        return None, 0
    best = (None, 0.0)
    for code, desc in candidates:
        dn = normalize_text(desc)
        if not dn:
            continue
        if dn in qn or qn in dn:
            return code, 100
        # fallback to difflib
        import difflib
        score = int(difflib.SequenceMatcher(None, qn, dn).ratio() * 100)
        if score > best[1]:
            best = (code, score)
    if best[1] >= threshold:
        return best[0], best[1]
    return None, best[1]


# --- Validation logic adapted from validate_qualification_place.py ---
def classify_qualification(q: str):
    ql = q.lower()
    if any(x in ql for x in ['doctor', 'doctoral', 'phd']):
        return 'doctorate'
    if 'master' in ql and 'degree' in ql or re.search(r"\bmaster\b", ql):
        return 'master'
    if any(x in ql for x in ['bachelor', 'first degree', 'degree', 'undergraduate']):
        return 'degree'
    if 'postgraduate diploma' in ql or ('postgraduate' in ql and 'diploma' in ql):
        return 'postgrad_diploma'
    if 'polytechnic' in ql or (re.search(r'\bdiploma\b', ql) and 'postgraduate' not in ql):
        return 'diploma'
    if any(x in ql for x in ['nitec', 'ite', 'wsq', 'vocational', 'certificate', 'skills certificate']):
        return 'vocational'
    if any(x in ql for x in ['gce a', 'a level', 'psle', 'primary', 'secondary', 'o level', 'n level']):
        return 'school'
    if 'polytechnic' in ql:
        return 'diploma'
    return 'unknown'


def classify_place(p: str):
    pl = p.lower()
    if 'outside' in pl:
        return 'outside'
    if any(x in pl for x in ['university', 'nus', 'ntu', 'smu', 'nanyang']):
        return 'university'
    if 'polytechnic' in pl or 'poly' in pl:
        return 'polytechnic'
    if any(x in pl for x in ['ite', 'nitec']):
        return 'ite'
    if 'local polytechnics' in pl:
        return 'polytechnic'
    return 'other'


def should_flag(qual_cat, place_cat, qual_text, place_text):
    reasons = []
    q = qual_cat
    p = place_cat
    qt = qual_text.lower()
    pt = place_text.lower()
    if q in ('diploma', 'vocational') and p == 'university':
        reasons.append('Diploma/vocational qualification reported from a university')
    if q in ('degree', 'postgrad_diploma', 'master', 'doctorate') and p in ('polytechnic', 'ite'):
        reasons.append('Degree or postgraduate qualification reported from polytechnic/ITE')
    if q == 'school' and p == 'university':
        reasons.append('Pre-university/school qualification reported from a university')
    if q == 'doctorate' and p in ('polytechnic', 'ite'):
        reasons.append('Doctoral degree reported from polytechnic/ITE')
    if p == 'outside' and any(x in qt for x in ['polytechnic', 'nitec', 'wsq', 'local polytechnic']):
        reasons.append('Singapore-specific qualification reported as obtained Outside of Singapore')
    if 'polytechnic' in qt and p == 'university':
        if 'diploma' in qual_text.lower():
            reasons.append('Polytechnic diploma reported from a university')
    return reasons


# --- Main pipeline ---
def main():
    parser = argparse.ArgumentParser(description='CLFS pipeline: map SSEC, validate qualification/place, clean None of the above')
    parser.add_argument('--input', '-i', required=True, help='Input CSV path')
    parser.add_argument('--ssec', '-s', required=True, help='SSEC classification workbook (.xlsx)')
    parser.add_argument('--rules', '-r', required=True, help='Rules JSON (CLFS_rules_and_routing.json)')
    parser.add_argument('--threshold', '-t', type=int, default=85, help='Fuzzy match threshold')

    args = parser.parse_args()
    repo = Path.cwd()
    input_path = Path(args.input)
    ssec_path = Path(args.ssec)
    rules_path = Path(args.rules)

    if not input_path.exists():
        print('Input CSV not found:', input_path)
        sys.exit(2)
    if not ssec_path.exists():
        print('SSEC workbook not found:', ssec_path)
        sys.exit(2)
    if not rules_path.exists():
        print('Rules JSON not found:', rules_path)
        sys.exit(2)

    # Detect header row and read CSV
    header_row = detect_csv_header(input_path)
    if header_row is None:
        df = pd.read_csv(input_path, dtype=object)
    else:
        df = pd.read_csv(input_path, header=header_row, dtype=object)

    total = len(df)

    # Step A: add SSEC Code column next to Highest Academic Qualification
    target_col = locate_target_column(df.columns)
    if target_col is None:
        print("Could not locate 'Highest Academic Qualification' column in input")
        sys.exit(2)

    # load SSEC candidates
    candidates = load_ssec_candidates(ssec_path)
    if not candidates:
        print('Warning: no SSEC candidates extracted; all rows will be UNMAPPED')

    ssec_col = 'SSEC Code'
    cols = list(df.columns)
    if ssec_col in cols:
        cols.remove(ssec_col)
    insert_at = cols.index(target_col) + 1
    cols.insert(insert_at, ssec_col)
    df = df.reindex(columns=cols)

    ssec_mapped = 0
    ssec_unmapped = 0
    mapped = []
    for qual in df[target_col].astype(object).tolist():
        if pd.isna(qual) or str(qual).strip() == '':
            mapped.append('UNMAPPED')
            ssec_unmapped += 1
            continue
        code, score = best_match(str(qual), candidates, threshold=args.threshold)
        if code is None:
            mapped.append('UNMAPPED')
            ssec_unmapped += 1
        else:
            mapped.append(code)
            ssec_mapped += 1
    df[ssec_col] = mapped

    # Step B: validate qualification vs place
    # Find place column (attempt exact title first)
    place_col = None
    for c in df.columns:
        k = str(c).lower()
        if 'place of study' in k and 'highest' in k:
            place_col = c
            break
    if place_col is None:
        for c in df.columns:
            k = str(c).lower()
            if 'place' in k and 'study' in k and 'highest' in k:
                place_col = c
                break
    if place_col is None:
        print('Could not find place of study column')
        sys.exit(2)

    df['Validation Flag'] = ''
    df['Validation Reason'] = ''
    flags = 0
    highlight_cells = []  # list of (row_idx, col_name)

    for idx, row in df.iterrows():
        qual = '' if pd.isna(row.get(target_col, '')) else str(row.get(target_col, ''))
        place = '' if pd.isna(row.get(place_col, '')) else str(row.get(place_col, ''))
        qual_cat = classify_qualification(qual)
        place_cat = classify_place(place)
        reasons = should_flag(qual_cat, place_cat, qual, place)
        if reasons:
            flags += 1
            df.at[idx, 'Validation Flag'] = 'FLAG FOR REVIEW'
            df.at[idx, 'Validation Reason'] = '; '.join(reasons)
            highlight_cells.append((idx, target_col))
            highlight_cells.append((idx, place_col))

    # Step C: clean "None of the above" multi-select
    with open(rules_path, 'r', encoding='utf-8') as fh:
        rules = json.load(fh)
    # find question title and options
    needle = 'have you utilised any of the following'
    ms_title = None
    ms_options = []
    for field in rules.get('form', {}).get('form_fields', []):
        title = field.get('title','')
        if title and needle in title.lower():
            ms_title = title
            ms_options = field.get('fieldOptions') or field.get('choices') or []
            break
    # try to detect ms column
    ms_col = None
    if ms_title:
        for c in df.columns:
            if ms_title.strip().lower() in str(c).lower():
                ms_col = c
                break
    if ms_col is None:
        for c in df.columns:
            ck = str(c).lower()
            if 'have you utilised' in ck or ('programmes' in ck and 'upgrade' in ck) or 'programmes/initiatives' in ck:
                ms_col = c
                break

    none_label = 'None of the above'
    for opt in ms_options:
        if opt and 'none of the above' in str(opt).lower():
            none_label = str(opt)
            break

    corrections = 0
    if ms_col is not None:
        for idx, row in df.iterrows():
            val = row.get(ms_col, '')
            parts = split_multi_select(val)
            lower_parts = [p.lower() for p in parts]
            if any(p == 'none of the above' or 'none of the above' in p for p in lower_parts):
                # if other items selected as well, correct
                if len(parts) > 1 or (len(parts) == 1 and parts[0].lower() != none_label.lower() and 'none of the above' in parts[0].lower()):
                    df.at[idx, ms_col] = none_label
                    corrections += 1
                    highlight_cells.append((idx, ms_col))
                    existing = df.at[idx, 'Validation Reason']
                    added = 'None of the above corrected to exact label'
                    if existing:
                        df.at[idx, 'Validation Reason'] = existing + '; ' + added
                    else:
                        df.at[idx, 'Validation Reason'] = added
                    if not df.at[idx, 'Validation Flag']:
                        df.at[idx, 'Validation Flag'] = 'CORRECTION'

    # Write outputs
    out_csv = input_path.parent / f"{input_path.stem}_processed.csv"
    out_xlsx = input_path.parent / f"{input_path.stem}_processed.xlsx"
    df.to_csv(out_csv, index=False)

    # write xlsx then apply highlights
    df.to_excel(out_xlsx, index=False, sheet_name='CLFS_sample_input')
    wb = load_workbook(out_xlsx)
    ws = wb['CLFS_sample_input']
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header = [str(c) for c in df.columns]
    col_idx = {c: i+1 for i, c in enumerate(header)}
    for (row_idx, col_name) in highlight_cells:
        excel_row = row_idx + 2
        excel_col = col_idx.get(col_name)
        if excel_col is None:
            continue
        cell = ws.cell(row=excel_row, column=excel_col)
        cell.fill = yellow
    wb.save(out_xlsx)

    # Summary
    print('Total rows processed:', total)
    print('SSEC mapped:', ssec_mapped)
    print('SSEC unmapped:', ssec_unmapped)
    print('Qualification/place flags:', flags)
    print('"None of the above" corrections:', corrections)
    print('Wrote CSV:', out_csv)
    print('Wrote highlighted XLSX:', out_xlsx)


if __name__ == '__main__':
    main()
