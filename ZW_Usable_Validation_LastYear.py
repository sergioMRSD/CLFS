"""
ZW_Usable_Validation_LastYear.py

Orchestrator that loads CLFS_newformat.csv and answer.json, groups household members,
invokes validators from CLFS_Brain.py, applies corrections, writes a _validated.xlsx with
highlights and a _report.xlsx summary.

Usage:
    python3 ZW_Usable_Validation_LastYear.py   # runs main pipeline (will look for CLFS_newformat.csv and answer.json)

"""
from pathlib import Path
import json
from typing import List, Dict, Any
import pandas as pd  # type: ignore
from openpyxl.styles import PatternFill, Font  # type: ignore
from openpyxl import load_workbook  # type: ignore

from CLFS_Brain import (
    validate_identification_type,
    validate_residential_st,
    validate_h_sep_y,
    validate_activity_status,
    validate_i_l,
    validate_none_of_the_above_exclusive,
    validate_travel_time_format,
    validate_years_in_employment_consistency,
    validate_num_children,
    validate_oaw_income_threshold,
    validate_occupation_details,
    validate_employment_consistency,
    validate_seeking_work_logic,
    validate_duration_numeric,
    ValidationResult,
)

ROOT = Path(__file__).parent
CSV_PATH = ROOT / "CLFS_newformat.csv"
ANSWER_JSON = ROOT / "answer.json"

SKIPROWS = 5

YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def _find_household_id_column(df: pd.DataFrame) -> str:
    # In this CSV layout the first two columns are metadata; start search from column index 2
    cols = list(df.columns)[2:]
    candidates = [c for c in cols if c.lower() in ('household_id', 'hhid', 'household', 'hh_id', 'hhid_str', 'household id')]
    if candidates:
        return str(candidates[0])
    # fallback: pick the first non-metadata column if available, otherwise first column
    return str(cols[0]) if cols else str(df.columns[0])


def _extract_identification_options(answer_json: Dict[str, Any]) -> List[str]:
    # Heuristic: find any fieldOptions lists that contain 'Singapore Citizen'
    def walk(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k == 'fieldOptions' and isinstance(v, list):
                    yield v
                else:
                    yield from walk(v)
        elif isinstance(obj, list):
            for item in obj:
                yield from walk(item)
    for options in walk(answer_json):
        try:
            if any('Singapore Citizen' in str(opt) for opt in options):
                return [str(o) for o in options]
        except Exception:
            continue
    # fallback empty
    return []


def run_pipeline(csv_path: Path = CSV_PATH, answer_json_path: Path = ANSWER_JSON):
    if not csv_path.exists():
        print(f"Input CSV not found: {csv_path}")
        return
    # Read CSV using the real header row (row 6) and skip the top metadata rows
    # Try tab-separated first (common for exported surveys), then fallback to python engine
    try:
        df = pd.read_csv(csv_path, sep='\t', skiprows=SKIPROWS)
    except Exception as e:
        print(f"TSV parse failed: {e}; attempting fallback parsing with python engine and sep='\t' (skiprows=5)")
        try:
            df = pd.read_csv(csv_path, engine='python', sep='\t', skiprows=5)
        except Exception as e2:
            print(f"Fallback TSV parse failed: {e2}; attempting generic sep=None parse")
            try:
                df = pd.read_csv(csv_path, engine='python', sep=None, skiprows=5)
            except Exception as e3:
                print(f"Fallback generic parse failed: {e3}")
                return
    print(f"Loaded {len(df)} rows from {csv_path.name}")

    hh_col = _find_household_id_column(df)
    print(f"Using household id column: {hh_col}")

    # Load answer.json for identification options
    options = []
    if answer_json_path.exists():
        with open(answer_json_path, 'r', encoding='utf-8') as f:
            aj = json.load(f)
            options = _extract_identification_options(aj)
    else:
        print("Warning: answer.json not found; identification validation will be less strict")

    # Prepare book-keeping
    errors = []  # list of dicts: row, col, rule, message
    corrections = []

    # We'll operate on a copy to apply corrections
    mod_df = df.copy()

    # Group by household and iterate
    for hh, grp in df.groupby(hh_col):
        # collect household ethnicities for language rule — heuristic: look for column names containing 'ethnic'
        eth_cols = [c for c in df.columns if 'ethnic' in c.lower()]
        household_ethnicities = []
        if eth_cols:
            household_ethnicities = [str(x) for x in grp[eth_cols].stack().dropna().astype(str).tolist()]

        for idx, row in grp.iterrows():
            # Example candidate columns (best-effort): Identification Type, Where are you currently staying?, Marital Status,
            # Labour Force Status, Are you actively looking for a new job?, PostalCode, TravelTime, TotalYearsEmployed,
            # YearsCurrentJob, Age, AgeStartedEmployment, NumChildren, PrimaryLanguage, EmploymentStatus, MonthlyIncome

            # we try to map a few common column names (case-insensitive)
            def _get(col_candidates):
                # Ignore first two metadata columns when searching for fields
                columns_to_search = list(df.columns)[2:]
                lower_candidates = [x.lower() for x in col_candidates]
                for c in columns_to_search:
                    if c.lower() in lower_candidates:
                        return row[c], c
                return None, None

            identification, identification_col = _get(['Identification Type', 'identification_type', 'id_type'])
            res_st, res_st_col = _get(['Where are you currently staying?', 'where_currently_staying', 'residential_status'])
            marital, marital_col = _get(['Marital Status', 'marital_status'])
            labour, labour_col = _get(['Labour Force Status', 'labour_force_status', 'activity_status'])
            looking, looking_col = _get(['Are you actively looking for a new job?', 'looking_for_job', 'i_l'])
            # postal code check removed per final cleanup
            travel, travel_col = _get(['TravelTime', 'travel_time', 'travel_minutes'])
            job_title, job_title_col = _get(['Job Title', 'job_title'])
            total_years, total_years_col = _get(['TotalYearsEmployed', 'total_years'])
            years_current, years_current_col = _get(['YearsCurrentJob', 'years_current'])
            age, age_col = _get(['Age', 'age'])
            age_started, age_started_col = _get(['At what age did you start employment'])
            num_children, num_children_col = _get(['Number of children given birth to'])
            employment_status, employment_status_col = _get(['Employment Status', 'employment_status'])
            monthly_income, monthly_income_col = _get(['Last drawn GMI'])
            # Next-15 candidate fields (best-effort mapping)
            e_occ, e_occ_col = _get(['What kind of occupation were you looking for?'])
            w_desc, w_desc_col = _get(['Main tasks / duties'])
            emp_flag, emp_flag_col = _get(['_EMP_', 'is_employed', 'employed_flag'])
            e_empst, e_empst_col = _get(['E_EMPST', 'employment_status_detail'])
            empst, empst_col = _get(['_EMPST', 'emp_status'])
            u_l, u_l_col = _get(['U_L', 'actively_looking'])
            u_w, u_w_col = _get(['U_W', 'available_to_start'])
            i_w, i_w_col = _get(['I_W', 'inability_to_work'])
            e_w, e_w_col = _get(['E_W', 'other_availability_flag'])
            duration, duration_col = _get(['_DUR', 'duration', 'duration_years'])

            # Run the validators from CLFS_Brain
            # Survey rules
            vr = validate_identification_type(identification, allowed_options=options)
            id_col_for_report = identification_col or 'Identification Type'
            if not vr.is_valid:
                errors.append({'row': idx, 'col': id_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})
            elif vr.corrected_value is not None and vr.corrected_value != identification:
                target_col = identification_col or 'Identification Type'
                mod_df.at[idx, target_col] = vr.corrected_value
                corrections.append({'row': idx, 'col': target_col, 'from': identification, 'to': vr.corrected_value, 'rule': vr.rule_applied})

            vr = validate_residential_st(res_st)
            res_col_for_report = res_st_col or 'Where are you currently staying?'
            if not vr.is_valid:
                errors.append({'row': idx, 'col': res_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_h_sep_y(marital)
            marital_col_for_report = marital_col or 'Marital Status'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': marital_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_activity_status(labour)
            labour_col_name = labour_col or 'Labour Force Status'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': labour_col_name, 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_i_l(looking)
            looking_col_for_report = looking_col or 'Are you actively looking for a new job?'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': looking_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            # Next-15 validators
            vr = validate_occupation_details(employment_status, e_occ, w_desc)
            occ_col_for_report = job_title_col or e_occ_col or w_desc_col or 'Job Title'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': occ_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_employment_consistency(emp_flag, e_empst, empst, labour)
            emp_cons_col_for_report = emp_flag_col or e_empst_col or empst_col or labour_col or 'Labour Force Status'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': emp_cons_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_seeking_work_logic(u_l, u_w, i_w, e_w)
            seeking_col_for_report = u_l_col or u_w_col or i_w_col or e_w_col or 'Are you actively looking for a new job?'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': seeking_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_duration_numeric(duration)
            duration_col_for_report = duration_col or 'Total Duration'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': duration_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            # Logic rules
            ms_val, ms_col = _get(['SomeMultiSelectQuestion','multi_select'])
            vr = validate_none_of_the_above_exclusive(ms_val)
            ms_col_for_report = ms_col or 'SomeMultiSelectQuestion'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': ms_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            # postal code validation removed

            vr = validate_travel_time_format(travel)
            travel_col_for_report = travel_col or 'TravelTime'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': travel_col_for_report, 'rule': vr.rule_applied, 'message': vr.message})

            # Cross-field rules
            vr = validate_years_in_employment_consistency(total_years, years_current, age, age_started)
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': 'EmploymentYears', 'rule': vr.rule_applied, 'message': vr.message})

            vr = validate_num_children(num_children, age)
            num_children_col_name = num_children_col or 'Number of children given birth to'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': num_children_col_name, 'rule': vr.rule_applied, 'message': vr.message})

            # language-vs-ethnicity validation removed

            # Financial rules
            vr = validate_oaw_income_threshold(employment_status, monthly_income)
            monthly_income_col_name = monthly_income_col or 'Last drawn GMI'
            if vr and not vr.is_valid:
                errors.append({'row': idx, 'col': monthly_income_col_name, 'rule': vr.rule_applied, 'message': vr.message})

    # Write validated dataframe to Excel
    out_validated = ROOT / (csv_path.stem + "_validated.xlsx")
    mod_df.to_excel(out_validated, index=False)

    # Apply highlights for errors and corrections in the validated workbook
    wb = load_workbook(out_validated)
    ws = wb.active
    if ws is None:
        print(f"Warning: workbook has no active sheet; skipping highlights for {out_validated.name}")
    else:
        # Map dataframe column name to excel col index
        col_to_idx = {c: i + 1 for i, c in enumerate(mod_df.columns)}

        # Yellow highlight for any cell that failed a validation rule
        for e in errors:
            excel_row = int(e['row']) + 2  # pandas index -> excel row (header row + 1)
            c = col_to_idx.get(e['col'], None)
            if c is None:
                continue
            cell = ws.cell(row=excel_row, column=c)
            cell.fill = YELLOW_FILL

        # Red text for any auto-corrected value (mod_df differs from original df)
        red_font = Font(color='FFFF0000')
        for idx in mod_df.index:
            for col in mod_df.columns:
                if col not in df.columns:
                    continue
                try:
                    orig_val = df.at[idx, col]
                except Exception:
                    orig_val = None
                new_val = mod_df.at[idx, col]
                if orig_val is not None and str(orig_val) != str(new_val):
                    excel_row = int(idx) + 2
                    c = col_to_idx.get(col, None)
                    if c is None:
                        continue
                    ws.cell(row=excel_row, column=c).font = red_font

        wb.save(out_validated)
    print(f"Wrote validated workbook: {out_validated.name}")

    # Build the structured validation report workbook with three sheets: Summary, Details, Complete Dataset
    import collections

    # Summary: Rule ID, Column Name, Error Message, Count
    summary_counts = collections.Counter([(e.get('rule'), e.get('col'), e.get('message')) for e in errors if e.get('rule')])
    df_summary = pd.DataFrame([
        {'Rule ID': r, 'Column Name': c, 'Error Message': m, 'Count': cnt}
        for (r, c, m), cnt in summary_counts.items()
    ])

    # Helper to find response ID and member name columns heuristically
    def _find_col(dfobj, keywords):
        for k in keywords:
            for col in dfobj.columns:
                if k in col.lower():
                    return col
        return None

    resp_col = _find_col(df, ['responseid', 'response id', 'response_id', 'respid', 'resp_id', 'response', 'resp'])
    name_col = _find_col(df, ['membername', 'member name', 'name', 'full name', 'fullname'])

    # Details: File Name, Row Number (original CSV), Response ID, Member Name, Rule ID, Column Name, Error Message
    details_rows = []
    for e in errors:
        idx = int(e['row'])
        original_csv_row = idx + SKIPROWS + 2  # account for header and skiprows so original CSV row matches
        resp_val = df.at[idx, resp_col] if resp_col and resp_col in df.columns else ''
        name_val = df.at[idx, name_col] if name_col and name_col in df.columns else ''
        details_rows.append({
            'File Name': csv_path.name,
            'Row Number': original_csv_row,
            'Response ID': resp_val,
            'Member Name': name_val,
            'Rule ID': e.get('rule'),
            'Column Name': e.get('col'),
            'Error Message': e.get('message'),
        })

    df_details = pd.DataFrame(details_rows)

    out_report = ROOT / (csv_path.stem + "_validation_report.xlsx")
    with pd.ExcelWriter(out_report) as writer:
        # ensure sheets are created in requested order
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_details.to_excel(writer, sheet_name='Details', index=False)
        df.to_excel(writer, sheet_name='Complete Dataset', index=False)

    print(f"Wrote report workbook: {out_report.name}")


if __name__ == '__main__':
    run_pipeline()
