#!/usr/bin/env python3
"""Apply corrections from validation report to the 'Complete Dataset' column

Usage:
    python scripts/apply_report_corrections.py [input_csv] [output_xlsx]

Defaults:
    input_csv: output/CLFS_contextually_wrong_answers_validation_report.csv
    output_xlsx: output/CLFS_contextually_wrong_answers_validation_applied.xlsx

The script will:
 - Load the CSV as pandas DataFrame
 - For each row, if 'corrections', 'corrections_2', or 'corrections_3' has a non-empty value,
   overwrite the 'Complete Dataset' cell with the first non-empty correction (priority: corrections, corrections_2, corrections_3)
 - Save the final table to Excel
 - Highlight the changed cells in the 'Complete Dataset' column with an orange fill (FFA500)
"""
from pathlib import Path
import sys
try:
    import pandas as pd  # type: ignore[import]
except Exception:  # pragma: no cover - runtime will warn and exit if missing
    pd = None

try:
    from openpyxl import load_workbook  # type: ignore[import]
    from openpyxl.styles import PatternFill  # type: ignore[import]
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]
except Exception:  # pragma: no cover
    load_workbook = None
    PatternFill = None
    Worksheet = None

from typing import Optional, List, cast


def _scalar_to_str(val) -> str:
    """Convert a pandas/numpy scalar or Python value to a safe string for comparison.

    Returns empty string for NA/None.
    """
    try:
        if val is None:
            return ""
        # pandas NA / numpy nan (only call if pandas imported)
        if pd is not None:
            try:
                if pd.isna(val):
                    return ""
            except Exception:
                pass
    except Exception:
        pass

    # numpy/pandas scalar
    try:
        if hasattr(val, "item"):
            try:
                return str(val.item())
            except Exception:
                pass
    except Exception:
        pass

    # Sequence-like objects (Series, Index, list, tuple) - try to grab first element
    try:
        if hasattr(val, '__getitem__') and not isinstance(val, (str, bytes)):
            try:
                return str(val[0])
            except Exception:
                return str(val)
    except Exception:
        pass

    return str(val)


def main(argv: Optional[List[str]] = None) -> int:
    argv = argv or sys.argv[1:]

    input_csv = Path(argv[0]) if len(argv) >= 1 else Path("output") / "CLFS_contextually_wrong_answers_validation_report.csv"
    output_xlsx = Path(argv[1]) if len(argv) >= 2 else Path("output") / "CLFS_contextually_wrong_answers_validation_applied.xlsx"

    if not input_csv.exists():
        print(f"Error: input CSV not found: {input_csv}")
        return 2

    # Ensure runtime deps are available
    if pd is None:
        print("Error: pandas is not installed. Please install pandas (e.g., pip install pandas openpyxl) and retry.")
        return 10
    if load_workbook is None:
        print("Error: openpyxl is not installed. Please install openpyxl (e.g., pip install openpyxl) and retry.")
        return 11

    # Read input (CSV or Excel). Try to be flexible with sheets.
    try:
        if input_csv.suffix.lower() in {'.xlsx', '.xls', '.xlsm'}:
            # Try primary sheet first
            try:
                df = pd.read_excel(input_csv, sheet_name=0)
            except Exception:
                # Fall back to reading the 'Details' sheet if present
                try:
                    df = pd.read_excel(input_csv, sheet_name='Details')
                except Exception:
                    # As last resort, try the 'Complete Dataset' sheet
                    df = pd.read_excel(input_csv, sheet_name='Complete Dataset')
        else:
            df = pd.read_csv(input_csv, encoding="utf-8-sig")
    except Exception as e:
        print(f"Error reading input {input_csv}: {e}")
        return 3

    # If the input was Excel with separate 'Details' and 'Complete Dataset' sheets,
    # prefer to operate by applying corrections from the Details sheet into the
    # Complete Dataset sheet. Otherwise, if the loaded df contains a
    # 'Complete Dataset' column, treat it as the working table.
    is_excel = input_csv.suffix.lower() in {'.xlsx', '.xls', '.xlsm'}

    if is_excel:
        # load both sheets explicitly
        xls = pd.ExcelFile(input_csv)
        if 'Details' not in xls.sheet_names or 'Complete Dataset' not in xls.sheet_names:
            print("Error: Excel input missing required sheets 'Details' and/or 'Complete Dataset'")
            return 4

        details_df = pd.read_excel(input_csv, sheet_name='Details')
        complete_df = pd.read_excel(input_csv, sheet_name='Complete Dataset')

        # Determine correction columns available in details
        correction_cols = [c for c in ['corrections', 'corrections_2', 'corrections_3'] if c in details_df.columns]
        if not correction_cols:
            print("No correction columns found in 'Details' sheet. Nothing to apply.")
            # Still write a copy of the workbook
            out = output_xlsx
            # copy original workbook
            from shutil import copyfile
            copyfile(input_csv, out)
            print(f"Wrote copy of original workbook to: {out}")
            return 0

        changes = []  # list of (excel_row_idx (1-based), col_name)
        audit_records: list[dict] = []

        for _, drow in details_df.iterrows():
            # Each details row includes a 'row' (1-based) and one or more 'column' fields
            try:
                raw_row = drow.get('row')
                if raw_row is None:
                    continue
                raw_row_s = str(raw_row).strip()
                if not raw_row_s:
                    continue
                try:
                    target_row = int(float(raw_row_s)) - 1
                except Exception:
                    continue
            except Exception:
                continue
            if target_row < 0 or target_row >= len(complete_df):
                continue

            # Build list of target column names: column, column_2, column_3...
            target_columns = []
            if 'column' in details_df.columns and pd.notna(drow.get('column')):
                # split on '&' if multiple names combined
                base_cols = [c.strip() for c in str(drow.get('column')).split('&') if c.strip()]
                target_columns.extend(base_cols)
            # Explicit numbered column_n fields
            i = 2
            while f'column_{i-1}' in details_df.columns or f'column_{i}' in details_df.columns:
                key = f'column_{i-1}' if f'column_{i-1}' in details_df.columns else f'column_{i}'
                if key in details_df.columns and pd.notna(drow.get(key)):
                    target_columns.append(str(drow.get(key)).strip())
                i += 1

            # Now apply corrections in order
            for idx_c, corr_col in enumerate(correction_cols):
                corr_val = drow.get(corr_col)
                if pd.isna(corr_val):
                    continue
                corr_str = str(corr_val).strip()
                if not corr_str:
                    continue

                # Find the matching target column for this correction
                col_name = None
                if idx_c < len(target_columns):
                    col_name = target_columns[idx_c]
                else:
                    # Fallback: use the base 'column' name (first) if present
                    if target_columns:
                        col_name = target_columns[0]

                if not col_name:
                    continue

                # If column name includes '&', pick first (already split above)
                col_name = col_name.strip()

                # Only apply if the column exists in complete_df
                if col_name not in complete_df.columns:
                    # try to find a close match (case-insensitive)
                    match = next((c for c in complete_df.columns if str(c).strip().lower() == col_name.lower()), None)
                    if match:
                        col_name = match
                    else:
                        continue

                # Ensure we get a single integer index for the column (handles duplicate columns)
                try:
                    loc = int(list(complete_df.columns).index(col_name))
                except Exception:
                    # fallback to get_loc
                    loc = complete_df.columns.get_loc(col_name)

                col_label = complete_df.columns[loc]
                # Ensure column is object dtype to allow string replacements without dtype warnings
                try:
                    complete_df[col_label] = complete_df[col_label].astype(object)
                except Exception:
                    # If astype fails, ignore and proceed
                    pass

                # Fetch the old value as a scalar using label-based access
                try:
                    old_val = complete_df.at[target_row, col_label]
                except Exception:
                    old_val = complete_df.iloc[target_row, loc]
                old_str = _scalar_to_str(old_val)
                if old_str != corr_str:
                    # Assign using .at (label-based) for scalar set
                    try:
                        complete_df.at[target_row, col_label] = corr_str
                    except Exception:
                        complete_df.iloc[target_row, loc] = corr_str
                    changes.append((target_row, col_name))
                    # record audit
                    audit_records.append({
                        "excel_row": target_row + 2,
                        "df_index": target_row,
                        "column": col_name,
                        "old_value": old_val,
                        "new_value": corr_str,
                    })

        # Write back into a copy of the original workbook and highlight changes
        wb = load_workbook(input_csv)
        ws = wb['Complete Dataset']
        assert ws is not None

        # Build header -> column index mapping for the sheet
        if ws is None:
            print("Error: worksheet 'Complete Dataset' not found")
            return 6
        # Use values_only to robustly read header row as plain values
        try:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except Exception:
            print("Error reading header row from 'Complete Dataset' sheet")
            return 7
        header_cells = [str(v).strip() if v is not None else "" for v in first_row]
        header_map = {name: idx + 1 for idx, name in enumerate(header_cells) if name}

        if PatternFill is not None:
            orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        else:
            orange = None

        for (trow, col_name) in changes:
            excel_row = trow + 2
            col_idx = header_map.get(col_name)
            if col_idx is None:
                # try case-insensitive match
                for k, v in header_map.items():
                    if k.lower() == col_name.lower():
                        col_idx = v
                        break
            if col_idx is None:
                continue
            assert ws is not None
            cell = ws.cell(row=excel_row, column=col_idx)
            # use a scalar value from complete_df
            try:
                val = complete_df.iat[trow, list(complete_df.columns).index(col_name)]
            except Exception:
                try:
                    val = complete_df.iloc[trow, complete_df.columns.get_loc(col_name)]
                except Exception:
                    val = None
            # normalize numpy / pandas scalar types to native Python scalars for openpyxl
            try:
                if val is None:
                    pass
                elif isinstance(val, (str, bytes, int, float, bool)):
                    pass
                else:
                    # try numpy scalar
                    if hasattr(val, 'item'):
                        try:
                            val = val.item()
                        except Exception:
                            pass
                    # try pandas Series/Index
                    if hasattr(val, 'iloc') and not isinstance(val, (str, bytes)):
                        try:
                            val = val.iloc[0]
                        except Exception:
                            val = str(val)
            except Exception:
                # as a last resort, stringify
                try:
                    val = str(val)
                except Exception:
                    val = None
            cell.value = val  # type: ignore[assignment]
            if orange is not None:
                cell.fill = orange

        out_path = output_xlsx
        wb.save(out_path)
        wb.close()

        # write audit CSV
        try:
            audit_path = Path("output") / "applied_corrections_audit.csv"
            if audit_records:
                pd.DataFrame(audit_records).to_csv(audit_path, index=False, encoding="utf-8-sig")
            else:
                # write empty audit with headers
                pd.DataFrame(columns=["excel_row", "df_index", "column", "old_value", "new_value"]).to_csv(audit_path, index=False, encoding="utf-8-sig")
            print(f"Wrote audit CSV to: {audit_path}")
        except Exception:
            print("Warning: failed to write audit CSV")

        print(f"Applied corrections to {len(changes)} cells and saved: {out_path}")
        return 0

    # Non-excel fallback: if df contains 'Complete Dataset' column
    if 'Complete Dataset' not in df.columns:
        print("Error: 'Complete Dataset' column not found in input CSV or sheet")
        return 4

    correction_cols = [c for c in ['corrections', 'corrections_2', 'corrections_3'] if c in df.columns]
    if not correction_cols:
        print("No correction columns found ('corrections', 'corrections_2', 'corrections_3'). Nothing to apply.")
        # Still write out Excel copy
        df.to_excel(output_xlsx, index=False, engine='openpyxl')
        print(f"Wrote output (no changes): {output_xlsx}")
        return 0

    changed_rows = []

    # Iterate rows and apply first available correction to the Complete Dataset column
    for idx, row in df.iterrows():
        new_val = None
        for col in correction_cols:
            try:
                val = row.get(col)
            except Exception:
                val = None
            if pd.isna(val):
                continue
            s = str(val).strip()
            if s:
                new_val = s
                break

        if new_val is not None:
            # Only record change if the value actually differs (avoid false positives)
            old = df.at[idx, 'Complete Dataset']
            old_str = '' if pd.isna(old) else str(old)
            if old_str != new_val:
                df.at[idx, 'Complete Dataset'] = new_val
                changed_rows.append(idx)

    # Save to Excel first
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_xlsx, index=False, engine='openpyxl')

    # If no changes, we're done
    if not changed_rows:
        print("No changes were applied.")
        print(f"Output written to: {output_xlsx}")
        return 0

    # Open workbook and highlight changed cells in 'Complete Dataset' column
    wb = load_workbook(output_xlsx)
    ws = wb.active
    assert ws is not None

    # Find the column index (1-based) for 'Complete Dataset'
    # use values_only row to be robust
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v) if v is not None else "" for v in first_row]
    try:
        col_idx = headers.index('Complete Dataset') + 1
    except ValueError:
        print("Error: 'Complete Dataset' header not found in written Excel file")
        wb.save(output_xlsx)
        wb.close()
        return 5

    if PatternFill is not None:
        orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    else:
        orange = None

    for r in changed_rows:
        excel_row = r + 2  # pandas row 0 => excel row 2 (header in row 1)
        cell = ws.cell(row=excel_row, column=col_idx)
        if orange is not None:
            cell.fill = orange  # type: ignore[assignment]

    wb.save(output_xlsx)
    wb.close()

    print(f"Applied corrections to {len(changed_rows)} rows and saved: {output_xlsx}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
