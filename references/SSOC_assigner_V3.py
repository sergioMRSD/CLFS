# -*- coding: utf-8 -*-
"""
Created on Sept 01 15:18:21 2025

@author: momsgk2
"""

# =========================================
# SSOC matcher � Duties-first, always 5-digit, dynamic scoring with guardrails
# + education-weighting
# + disambiguation for arts/hospitality/gaming + electrician/painter/construction
# + managerial anchoring; title�sector conflict penalty; title-only fallback
# ======== Optimised (caching + precomputes + TF-IDF topK) =====================
# ======== Forced assignment ON by default (no X1000) ==========================
# ======== Batch mode: process ALL Excel files in a folder =====================
# ======== Occupation Group (Annex D) hint support =============================
# ======== [NEW] Detailed Top-5 Scoring Report feature =========================
# =========================================

import os, re, sys, datetime as _dt, argparse, shutil, glob
from typing import List, Dict, Tuple, Optional, Set, Callable
from difflib import SequenceMatcher
from functools import lru_cache

try:
    import pandas as pd
except Exception as e:
    raise RuntimeError("This script needs pandas. Install with: pip install pandas openpyxl") from e

try:
    from rapidfuzz import fuzz
    _HAS_RF = True
except Exception:
    _HAS_RF = False

# Optional fast candidate shortlisting
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    _HAS_SK = True
except Exception:
    _HAS_SK = False

# ====== CONFIG DEFAULTS ======
DEFAULT_DEFINITIONS_FILE = r"C:\Users\MOMSGK2\Desktop\OED\SSOC_coder_from_yeefei\ssoc2024-detailed-definitions.xlsx"

# DEFAULT_JOBS_FILE can be a single Excel file OR a folder (auto-treated as jobs-dir)
DEFAULT_JOBS_FILE = r"C:\Users\MOMSGK2\Desktop\OED\SSOC_coder_from_yeefei\ssoc_assigner\Operating_Table"

# Path to the Industry_Strata.xlsx file
DEFAULT_SSIC_LIST_FILE = r"C:\Users\MOMSGK2\Desktop\OED\SSOC_coder_from_yeefei\SSIC_List_2Sep2025.xlsx"
DEFAULT_SSIC_DEFS_FILE = r"C:\Users\MOMSGK2\Desktop\OED\SSOC_coder_from_yeefei\ssic2025a-detailed-definitions.xlsx"

# Manager and Engineer alternative titles file
DEFAULT_EXPERT_MAP_FILE = r"C:\Users\MOMSGK2\Desktop\OED\SSOC_coder_from_yeefei\Library_of_SSOC_eng_manager.xlsx"

DEFAULT_OUTPUT_FILE      = None

DEFAULT_DEF_SHEET      = None
DEFAULT_DEF_SKIP_ROWS  = 0

DEFAULT_JOBS_SHEET      = None
DEFAULT_JOBS_HEADER_ROW = 5
DEFAULT_TITLE_COL_NAME  = "Main Job Title*"

DEFAULT_DUTIES_COL_NAME = None
DEFAULT_DUTIES_COL_INDEX= 8 # column I

DEFAULT_EDU_COL_NAME    = None  # auto-detect

DEFAULT_TITLE_COL_INDEX = None
DEFAULT_EDU_COL_INDEX   = None

DEFAULT_MIN_SCORE = 5.0    # out of 100
DEFAULT_DEBUG     = False
# ============================

# ---------- normalisation + tokenisation ----------
_PUNCT = re.compile(r"[^\w\s]+", flags=re.UNICODE)

_STOP = {
    # Genuine boilerplate and filler words
    "responsible", "responsibilities", "duty", "duties", "works", "work", "working", 
    "perform", "performing", "including", "include", "includes", "ensure", "ensuring", 
    "handle", "handling", "related", "etc", "various", "tasks", "high", "quality", 
    "according", "specifications", "required", "standards", "time", "company",
    "superiors", "based", "prepare", "preparing", "progress","senior", "junior", "lead", "sr", 
    "jr"
}

_ACTION = {
    "paint","painting","spray","roller","brush",
    "drive","driving","deliver","delivery","vehicle","goods","fetch","transport",
    "draw","drawing","draft","drafting","autocad","cad","blueprint","plan","planning",
    "survey","surveying","quantify","measurement","measurements","boq",
    "install","installation","assemble","assembly","maintain","maintenance","repair","troubleshoot",
    "weld","welding","operate","operation","machinery","commission","commissioning","calibrate","calibration",
    "cook","cooking","bake","baking","butcher","cut","cutting","plaster","tile","glaze",
    "inspect","inspection","test","testing","audit","auditing",
    "design","designing","configure","configuration","monitor","monitoring",
    "teach","teaching","care","caring","nurse","nursing","treat","treatment","diagnose","diagnosis",
    "clean","cleaning","sanitize","sanitise","wash","washing","polish","polishing","pack","packing","pick","picking"
}

'''
@lru_cache(maxsize=500_000)
def _get_stem(word: str) -> str:
    """A simple, lightweight stemmer to handle common English word endings."""
    # Order is important: handle longer suffixes first to avoid errors
    if len(word) > 5 and word.endswith('ing'):
        return word[:-3]
    if len(word) > 4 and word.endswith('er'):
        return word[:-2]
    # Handle plurals, but avoid changing short words like 'is' or 'ss' words
    if len(word) > 3 and word.endswith('s') and not word.endswith('ss'):
        return word[:-1]
    return word
'''

@lru_cache(maxsize=200_000)
def _normalize(t: str) -> str:
    if t is None: return ""
    t = t.lower()
    
    # 2. NOW, handle any specific separators that might have been letters (like '_').
    t = re.sub(r"[_/,\-()&]+", " ", t)
    
    # === CORRECTED ORDER OF OPERATIONS ===
    # 1. First, do the aggressive cleaning to remove ALL non-letter, non-space junk.
    t = re.sub(r'[^a-z\s]', '', t)
 
    # 3. Finally, clean up any extra whitespace that was created.
    t = re.sub(r"\s+", " ", t).strip()
    
    return t


@lru_cache(maxsize=200_000)
def _tokens_cached(t: str) -> tuple:
    toks = [w for w in _normalize(t).split() if w and w not in _STOP and not w.isdigit()]
    return tuple(toks)

def _tokens(t: str) -> List[str]:
    return list(_tokens_cached(t))

@lru_cache(maxsize=200_000)
def _bigrams_from_text(t: str) -> frozenset:
    toks = _tokens_cached(t)
    return frozenset(f"{toks[i]}_{toks[i+1]}" for i in range(len(toks)-1))

def _bigrams(toks: List[str]) -> Set[str]:
    return {"{}_{}".format(toks[i], toks[i+1]) for i in range(len(toks)-1)}

def _diff_ratio(a: str, b: str) -> float:
    a, b = _normalize(a), _normalize(b)
    if _HAS_RF:
        return fuzz.WRatio(a, b) / 100.0
    return SequenceMatcher(None, a, b).ratio()

def _diff_ratio_normed(a_norm: str, b_norm: str) -> float:
    if _HAS_RF:
        return fuzz.WRatio(a_norm, b_norm) / 100.0
    return SequenceMatcher(None, a_norm, b_norm).ratio()

def _overlap_measure_sets(A: Set[str], B: Set[str]) -> Tuple[float, float, int]:
    if not A or not B: return 0.0, 0.0, 0
    inter = A & B
    set_ov = len(inter) / float(min(len(A), len(B))) if min(len(A), len(B)) > 0 else 0.0
    jac    = len(inter) / float(len(A | B)) if len(A | B) > 0 else 0.0
    acts   = len(inter & _ACTION)
    return set_ov, jac, acts

def _overlap_measure(a_toks: List[str], b_toks: List[str]) -> Tuple[float, float, int]:
    A, B = set(a_toks), set(b_toks)
    return _overlap_measure_sets(A, B)

def _bigram_overlap_sets(A: Set[str], B: Set[str]) -> float:
    if not A or not B: return 0.0
    return len(A & B) / float(min(len(A), len(B))) if min(len(A), len(B)) > 0 else 0.0

def _bigram_overlap(a_toks: List[str], b_toks: List[str]) -> float:
    A, B = _bigrams(a_toks), _bigrams(b_toks)
    if not A or not B: return 0.0
    return len(A & B) / float(min(len(A), len(B))) if min(len(A), len(B)) > 0 else 0.0


# ---------- sector anchors ----------
SECTOR_ANCHORS: Dict[str, Set[str]] = {
    "managers":{"manager","managing","director","chief","executive","ceo","coo","cfo","cio","cto","chairman","board",
                "general","governance","policy","strategic","strategy","planning","budget","kpi","targets"},
    "construction":{"construction","building","worksite","site","scaffolding","concrete","formwork","rebar","excavation",
                    "tunnel","pile","piling","draught","draft","autocad","bim","fitout","renovation","tiling","plaster",
                    "glazier","roof","roofer","carpentry","joinery","brick","masonry","drywall","mep","plant","crane",
                    "rigger","architectural","structural","clerk_of_works","quantity","boq","drafter","foreman","superintendent"},
    "mfg":{"manufacturing","production","factory","cnc","machining","lathe","milling","tooling","assembly","assembler","process",
           "commissioning","calibration","maintenance","quality","qa","qc","line","plant","operator"},
    "mfg_electronics":{"semiconductor","wafer","fab","cleanroom","esd","pcb","smt","solder","bonding","die","ic","mems"},
    "mfg_chemical":{"chemical","refinery","petrochemical","reactor","polymer","adhesive","paint","coating","distillation","blending"},
    "mfg_food":{"central","kitchen","slaughter","butcher","baking","brewery","distillery","dairy","confectionery","haccp","halal"},
    "transport":{"driver","driving","deliver","delivery","route","dispatch","passengers","taxi","private","hire","van","lorry",
                 "truck","trailer","prime","mover","bus","train","rail","mrt","station","cargo","freight"},
    "logistics":{"logistics","warehouse","storekeeper","storeman","inventory","stock","picking","packing","forklift","reachtruck",
                 "yard","dc","hub","port","quay","container","manifest","awb","bonded"},
    "aviation_marine":{"pilot","aircraft","airline","cabin","steward","airport","runway","marine","ship","vessel","deckhand","tug"},
    "retail":{"retail","shop","store","outlet","cashier","pos","merchandising","category","showroom","sales","salesperson",
              "sales assistant","shop assistant","customer","service","counter","boutique","mall","department"},
    "hospitality":{"hotel","housekeeping","linen","guest","room","butler","steward","concierge","banquet","lodging","resort","front office","guest relations"},
    "travel":{"travel","tourism","tour","ticketing","visa","booking","hotel","resort",
              "itinerary","tourist","guest services","attraction","cruise"},
    "fnb":{"restaurant","cafe","catering","kitchen","cook","chef","pastry","baker","barista","bartender","stewarding","menu"},
    "ict":{"software","developer","programmer","devops","application","system","database","data","network","server","cloud",
           "security","cyber","infrastructure","telecom","ai","ml","testing","qa","product manager","architect","api",
           "frontend","backend","fullstack","ui","ux","scrum","agile"},
    "arts_media":{"gallery","museum","curator","artist","designer","graphic","multimedia","animation","photography","broadcast",
                  "radio","television","film","editor","content","journalist","reporter","pr","copywriter","printing","orchestra","choir"},
    "engineering":{"engineer","mechanical","electrical","electronics","civil","chemical","environmental","biomedical",
                   "process","quality","industrial","production","maintenance","commissioning","calibration","design","workshop"},
    "finance":{"accounting","accounts","accountant","audit","auditor","tax","treasury","compliance","risk","bank","banking",
               "loan","credit","underwriting","portfolio","fund","valuation","actuarial","insurer","insurance","payroll","ledger"},
    "legal":{"law","lawyer","legal","solicitor","advocate","counsel","paralegal","court","litigation","regulatory"},
    "public_admin":{"ministry","statutory","board","government","regulatory","policy","planning","licensing","immigration",
                    "customs","public service","civil service"},
    "education":{"school","teacher","teaching","student","classroom","lecturer","trainer","polytechnic","university","tutor","curriculum","syllabus","preschool"},
    "healthcare":{"patient","ward","clinic","hospital","nursing","physician","dentist","pharmacy","therapist","diagnostic",
                  "radiography","laboratory","rehabilitation","paramedic","medical","surgeon","doctor","allied health"},
    "social":{"social","counsellor","counselling","community","youth","family","casework","welfare","outreach","volunteer"},
    "cleaning":{"cleaning","cleaner","janitor","sanitation","disinfection","premises","landscape","conservancy","sweeping","mopping","dishwasher"},
    "security":{"security","guard","patrol","cctv","command centre","incident","auxiliary","police","prison","investigator","lifeguard","fire rescue"},
    "agri":{"farm","agriculture","horticulture","landscape","nursery","aquaculture","poultry","livestock","gardener","tree"},
    "beauty":{"beautician","makeup","manicurist","pedicurist","spa","massage","therapist","hairdresser","barber","cosmetology"},
    "waste_env":{"waste","recycling","material recovery","grease","collection","environmental","sanitarian","hygiene"},
    "arch_design":{"architect","architecture","urban","town","planner","planning","survey","surveyor","cad","bim","draughtsman","draughtsperson","drafter","interior","product design","landscape architect"},
    "electrical":{"electrician","electrical","wiring","switchboard","conduit","cable pulling","testing and commissioning"},
    "fitness_instruction":{"instructor","fitness","coach","gym","studio","personal trainer",
                       "martial arts","yoga","pilates","zumba","training","mentorship"}
}

def _sector_cues_from_text(text: str) -> Set[str]:
    toks_list = _tokens(text)
    toks_set  = set(toks_list)
    lower     = _normalize(text)
    cues: Set[str] = set()
    for sector, anchors in SECTOR_ANCHORS.items():
        for a in anchors:
            a_norm = _normalize(a.replace("_", " "))
            if " " in a_norm:
                if a_norm in lower:
                    cues.add(sector)
            else:
                if a_norm in toks_set:
                    cues.add(sector)
    return cues

ROLE_ANCHORS = {
    "driver","painter","drafter","draftsman","draftsperson","installer","fitter","welder",
    "cook","chef","butcher","baker","barista","bartender",
    "supervisor","foreman","manager","director","coordinator","engineer","technician",
    "accountant","accounts","payroll","auditor","tax","admin","clerk","assistant","receptionist",
    "teacher","lecturer","nurse","therapist","security","guard","cleaner","housekeeper","steward","butler",
    "salesman","sales","promoter","executive","storekeeper","storeman","principal","doctor","lawyer","architect"
}

def _role_anchor_overlap(query_text: str, candidate_title_and_blob: str) -> int:
    q = set(_tokens(query_text)) & ROLE_ANCHORS
    c = set(_tokens(candidate_title_and_blob)) & ROLE_ANCHORS
    return len(q & c)

# ---------- Excel I/O ----------
def load_expert_map(path: str, debug=False) -> Dict[str, Tuple[str, str]]:
    """
    Loads a curated map of original job titles to their correct SSOC code and title
    by reading both the "Manager" and "Engineer" sheets from the expert file.
    """
    if not path or not os.path.exists(path):
        if debug: print("[INFO] Expert map file not provided or not found. Skipping.")
        return {}
    
    try:
        # === NEW: Load ALL sheets from the Excel file into a dictionary of DataFrames ===
        all_sheets = pd.read_excel(path, engine="openpyxl", sheet_name=None, header=0)
        # ==============================================================================
        
        expert_map = {}
        sheets_to_process = ["Manager", "Engineer"] # Define the specific sheets we care about

        for sheet_name in sheets_to_process:
            if sheet_name not in all_sheets:
                if debug: print(f"[Expert Map] Sheet '{sheet_name}' not found in file. Skipping.")
                continue

            df = all_sheets[sheet_name]
            df.columns = [_norm_cell(c) for c in df.columns]
            
            # Find columns by likely names for this sheet
            orig_title_col = [c for c in df.columns if "original" in c and "title" in c][0]
            ssoc_code_col = [c for c in df.columns if "correct" in c and "ssoc" in c][0]
            ssoc_title_col = [c for c in df.columns if "job" in c and "ssoc" in c][0]

            for _, r in df.iterrows():
                orig_title = str(r.get(orig_title_col, "")).strip()
                ssoc_code = str(r.get(ssoc_code_col, "")).strip()
                ssoc_title = str(r.get(ssoc_title_col, "")).strip()

                if orig_title and ssoc_code:
                    norm_orig_title = _normalize(orig_title)
                    expert_map[norm_orig_title] = (ssoc_code, ssoc_title)
        
        if debug: print(f"[Expert Map] Loaded {len(expert_map)} curated title mappings from {len(all_sheets)} sheet(s).")
        return expert_map
    except Exception as e:
        print(f"[ERROR] Failed to load expert map from {path}: {e}", file=sys.stderr)
        return {}
    
def _strip_punct_lower(s: str) -> str:
    return re.sub(r"[^\w\s]", "", (s or "")).strip().lower()

def _norm_cell(x):
    s = "" if x is None else str(x)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def _canon_headers(row: List[str]) -> List[str]:
    return [_norm_cell(h) for h in row]

def _tok_overlap(a: str, b: str) -> float:
    A, B = set(_tokens(a)), set(_tokens(b))
    if not A or not B: return 0.0
    return len(A & B) / float(max(1, min(len(A), len(B))))

def _read_excel_any(path: str, sheet=None, header=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    ext = os.path.splitext(path.lower())[1]
    kw = {"sheet_name": sheet if sheet is not None else 0, "header": header}
    if ext == ".xlsx":
        return pd.read_excel(path, engine="openpyxl", **kw)
    elif ext == ".xls":
        try:
            return pd.read_excel(path, engine="xlrd", **kw)
        except Exception as e:
            raise RuntimeError("Reading .xls requires xlrd==1.2.0 or convert the file to .xlsx") from e
    elif ext == ".xlsm":
        return pd.read_excel(path, engine="openpyxl", **kw)
    return pd.read_excel(path, **kw)


def load_uen_to_ssic_map(path: str, debug=False) -> Dict[str, str]:
    """Loads the UEN-to-SSIC mapping from the SSIC_List file using column names."""
    if not path or not os.path.exists(path):
        if debug: print("[INFO] UEN-to-SSIC list file not found. Skipping industry context.")
        return {}
    try:
        # Read the Excel file using the first row as the header
        df = pd.read_excel(path, engine="openpyxl", header=0)
        
        # Select only the 'UEN' and 'SSIC2020' columns
        df = df[['UEN', 'SSIC2020']]
        
        # Rename columns for consistency
        df.columns = ['uen', 'ssic']
        
        # Ensure data types are strings
        df['uen'] = df['uen'].astype(str)
        df['ssic'] = df['ssic'].astype(str)
        
        # Create the dictionary for fast lookup
        uen_map = pd.Series(df.ssic.values, index=df.uen).to_dict()
        
        if debug: print(f"[Industry] Loaded {len(uen_map)} UEN-to-SSIC mappings.")
        return uen_map
    except Exception as e:
        print(f"[ERROR] Failed to load UEN-to-SSIC map from {path}: {e}", file=sys.stderr)
        return {}

def load_ssic_definitions(path: str, debug=False) -> Dict[str, str]:
    """
    Loads the detailed SSIC definitions using fixed column indexes for maximum speed and reliability.
    It reads the header from row 5, includes specific columns in the context blob,
    explicitly excludes Column E ("Cross References"), and cleans up "<Blank>" values.
    """
    if not path or not os.path.exists(path):
        if debug: print("[INFO] SSIC definitions file not found. Skipping industry context.")
        return {}
    try:
        # 1. Read the file with the header on row 5 (index 4)
        df = pd.read_excel(path, engine="openpyxl", header=4)

        # 2. Define the exact column indexes to use
        # Column A: SSIC 2025 (The code)
        code_col_idx = 0
        # Columns B, C, D, F: Title, Groups, Definitions, Examples
        desc_col_indices = [1, 2, 3, 5] 
        # Column E (index 4) is now implicitly excluded.
        
        ssic_map = {}
        exclusion_pattern = re.compile(r'(\bsee\b|\s*�)\s*\d{5}\b')

        for _, r in df.iterrows():
            # Use .iloc for fast, direct index-based access
            ssic_code = str(r.iloc[code_col_idx]).strip()
            if not ssic_code: continue
            
            # 3. Build the text blob from the specified columns, ignoring "<Blank>"
            parts = []
            for idx in desc_col_indices:
                # Check if the column index is valid for this row
                if idx < len(r):
                    val = r.iloc[idx]
                    if pd.notna(val) and str(val).strip().lower() != '<blank>':
                        parts.append(str(val))
            
            blob = " ".join(parts)

            # Smartly truncate the blob to remove irrelevant exclusion text
            match = exclusion_pattern.search(blob)
            if match:
                blob = blob[:match.start()]
            
            ssic_map[ssic_code] = blob
            
        if debug: print(f"[Industry] Loaded {len(ssic_map)} detailed SSIC definitions using fixed indexes.")
        return ssic_map
    except Exception as e:
        print(f"[ERROR] Failed to load SSIC definitions from {path}: {e}", file=sys.stderr)
        return {}
    
# ---------- definitions loader ----------
HEADER_ALIASES_CODE  = {"ssoc 2024","ssoc","ssoc code","ssoc2024","ssoc_2024","code","occupation code"}
HEADER_ALIASES_TITLE = {"ssoc 2024 title","title","occupation title","job title","ssoc title"}
AUX_HEADERS = [
    "groups classified under this code","detailed definitions","tasks","notes",
    "examples of job classified under this code","examples of job classified elsewhere",
    "definition","examples","exclusions","description","responsibilities"
]

def _find_header_row(df_any_header_none) -> int:
    max_scan = min(15, len(df_any_header_none)); candidate = 0
    for ridx in range(max_scan):
        row_vals = [_norm_cell(v) for v in list(df_any_header_none.iloc[ridx, :]) ]
        if not any(row_vals): continue
        row_set = set(row_vals)
        if (row_set & HEADER_ALIASES_CODE) or (row_set & HEADER_ALIASES_TITLE): return ridx
        filled = sum(1 for v in row_vals if v)
        if filled >= max(3, int(0.4 * len(row_vals))): candidate = ridx
    return candidate

def _get_col_by_alias(columns: List[str], aliases: Set[str]) -> Optional[str]:
    cols_norm = {_norm_cell(c): c for c in columns}
    for alias in aliases:
        if alias in cols_norm: return cols_norm[alias]
    best_col, best_sc = None, 0.0
    for cnorm, corig in cols_norm.items():
        for alias in aliases:
            sc = _tok_overlap(cnorm, alias)
            if sc > best_sc:
                best_col, best_sc = corig, sc
    return best_col if best_sc >= 0.6 else None

def _split_and_normalize_titles(text: str) -> List[str]:
    """Helper to split titles by '/' and normalize each part."""
    if "/" not in text:
        # Normalize even if there's no slash, to handle single titles correctly.
        return [_normalize(text)]
    # If a slash is present, split and normalize each part.
    return [_normalize(part) for part in text.split('/')]

def load_definitions(path: str, def_sheet, def_skip_rows: int, debug=False) -> Tuple[List[Dict[str, str]], Dict[str, Dict]]:
    """
    Loads the SSOC definitions, intelligently parsing main titles and alternative titles
    (from the "Examples" column) that contain slashes to create a comprehensive lookup map.
    """
    ext = os.path.splitext(path.lower())[1]
    if ext in (".xlsx", ".xls", ".xlsm"):
        df_raw = _read_excel_any(path, sheet=def_sheet, header=None)
        if def_skip_rows: df_raw = df_raw.iloc[def_skip_rows:].reset_index(drop=True)
        header_row = _find_header_row(df_raw)
        df = _read_excel_any(path, sheet=def_sheet, header=header_row + def_skip_rows)
    else:
        df = _read_excel_any(path, sheet=def_sheet, header=0)

    columns = list(df.columns)
    headers_norm = _canon_headers(columns)
    code_col  = _get_col_by_alias(columns, HEADER_ALIASES_CODE)
    title_col = _get_col_by_alias(columns, HEADER_ALIASES_TITLE)
    
    # Find the specific "Examples" column
    HEADER_ALIASES_EXAMPLES = {"examples of job classified under this code", "examples"}
    example_col = _get_col_by_alias(columns, HEADER_ALIASES_EXAMPLES)

    aux_cols = {h: columns[headers_norm.index(h)] for h in AUX_HEADERS if h in headers_norm}

    rows: List[Dict[str, str]] = []
    for _, r in df.iterrows():
        code = str(r.get(code_col, "")).strip()
        title = str(r.get(title_col, "")).strip()
        if not code and not title: continue
        rec: Dict[str, str] = {"code": code, "title": title}

        # --- Create a comprehensive list of all title variations ---
        all_title_variations = []
        # Add variations from the main title (e.g., "Tailor/Dressmaker")
        all_title_variations.extend(_split_and_normalize_titles(title))

        # Parse and add variations from the "Examples" column
        rec['alternative_titles'] = []
        if example_col:
            val = r.get(example_col, "")
            if pd.notna(val):
                # This regex now splits by comma, semicolon, OR the non-standard bullet points.
                raw_alt_titles = re.split(r'[;,]|\s*�.*?\s*', str(val))
                for alt_title in raw_alt_titles:
                    # Also split any alternative titles that contain a slash
                    split_alts = _split_and_normalize_titles(alt_title)
                    rec['alternative_titles'].extend(split_alts)
        
        all_title_variations.extend(rec['alternative_titles'])
        # Store the unique, clean list of all possible titles
        rec['all_title_variations'] = sorted(list(set(filter(None, all_title_variations))))
        # -----------------------------------------------------------

        parts = []
        for h, col in aux_cols.items():
            val = r.get(col, "")
            if pd.notna(val):
                sval = str(val)
                rec[h] = sval
                parts.append(sval)

        blob = " | ".join(p for p in parts if p) or title
        rec["search_text"] = blob

        # Precompute other values
        rec["title_norm"]       = _normalize(title)
        rec["blob_norm"]        = _normalize(blob)
        rec["title_tokens_set"] = set(_tokens(rec["title_norm"]))
        rec["blob_tokens_set"]  = set(_tokens(rec["blob_norm"]))
        rec["blob_bigrams"]     = _bigrams_from_text(rec["blob_norm"])
        rec["sector_cues"]      = _sector_cues_from_text(rec["title_norm"] + " " + rec["blob_norm"])
        rec["is_5d"]            = rec["code"].isdigit() and len(rec["code"]) == 5

        rows.append(rec)

    # --- Create the expanded title-to-record lookup map ---
    title_to_record_map = {}
    for r in rows:
        # Add all variations (main, split, and alternative) to the map.
        for variation in r.get('all_title_variations', []):
            title_to_record_map[variation] = r
        
        # Ensure the original, un-split normalized title is also a key
        # This handles the case where "Tailor/Dressmaker" might be searched for literally.
        if r.get("title_norm"):
            title_to_record_map[r["title_norm"]] = r
    # ----------------------------------------------------

    if debug:
        print(f"[Definitions] Loaded {len(rows)} SSOC records.")
        print(f"[Title Map] Created a lookup map with {len(title_to_record_map)} total title variations.")
    
    return rows, title_to_record_map
# ---------- jobs loader ----------
def _choose_col_by_name_or_index(columns: List[str], name: Optional[str], idx: Optional[int]) -> Optional[str]:
    if name:
        target = _strip_punct_lower(name)
        for c in columns:
            if _strip_punct_lower(str(c)) == target: return c
        for c in columns:
            if target and target in _strip_punct_lower(str(c)): return c
    if idx is not None and 0 <= idx < len(columns): return columns[idx]
    return None

def _fuzzy_find_column(columns: List[str], patterns: List[str]) -> Optional[str]:
    best, best_sc = None, 0.0
    for c in columns:
        cn = _strip_punct_lower(str(c))
        for p in patterns:
            sc = _tok_overlap(cn, p)
            if sc > best_sc:
                best, best_sc = c, sc
    return best if best_sc >= 0.35 else None

def load_jobs_separate(path: str, sheet, header_row: int,
                       title_name: Optional[str], edu_name: Optional[str],
                       title_idx: Optional[int], edu_idx: Optional[int],
                       debug=False) -> Tuple[List[int], List[str], List[str], List[str], List[str], str]:
    uen_for_file = ""
    try:
        df_raw = pd.read_excel(path, header=None, engine="openpyxl")
        uen_for_file = str(df_raw.iloc[3, 2]).strip()
        if debug: print(f"[Jobs] Found UEN in cell C4: {uen_for_file}")
    except Exception as e:
        if debug: print(f"[WARN] Could not read UEN from cell C4: {e}", file=sys.stderr)
    
    df = _read_excel_any(path, sheet=sheet, header=header_row)
    columns = list(df.columns)
    if debug:
        print("[Jobs] using sheet:", sheet if sheet is not None else 0)
        print("[Jobs] header row (1-based):", header_row + 1)
        print("[Jobs] columns:", columns)

    # Primary duties column is AP (index 41), Fallback is I (index 8)
    duties_col_primary_idx = 41 
    col_duties_primary = columns[duties_col_primary_idx] if duties_col_primary_idx < len(columns) else None
    duties_col_fallback_idx = 8
    col_duties_fallback = columns[duties_col_fallback_idx] if duties_col_fallback_idx < len(columns) else None

    if debug:
        print(f"[Jobs] Primary duties column (by index {duties_col_primary_idx}): {col_duties_primary}")
        print(f"[Jobs] Fallback duties column (by name 'Main Job Duties'): {col_duties_fallback}")

    col_title  = _choose_col_by_name_or_index(columns, title_name,  title_idx)
    col_edu = _choose_col_by_name_or_index(columns, edu_name, edu_idx)
    if not col_edu:
        col_edu = _fuzzy_find_column(columns, patterns=["highest education attained", "highest edu attained", "highest qualification", "education level", "qualification (annex a)", "annex a"])

    col_group = _fuzzy_find_column(columns, patterns=["occupation group (see annex d)", "occupation group", "group (annex d)", "annex d"])
    
    col_fullname = _fuzzy_find_column(columns, patterns=["full name", "name of employee", "employee name", "name"])

    if not col_title and not (col_duties_primary or col_duties_fallback):
        return [], [], [], [], [], ""

    t = df[col_title]  if col_title  else pd.Series([""] * len(df))
    d_primary = df[col_duties_primary] if col_duties_primary else pd.Series([None] * len(df))
    d_fallback = df[col_duties_fallback] if col_duties_fallback else pd.Series([None] * len(df))
    
    duties_series = d_primary.fillna(d_fallback)

    e = df[col_edu]    if col_edu    else pd.Series([""] * len(df))
    g = df[col_group]  if col_group  else pd.Series([""] * len(df))
    fn = df[col_fullname] if col_fullname else pd.Series([])

    titles  = [("" if pd.isna(x) else str(x)).strip() for x in t.tolist()]
    duties  = [("" if pd.isna(x) else str(x)).strip() for x in duties_series.tolist()]
    edus    = [("" if pd.isna(x) else str(x)).strip() for x in e.tolist()]
    groups  = [("" if pd.isna(x) else str(x)).strip() for x in g.tolist()]
    full_names = [("" if pd.isna(x) else str(x)).strip() for x in fn.tolist()]
    data_row_indices = list(range(header_row + 1, header_row + 1 + len(df)))
    
    stop_index = None
    if full_names:
        for i, name in enumerate(full_names):
            if not name:
                stop_index = i
                if debug:
                    print(f"[INFO] Empty name found at data row {i + 1}. Stopping processing there.")
                break
    
    if stop_index is not None:
        titles = titles[:stop_index]
        duties = duties[:stop_index]
        edus = edus[:stop_index]
        groups = groups[:stop_index]
        data_row_indices = data_row_indices[:stop_index]
    
    return data_row_indices, titles, duties, edus, groups, uen_for_file
# ---------- X-codes ----------
X_TITLE_MAP = {
    "X1000": "Worker reporting unidentifiable or inadequately described occupation",
    "X2000": "Worker not reporting any occupation",
    "X3000": "Singapore armed forces personnel",
    "X4000": "Foreign armed forces personnel",
    "X5000": "Foreign diplomatic personnel",
}

_SG_ARMED_RX   = re.compile(r"\b(?:saf|singapore armed forces|singapore army|rsn|rsaf|mindef)\b", re.I)
_FOR_ARMED_RX  = re.compile(r"\b(?:foreign armed forces|foreign military|u\.s\.|us army|royal navy|adf|pla|idf)\b", re.I)
_DIPLO_RX      = re.compile(r"\b(?:ambassador|high commissioner|attach[eé]|consul|consular|diplomat|diplomatic|embassy|charg[eé]\s+d'affaires|charge d affaires)\b", re.I)

def _xcode_checker(duties_text: str, title_text: str) -> Optional[Tuple[str, str]]:
    d = _normalize(duties_text)
    t = _normalize(title_text)
    both = f"{d} {t}".strip()
    if not d and not t:
        return "X2000", X_TITLE_MAP["X2000"]
    hits = 0
    if _SG_ARMED_RX.search(both):   hits += 2
    if _FOR_ARMED_RX.search(both):  hits += 2
    if _DIPLO_RX.search(both):      hits += 2
    if hits >= 2:
        if _SG_ARMED_RX.search(both):  return "X3000", X_TITLE_MAP["X3000"]
        if _FOR_ARMED_RX.search(both): return "X4000", X_TITLE_MAP["X4000"]
        if _DIPLO_RX.search(both):     return "X5000", X_TITLE_MAP["X5000"]
    return None

# ---------- scoring helpers & guards ----------
# ---------- scoring helpers & guards ----------
_SUPERVISORY_TOKENS = {
    "supervisor", "supervise", "supervising", "supervision", 
    "foreman", "foremen", "manager", "managers", "managing", 
    "head", "chief", "lead", "leader", "oversee", "overseeing", 
    "coordinating", "coordination", "directing", "director",
    "superintendent" # <-- ADDED
}
_SUPERVISE_CUES_IN_QUERY = {"supervise","supervision","manage","oversee","lead","coordinate","assign","schedule","train","coach","report","budget","plan","crew","team"}
_VEHICLE_CUES = re.compile(r"\b(auto|car|vehicle|motor|automotive|bodyshop|panel|bumper|spray booth)\b", re.I)
_BUILDING_PAINT_CUES = re.compile(r"\b(wall|walls|ceiling|ceilings|facade|façade|interior|exterior|building|structure|premises|floor|room|units?)\b", re.I)
_ARTS_MEDIA_CUES = re.compile(r"\b(orchestra|choir|stage|film|theatre|theater|shoot|broadcast|studio|gallery|museum|curator|composer|conductor|perform(ing|ance)|band)\b", re.I)
_HOSPITALITY_CUES = re.compile(r"\b(hotel|resort|guest|housekeeping|butler|steward|banquet|front office|concierge|lodging)\b", re.I)
_GAMING_CUES = re.compile(r"\b(casino|gaming|pit boss|tables?|jackpot|slots?)\b", re.I)
_SPORTS_CUES = re.compile(r"\b(sport(s)?|stadium|arena|gym|fitness|leisure|recreation|club)\b", re.I)
_MARKETING_CUES = re.compile(r"\b(marketing|brand|branding|campaign|advertis(e|ing)|digital\s+marketing)\b", re.I)
_CONSTRUCTION_LABOUR_CUES = re.compile(r"\b(work\s?site|worksite|cleaning work\s?sites?|remove (site )?obstructions|debris|demolition|trench|scaffold|construction|site)\b", re.I)
_ELECTRICIAN_CUES = re.compile(r"\b(electrician|electrical wiring|install(ing|ation) electrical|switchboard|cable pulling|cabling)\b", re.I)
_DRAFTER_CUES = re.compile(r"\b(autocad|auto-cad|cad|drafter|drafts?man|draftsperson|bim|shop drawings?|technical drawings?|layouts?)\b", re.I)
_ADMIN_MANAGER_CUES = re.compile(r"\b(manage(s|d|ment)? (admin(istration)?|team|department)|oversee admin|lead admin)\b", re.I)
_CONSULTING_CUES = re.compile(r"\b(consult(ant|ing)|advisory|strategy engagement|diagnostic study)\b", re.I)
_ATTRACTIONS_CUES = re.compile(r"\b(attraction(s)?|theme park|zoo|park manager|nature park|botanic)\b", re.I)
_POLITICAL_CUES = re.compile(r"\b(party (official|organisation|organization)|political party|secretariat|central committee|grassroots|constituency)\b", re.I)
_JUNIOR_TITLE_CUES = {"executive", "assistant", "asst", "associate", "clerk", "coordinator","junior","attendant"}  
_WELLNESS_CUES = re.compile(r"\b(wellness|spa|fitness|health|therapy|recreation)\b", re.I)


ROLE_CLUSTERS: Dict[str, Set[str]] = {
    # For roles involving project management, planning, and operational coordination.
    "project_management": {
        "project", "planning", "schedule", "scheduling", "coordination", "delivery", 
        "timeline", "milestones", "operations"
    },
    # For roles focused on sales, business growth, and client acquisition.
    "sales_business_development": {
        "sales", "business development", "bizdev", "account management", "client acquisition", 
        "revenue", "growth", "leads", "pipeline"
    },
    # For data-centric roles involving analysis and reporting.
    "data_analysis": {
        "data", "analytics", "reporting", "dashboard", "insights", "metrics", "kpi", 
        "business intelligence", "bi"
    },
    # For roles that provide direct support and service to customers.
    "customer_support": {
        "customer service", "support", "helpdesk", "client relations", "technical support", 
        "issue resolution", "troubleshooting"
    },
    # For financial and accounting functions.
    "finance_accounting": {
        "finance", "accounting", "bookkeeping", "ledger", "invoicing", "payroll",
        "accounts payable", "ap", "accounts receivable", "ar", "financial reporting", "audit", "tax"
    },
    # For marketing, communications, and public relations roles.
    "marketing_communications": {
        "marketing", "communications", "comms", "public relations", "pr", "branding", 
        "campaigns", "content", "social media", "digital marketing"
    },
    # For roles focused on quality control and standards compliance.
    "quality_assurance": {
        "quality", "qa", "qc", "quality control", "quality assurance", "testing", 
        "inspection", "compliance", "standards"
    },
    # For logistics, purchasing, and supply chain management.
    "supply_chain_logistics": {
        "logistics", "supply chain", "procurement", "purchasing", "sourcing", "inventory", 
        "shipping", "freight", "warehouse"
    },
    # For creative roles involving visual and user-focused design.
    "creative_design": {
        "design", "graphic", "ui", "ux", "user interface", "user experience", "visual", 
        "creative", "illustration"
    }
}


# SSOC codes for corporate managers (Finance, HR, IT) that are often incorrectly matched.
_CORPORATE_MANAGER_CODES = {
    "12111", # Finance Manager
    "12121", # Human Resource Manager
    "13301", # Information and Communications Technology Manager
    "24111", # Accountant (often appears for Finance Manager)
}

# Keywords that indicate a specific corporate function (finance, HR, IT).
_CORPORATE_FUNCTION_KEYWORDS = {
    "finance", "financial", "accounting", "audit", "tax", "treasury", "risk",
    "hr", "human resource", "payroll", "recruitment", "talent",
    "it", "ict", "software", "network", "cybersecurity", "infrastructure"
}

# Keywords that indicate a hands-on, operational (non-corporate) context.
_OPERATIONAL_CONTEXT_KEYWORDS = {
    "workshop", "factory", "site", "production", "plant", "construction", 
    "repair", "maintenance", "vehicle", "machinery"
}

# Keywords for textile/garment machine operations to differentiate from metal/wood machining.
_TEXTILE_MACHINE_KEYWORDS = {"sewing", "garment", "textile", "fabric", "stitching", "embroidery", "overlock"}

_MARINE_CONTEXT_KEYWORDS = {
    "marine", "maritime", "ship", "ships", "vessel", "vessels", 
    "shipyard", "deck", "hull", "offshore"
}

# Specific SSOC codes for machine operators that require this special context check.
_MACHINE_OPERATOR_PENALTY_CODES = {
    "72231": "Machine-tool setter-operator", # Should NOT have textile words
    "81531": "Sewing machine operator"       # Should NOT have machining words
}
# Specific technical keywords that are strong indicators of a particular job function.
_TECHNICAL_KEYWORDS = {
    # CAD/CAM & Machining
    "cnc", "cam", "cad", "machining", "machinist", "lathe", "milling", "grinding", "fabrication", 
    "welding", "g-code", "m-code", "mastercam", "solidworks", "autocad", "unigraphics",
    "3d model", "setter", "operator", "programmer"
}

# Generic engineering words that should be ignored when differentiating between disciplines.
_GENERIC_ENGINEERING_TOKENS = {
    "engineer", "engineering", "technical", "technician", "design", "develop", 
    "specifications", "testing", "analysis", "solutions", "consultancy"
}


# Specific technical keywords that are strong indicators of a particular job function.
HIGH_VALUE_KEYWORDS = {
    "system", "systems", "software", "embedded", "network", "database", "cloud",
    "security", "cyber", "mechanical", "electrical", "civil", "structural",
    "chemical", "optical", "materials",
    # === NEW: Machining & CAD/CAM Keywords ===
    "cnc", "cam", "cad", "machining", "machinist", "lathe", "milling", "fabrication", 
    "welding", "g-code", "solidworks", "autocad"
}

_DRAFTER_DISCIPLINES = {
    "civil": ({"civil", "structural", "construction", "building"}, "31183"),
    "electrical": ({"electrical", "power", "switchboard"}, "31182"),
    "mechanical": ({"mechanical", "piping", "hvac", "acmv", "workshop"}, "31181"),
    "architectural": ({"architectural", "floor plan", "elevation", "interior", "architect"}, "31184"),
}

# Regex to find subordinate phrases like "assisting director" or "reporting to the manager"
_SUBORDINATE_PHRASE_RX = re.compile(r"\b(assist(ing)?|support(ing)?|report(ing)?\s+to?)\s+(\w+\s+){0,2}(director|manager|ceo|chief)\b", re.I)
# The specific SSOC codes for top-level directors that this penalty should apply to
_DIRECTOR_LEVEL_CODES = {"11201", "11203"}
# Generic words in titles that do not describe a job's function and should be ignored for title-keyword boosting.
_GENERIC_TITLE_KEYWORDS = {
    "associate", "assistant", "executive", "officer", "specialist", 
    "professional", "staff", "clerk", "assoc", "prof",
    "senior", "junior", "lead", "principal"
}
# Keywords for specific engineering disciplines to prevent cross-contamination.
_SOFTWARE_SYSTEMS_KEYWORDS = {"system", "systems", "software", "embedded", "network", "database", "cloud", "security", "cyber", "application", "it"}
_PHYSICAL_SCIENCES_KEYWORDS = {"materials", "chemical", "civil", "structural", "optical", "mechanical", "mining", "petroleum"}
_SAFETY_DISCIPLINES = {
    "occupational_health": ({"occupational", "health", "wsh", "workplace", "ergonomics", "risk assessment"}, "32572"),
    "industrial_engineering": ({"industrial", "factory", "plant", "manufacturing", "machinery", "construction", "engineer"}, "21493"),
    "fire": ({"fire", "alarm", "sprinkler", "extinguisher", "evacuation", "scdf"}, "31711"),
    "product_vehicle": ({"vehicle", "automotive", "product", "process", "quality", "component"}, "31720"),
}
_DIGITAL_DESIGN_KEYWORDS = {
    "digital", "graphic", "multimedia", "web", "ui", "ux", "marketing", 
    "visual content", "branding", "illustrator", "photoshop", "figma"
}

def _design_discipline_penalty(candidate_code: str, query_text: str) -> float:
    """
    Applies a heavy penalty to Interior Designer if the query contains keywords
    related to digital, graphic, or marketing design.
    """
    # This penalty is specifically targeted at the Interior Designer code.
    if candidate_code != "34321":
        return 1.0

    q_toks = set(_tokens(query_text))

    # If the job description contains ANY digital design keywords, it's a clear mismatch.
    if not q_toks.isdisjoint(_DIGITAL_DESIGN_KEYWORDS):
        return 0.10 # Apply a massive 90% penalty

    return 1.0

def _machinist_drafter_penalty(candidate_code: str, query_text: str) -> float:
    """
    Applies a heavy penalty to drafter candidates if the query text contains
    strong keywords related to hands-on machining.
    """
    # This penalty only applies to the Draughtsperson sub-group (3118x)
    if not candidate_code.startswith("3118"):
        return 1.0

    q_toks = set(_tokens(query_text))
    
    # Define a specific subset of keywords that are unambiguous signals for machining.
    machining_cues = {"cnc", "machining", "machinist", "lathe", "milling", "setter"}

    # If the job description contains ANY of these strong machining keywords, it is
    # highly unlikely to be a pure drafting role. Penalize heavily.
    if not q_toks.isdisjoint(machining_cues):
        return 0.15 # Apply a massive 85% penalty

    return 1.0

def _safety_discipline_handler(candidate_code: str, query_text: str) -> float:
    """
    If the query is for a safety role, this boosts the correct safety sub-type and
    penalizes incorrect ones based on contextual keywords.
    """
    q_toks = set(_tokens(query_text))
    
    # This logic only activates if the query contains "safety".
    if "safety" not in q_toks:
        return 1.0

    # Score each discipline based on keyword matches
    scores = {}
    for discipline, (keywords, _) in _SAFETY_DISCIPLINES.items():
        scores[discipline] = len(q_toks & keywords)

    # If no specific discipline keywords are found, do nothing.
    if not any(scores.values()):
        # Default to a general safety officer code if no other context is found
        if candidate_code == "32573": return 1.20 # Boost for the generalist
        return 1.0
    
    # Find the winning discipline (the one with the most matching keywords)
    winning_discipline = max(scores, key=scores.get)
    winning_ssoc_code = _SAFETY_DISCIPLINES[winning_discipline][1]

    # Apply a strong boost to the winner and a penalty to the losers
    if candidate_code == winning_ssoc_code:
        return 1.50 # Strong 50% boost for the correct discipline
    elif candidate_code in [code for _, code in _SAFETY_DISCIPLINES.values()]:
        return 0.20 # Heavy 80% penalty for the wrong discipline
        
    return 1.0

def _marine_context_penalty(candidate_rec: Dict, query_text: str) -> float:
    """
    Applies a heavy penalty if the query has a clear marine context, but the
    candidate SSOC code is not in the aviation_marine sector.
    """
    q_toks = set(_tokens(query_text))

    # First, check if the query has any marine keywords. If not, do nothing.
    if q_toks.isdisjoint(_MARINE_CONTEXT_KEYWORDS):
        return 1.0

    # The query is clearly about a marine role. Now check the candidate.
    candidate_cues = candidate_rec.get("sector_cues", set())
    
    # If the candidate is NOT in the correct sector, apply a massive penalty.
    if "aviation_marine" not in candidate_cues:
        return 0.05 # Apply a 95% penalty

    # The context is marine and the candidate is in the right sector. No penalty.
    return 1.0

def _score_title_similarity(input_title_text: str, candidate_rec: Dict) -> float:
    """
    Calculates a title similarity score by testing the input against EACH possible
    title variation (e.g., "General practitioner" and "Physician") and taking the best score.
    """
    input_toks = set(_tokens(input_title_text))
    filtered_input_toks = input_toks - _GENERIC_TITLE_KEYWORDS
    
    if not filtered_input_toks:
        return 0.0

    # Get the list of all possible valid titles for this SSOC code
    all_title_variations = candidate_rec.get("all_title_variations", [])
    if not all_title_variations:
        return 0.0 # No valid titles to compare against

    max_score = 0.0

    # === NEW: Loop through each possibility and find the best match ===
    for variation in all_title_variations:
        cand_toks = set(_tokens(variation))
        filtered_cand_toks = cand_toks - _GENERIC_TITLE_KEYWORDS

        if not filtered_cand_toks:
            continue

        # Use the robust Jaccard Index for a fair comparison
        intersection = filtered_input_toks.intersection(filtered_cand_toks)
        union = filtered_input_toks.union(filtered_cand_toks)

        if not union:
            continue

        jaccard_score = len(intersection) / float(len(union))
        
        # Keep track of the highest score found so far
        if jaccard_score > max_score:
            max_score = jaccard_score
    # ===============================================================
    
    return max_score

def _drafter_discipline_handler(candidate_code: str, query_text: str) -> float:
    """
    If the query is for a drafter, this boosts the correct drafter sub-type and
    penalizes incorrect ones based on contextual keywords.
    """
    q_norm = _normalize(query_text)
    
    # === CORRECTED: Use the regex pattern to search the text ===
    # This logic now correctly activates if a drafter-related keyword is found.
    if not _DRAFTER_CUES.search(q_norm):
        return 1.0
    # ==========================================================

    q_toks = set(_tokens(q_norm))
    
    # Score each discipline based on keyword matches
    scores = {}
    for discipline, (keywords, _) in _DRAFTER_DISCIPLINES.items():
        scores[discipline] = len(q_toks & keywords)

    # Find the winning discipline (the one with the most matching keywords)
    if not any(scores.values()): # No discipline keywords found
        return 1.0
    
    winning_discipline = max(scores, key=scores.get)
    winning_ssoc_code = _DRAFTER_DISCIPLINES[winning_discipline][1]

    # Apply a strong boost to the winner and a penalty to the losers
    if candidate_code == winning_ssoc_code:
        return 1.50 # Strong 50% boost for the correct discipline
    elif candidate_code in [code for _, code in _DRAFTER_DISCIPLINES.values()]:
        return 0.20 # Heavy 80% penalty for the wrong discipline
        
    return 1.0


def _validate_expert_match(title_text: str, duties_text: str, expert_map: Dict[str, Tuple[str, str]], all_defs: List[Dict]) -> Optional[Tuple[str, str]]:
    """
    Checks if a title exists in the expert map and validates it against the duties
    using the context-scorers for managers and engineers.
    """
    norm_title = _normalize(title_text)
    if norm_title not in expert_map:
        return None

    ssoc_code, ssoc_title = expert_map[norm_title]
    
    # Find the full record for the candidate SSOC code
    candidate_rec = next((rec for rec in all_defs if rec.get("code") == ssoc_code), None)
    if not candidate_rec:
        return ssoc_code, ssoc_title # Return as-is if we can't find the record

    combined_text = f"{title_text} {duties_text}"
    context_score = 0.0
    
    # If the expert title is a manager, validate with the manager context scorer
    if "manager" in norm_title:
        context_score = _score_manager_context(combined_text, candidate_rec.get("search_text", ""))
    
    # If the expert title is an engineer, validate with the engineer context scorer
    elif "engineer" in norm_title:
        context_score = _score_engineer_context(combined_text, candidate_rec.get("search_text", ""))

    # If it's not a manager/engineer, or if the context score is good, accept the match.
    # A score of > 0 means there is at least some contextual keyword overlap.
    if context_score > 0.0 or ("manager" not in norm_title and "engineer" not in norm_title):
        return ssoc_code, ssoc_title
        
    # If the context score is zero, the duties contradict the title. Reject the expert match.
    return None

def _corporate_manager_penalty(candidate_code: str, query_text: str, query_title: str) -> float:
    """
    Applies a heavy penalty to corporate manager codes when the job context is
    clearly operational OR when the context is too generic.
    """
    if candidate_code not in _CORPORATE_MANAGER_CODES:
        return 1.0

    q_toks = set(_tokens(query_text))

    # Penalty 1: If the context is clearly operational (e.g., a workshop), penalize.
    if not q_toks.isdisjoint(_OPERATIONAL_CONTEXT_KEYWORDS):
        return 0.15

    # === NEW: Penalty 2: If the title is generic AND the duties are also generic ===
    # This prevents "Manager" from defaulting to "Finance Manager".
    norm_title = _normalize(query_title)
    is_title_generic = (norm_title == "manager" or norm_title == "general manager")
    
    if is_title_generic:
        # Check if the duties lack ANY specific corporate function keywords.
        if q_toks.isdisjoint(_CORPORATE_FUNCTION_KEYWORDS):
            return 0.20 # Apply a heavy 80% penalty
    # =========================================================================

    return 1.0

def _score_engineer_context(query_text: str, candidate_text: str) -> float:
    """
    Calculates a similarity score based only on non-generic, discipline-specific keywords
    to help differentiate between types of engineers.
    """
    q_toks = set(_tokens(query_text))
    c_toks = set(_tokens(candidate_text))

    # Remove all generic engineering words from both sets
    q_diff_toks = q_toks - _GENERIC_ENGINEERING_TOKENS
    c_diff_toks = c_toks - _GENERIC_ENGINEERING_TOKENS

    # If there are no differentiating keywords left, we can't make a judgment.
    if not q_diff_toks or not c_diff_toks:
        return 0.0

    # Calculate the Jaccard similarity of the *remaining*, discipline-specific keywords.
    intersection = len(q_diff_toks & c_diff_toks)
    union = len(q_diff_toks | c_diff_toks)
    
    return intersection / float(union) if union > 0 else 0.0

def _machine_operator_context_penalty(candidate_code: str, query_text: str) -> float:
    """
    Applies a heavy penalty if there is a context mismatch between machining and textile keywords.
    """
    if candidate_code not in _MACHINE_OPERATOR_PENALTY_CODES:
        return 1.0

    q_toks = set(_tokens(query_text))

    # Case 1: Candidate is a Sewing Machine Operator, but query has CNC/machining terms.
    if candidate_code == "81531" and not q_toks.isdisjoint(_TECHNICAL_KEYWORDS):
        return 0.05 # Massive 95% penalty for clear mismatch

    # Case 2: Candidate is a CNC/Machine-Tool Operator, but query has sewing/textile terms.
    if candidate_code == "72231" and not q_toks.isdisjoint(_TEXTILE_MACHINE_KEYWORDS):
        return 0.05 # Massive 95% penalty for clear mismatch

    return 1.0

def _score_manager_context(query_text: str, candidate_text: str) -> float:
    """
    Calculates a similarity score based only on non-managerial, context-specific keywords.
    This helps differentiate between types of managers.
    """
    q_toks = set(_tokens(query_text))
    c_toks = set(_tokens(candidate_text))

    # Remove all generic supervisory/managerial words from both sets
    q_diff_toks = q_toks - _SUPERVISORY_TOKENS
    c_diff_toks = c_toks - _SUPERVISORY_TOKENS

    # If there are no differentiating keywords left, we can't make a judgment.
    if not q_diff_toks or not c_diff_toks:
        return 0.0

    # Calculate the Jaccard similarity of the *remaining*, context-rich keywords.
    intersection = len(q_diff_toks & c_diff_toks)
    union = len(q_diff_toks | c_diff_toks)
    
    return intersection / float(union) if union > 0 else 0.0

def _engineering_discipline_penalty(candidate_code: str, query_text: str) -> float:
    """
    Applies a heavy penalty if there is a clear mismatch between software/systems
    and physical science engineering disciplines.
    """
    # This filter only applies to a specific range of professional engineer codes (21400-21600 block)
    try:
        code_num = int(candidate_code)
        if not (21400 <= code_num <= 21600):
            return 1.0
    except (ValueError, TypeError):
        return 1.0

    q_toks = set(_tokens(query_text))

    # Check for a mismatch
    has_software_cues = not q_toks.isdisjoint(_SOFTWARE_SYSTEMS_KEYWORDS)
    has_physical_cues = not q_toks.isdisjoint(_PHYSICAL_SCIENCES_KEYWORDS)

    # If the query is about software, but the candidate is a physical science engineer, penalize heavily.
    if has_software_cues and not has_physical_cues:
        if candidate_code in {"21492", "21497", "21460", "21451", "21421"}: # Materials, Optical, Mining, Chemical, Civil
            return 0.10 # Apply a 90% penalty

    # If the query is about physical sciences, but the candidate is a software/systems engineer, penalize heavily.
    if has_physical_cues and not has_software_cues:
        if candidate_code in {"21526", "21522"}: # Embedded Systems, Computer Engineer
            return 0.10 # Apply a 90% penalty

    return 1.0

def _subordinate_context_penalty(candidate_code: str, query_text: str) -> float:
    """
    Applies a heavy penalty if the candidate is a Director but the query text
    contains subordinate phrases like "assisting the director".
    """
    # This penalty only applies to top-level director codes.
    if candidate_code not in _DIRECTOR_LEVEL_CODES:
        return 1.0

    # If we find a subordinate phrase in the job title or duties...
    if _SUBORDINATE_PHRASE_RX.search(query_text):
        # ...it's a strong sign this person is NOT a director, so apply a huge penalty.
        return 0.15  # Apply an 85% penalty

    # No subordinate context found, so no penalty.
    return 1.0

# Codes for managers in specific industries that need strong contextual keywords to be considered valid.
_SPECIFIC_INDUSTRY_MANAGER_CODES = {
    "13210": "Manufacturing manager",
    "14121": "Restaurant manager",
    "14201": "Hotel manager",
    "14310": "Sports centre manager",
    "14323": "Discotheque/Karaoke/Nightclub manager",
    "14329": "Recreation centre manager n.e.c.",
    "14324": "Wellness centre manager" # Requires wellness/spa/fitness cues
}

# Helper keyword checks for the specific manager penalty
def _mfg_cues_for_penalty(q_norm: str) -> bool:
    return any(cue in q_norm for cue in ["manufactur", "production", "factory", "plant", "line"])

def _fnb_cues_for_penalty(q_norm: str) -> bool:
    return any(cue in q_norm for cue in ["restaurant", "f&b", "food", "beverage", "kitchen", "cafe"])

def _specific_manager_penalty(candidate_code: str, query_text: str) -> float:
    """
    Applies a heavy penalty to specific-industry manager codes if the query lacks
    corresponding industry keywords.
    """
    if candidate_code not in _SPECIFIC_INDUSTRY_MANAGER_CODES:
        return 1.0 # Not a specific manager, no penalty.

    q_norm = _normalize(query_text)
    
    # Check for required context. If missing, apply a huge 90% penalty.
    if candidate_code == "13211" and not _mfg_cues_for_penalty(q_norm): return 0.10
    if candidate_code == "14121" and not _fnb_cues_for_penalty(q_norm): return 0.10
    if candidate_code == "14201" and not _HOSPITALITY_CUES.search(q_norm): return 0.10
    if candidate_code == "14310" and not _SPORTS_CUES.search(q_norm): return 0.10
    
    # === NEW: Add the check for Wellness and Recreation managers ===
    if candidate_code in {"14323", "14329", "14324"} and not (_HOSPITALITY_CUES.search(q_norm) or _SPORTS_CUES.search(q_norm) or _WELLNESS_CUES.search(q_norm)):
        return 0.10
    # ===============================================================

    return 1.0 # Context was found, no penalty.

def _get_industry_multiplier(company_industry_description: str, candidate_rec: Dict) -> float:
    """
    Calculates a graduated industry context multiplier based on the strength of keyword overlap.
    """
    if not company_industry_description:
        return 1.0

    company_cues = _sector_cues_from_text(company_industry_description)
    candidate_cues = candidate_rec.get("sector_cues", set())

    if not company_cues or not candidate_cues:
        return 1.0

    intersection = company_cues & candidate_cues
    
    # Calculate overlap as a percentage of the candidate's keywords
    overlap_ratio = len(intersection) / float(len(candidate_cues))

    # === TUNED VALUES ===
    if overlap_ratio > 0.75: # Very strong overlap required for max boost
        return 1.40 # Was 1.50
    elif overlap_ratio > 0.40: # Moderate overlap
        return 1.20 # Was 1.20
    else: # Weak or no overlap
        return 1.0 # No boost for weak connections
    # ====================

def _seniority_penalty(candidate_title: str, query_text: str, candidate_blob: str) -> float:
    ctoks = set(_tokens(candidate_title)) | set(_tokens(candidate_blob))
    if not (ctoks & _SUPERVISORY_TOKENS):
        return 1.0
    if set(_tokens(query_text)) & _SUPERVISE_CUES_IN_QUERY:
        return 1.0
    qsecs = _sector_cues_from_text(query_text)
    csecs = _sector_cues_from_text(candidate_title + " " + candidate_blob)
    return 0.90 if (qsecs & csecs) else 0.65

def _title_seniority_conflict_penalty(input_title_text: str, input_duties_text: str, candidate_text: str) -> float:
    """
    Penalizes a senior/managerial SSOC candidate if the INPUT TITLE is junior,
    even if the INPUT DUTIES sound senior. This respects the client's provided title.
    """
    cand_toks = set(_tokens(candidate_text))
    # If the candidate code is not even a senior role, there's no conflict.
    if not (cand_toks & _SUPERVISORY_TOKENS):
        return 1.0

    title_toks = set(_tokens(input_title_text))
    duties_toks = set(_tokens(input_duties_text))

    # Check if the input title is junior (lacks manager words OR contains explicit junior cues)
    is_title_junior = not (title_toks & _SUPERVISORY_TOKENS) or (title_toks & _JUNIOR_TITLE_CUES)
    # Check if the duties sound senior
    are_duties_senior = bool(duties_toks & _SUPERVISE_CUES_IN_QUERY)

    # If the title is junior BUT the duties sound senior, we have a conflict.
    # Penalize the senior SSOC candidate.
    if is_title_junior and are_duties_senior:
        return 0.40  # Apply a heavy 60% penalty

    return 1.0 # No conflict, no penalty

def _get_cluster_boost(query_text: str, candidate_text: str) -> float:
    """
    Applies a boost if the query and candidate share conceptually related keywords.
    """
    q_toks = set(_tokens(query_text))
    c_toks = set(_tokens(candidate_text))
    
    for cluster_name, keywords in ROLE_CLUSTERS.items():
        # Check if BOTH the query and the candidate have a word from the same cluster
        if (q_toks & keywords) and (c_toks & keywords):
            # If they share a concept, apply a significant boost
            return 1.35 
            
    # If no shared cluster is found, do nothing.
    return 1.0

def _cross_domain_penalty(query_text: str, candidate_blob: str) -> float:
    qt = _sector_cues_from_text(query_text)
    ct = _sector_cues_from_text(candidate_blob)
    if not ct: return 1.0
    missing = ct - qt
    mult = 1.0

    # === TUNED VALUES: Penalties are now less severe ===
    if len(ct) >= 1 and len(missing) == len(ct): mult *= 0.85 # Was 0.68
    if len(missing) >= 1 and len(ct) >= 2:        mult *= 0.90 # Was 0.78
    # ===================================================
    
    if (("retail" in qt) or ("logistics" in qt)) and (("engineering" in ct) or ("ict" in ct)):
        mult *= 0.60
    if ("managers" in ct) and not (("managers" in qt) or ("education" in qt) or ("finance" in qt) or ("engineering" in qt)):
        mult *= 0.92
    return mult
_GUARDED_SECTORS = {"healthcare":0.45,"education":0.55,"hospitality":0.55,"diplomatic":0.35,"security":0.55,"arts_media":0.60, "travel":0.50}

def _sector_guard_penalty(query_text: str, candidate_text: str) -> float:
    mult = 1.0
    q = _normalize(query_text)
    c = _normalize(candidate_text)
    qsecs = _sector_cues_from_text(q)
    csecs = _sector_cues_from_text(c)

    if "healthcare" in csecs and "healthcare" not in qsecs:   mult *= _GUARDED_SECTORS["healthcare"]
    if "education"  in csecs and "education"  not in qsecs:   mult *= _GUARDED_SECTORS["education"]
    if "hospitality" in csecs and "hospitality" not in qsecs: mult *= _GUARDED_SECTORS["hospitality"]
    if "security"   in csecs and "security"   not in qsecs: mult *= _GUARDED_SECTORS["security"]
    if "arts_media" in csecs and "arts_media" not in qsecs: mult *= _GUARDED_SECTORS["arts_media"]
    if "travel"     in csecs and "travel"     not in qsecs: mult *= _GUARDED_SECTORS["travel"]

    if ("casino" in c or "gaming" in c) and not _GAMING_CUES.search(q):               mult *= 0.05
    if ("sports" in c or "sport " in c or "sports centre" in c) and not _SPORTS_CUES.search(q): mult *= 0.10
    if ("marketing" in c or "brand" in c or "advertis" in c) and not _MARKETING_CUES.search(q): mult *= 0.15
    return mult
    
def _title_duty_coherence_penalty(title_text: str, duties_text: str, candidate_uses_title_more: bool) -> float:
    tt = set(_tokens(title_text)); dt = set(_tokens(duties_text))
    if not tt or not dt: return 1.0
    inter = len(tt & dt) / float(min(len(tt), len(dt)))
    if candidate_uses_title_more and inter < 0.08:
        return 0.78
    return 1.0

def _role_anchor_boost(query_text: str, candidate_title_and_blob: str) -> float:
    overlap = _role_anchor_overlap(query_text, candidate_title_and_blob)
    q_tokens = set(_tokens(query_text))
    admin_hint = any(w in q_tokens for w in ["admin","administrative","executive","clerk","coordinator"])
    if overlap >= 3: return 1.10
    if overlap == 2: return 1.08
    if overlap == 1: return 1.05
    if admin_hint:   return 1.03
    return 1.00 

# ---- Occupation Group (Annex D) support -------------------------------------
_OCC_GROUP_ALIASES = {
    "legislators": "1", "senior officials": "1", "managers": "1",
    "professionals": "2",
    "associate professionals": "3", "technicians": "3",
    "clerical": "4", "clerical support workers": "4",
    "service": "5", "sales": "5", "service and sales": "5",
    "agricultural": "6", "agriculture": "6", "fishery": "6",
    "craftsmen": "7", "related trades": "7", "craft": "7",
    "plant": "8", "machine operators": "8", "assemblers": "8",
    "cleaners": "9", "labourers": "9", "laborers": "9", "related workers": "9",
    "armed forces": "X", "foreign diplomatic": "X", "diplomatic": "X", "personnel": "X",
}

def _parse_occ_group_hint(s: str) -> Optional[str]:
    """
    Returns: '1'..'9' or 'X' (armed/diplomatic) or None if unusable.
    Accepts forms like '2. Professionals', 'Professionals', 'Group 5', etc.
    """
    if not s:
        return None
    t = _normalize(str(s))
    if not t:
        return None
    m = re.match(r"^(\d{1,2})\b", t)
    if m:
        n = m.group(1)
        if n in set(list("123456789")):
            return n
        if n == "10":
            return "X"
    for key, val in _OCC_GROUP_ALIASES.items():
        if key in t:
            return val
    return None

def _group_hint_multiplier(cand_code: str, group_hint: Optional[str]) -> float:
    """Soft guard: if hint exists, boost matching major group; downweight others."""
    if not group_hint:
        return 1.0
    if group_hint == "X":
        return 1.0  # advisory only; X-codes handled via _xcode_checker
    if not cand_code or not cand_code[:1].isdigit():
        return 1.0
    maj = cand_code[0]
    if maj == group_hint:
        return 1.12
    else:
        return 0.55

# ---------- extra disambiguation multipliers ----------
def _candidate_context_multiplier(rec_code: str, title: str, duties: str) -> float:
    q = _normalize(f"{title} {duties}")
    mult = 1.0

    # === NEW: Machine Operator Disambiguation Penalty ===
    q_toks = set(_tokens(q))
    # If the job is clearly about CNC/machining, penalize sewing-related codes.
    if not q_toks.isdisjoint(HIGH_VALUE_KEYWORDS - {"sewing", "textile"}): # Check for non-textile tech words
        if rec_code == "81531": # Sewing machine operator
            mult *= 0.05 # Apply a 95% penalty

    # If the job is clearly about sewing/textiles, penalize machining-related codes.
    if not q_toks.isdisjoint(_TEXTILE_MACHINE_KEYWORDS):
        if rec_code == "72231": # Machine-tool setter-operator
            mult *= 0.05 # Apply a 95% penalty
    # ===================================================

    if "client" in q or "represent" in q:
        # Boost "Lawyer" if client is mentioned
        if rec_code == "26111":
            mult *= 1.30
        # Penalize "Judge" if client is mentioned
        if rec_code == "26121":
            mult *= 0.20
    if rec_code == "74110":
        if _ELECTRICIAN_CUES.search(q) or "electrician" in q:
            mult *= 1.25
    if rec_code == "71322":
        if _VEHICLE_CUES.search(q): 
            mult *= 1.12
        if _BUILDING_PAINT_CUES.search(q) and not _VEHICLE_CUES.search(q): 
            mult *= 0.25
    if rec_code == "71311":
        if _BUILDING_PAINT_CUES.search(q): 
            mult *= 1.15
        if _VEHICLE_CUES.search(q): 
            mult *= 0.60
    if rec_code == "93100":
        if _CONSTRUCTION_LABOUR_CUES.search(q) or "construction worker" in q: 
            mult *= 1.25
    if rec_code == "71331" and _CONSTRUCTION_LABOUR_CUES.search(q) and "maintenance" not in q:
        mult *= 0.45

    # include 31182 among drafter boosts
    if rec_code in {"31184","31183","31182","31181"} and _DRAFTER_CUES.search(q):
        mult *= 1.20

    if any(k in q for k in ["managing director","chief operating officer"," coo "]):
        if rec_code in {"11201","11203"}: mult *= 1.18
        if re.match(r"265\d{2}", rec_code) and not _ARTS_MEDIA_CUES.search(q): mult *= 0.20
        if rec_code == "14321" and not _GAMING_CUES.search(q): mult *= 0.25

    if "operations manager" in q:
        if rec_code == "13299" and not (_GAMING_CUES.search(q) or _HOSPITALITY_CUES.search(q) or _ARTS_MEDIA_CUES.search(q)):
            mult *= 1.12

    if "office manager" in q:
        if rec_code == "12112" and (_ADMIN_MANAGER_CUES.search(q) or "oversee" in q or "manage" in q): mult *= 1.18
        if rec_code == "41101": mult *= 0.60

    if ("mechanical technician" in q):
        if rec_code == "31151": mult *= 1.15
        if rec_code == "72310" and not _VEHICLE_CUES.search(q): mult *= 0.50

    if ("storekeeper" in q) or ("store keeper" in q) or ("storeman" in q):
        if rec_code in {"43212","43211"}: mult *= 1.25
        if rec_code.startswith("31") or rec_code.startswith("21"): mult *= 0.55

    if "workshop supervisor" in q or ("workshop" in q and "supervisor" in q):
        if rec_code == "72000": mult *= 1.15
        if re.match(r"5150[1-5]", rec_code): mult *= 0.25

    if "project officer" in q:
        if rec_code == "24213" and not (_ATTRACTIONS_CUES.search(q) or "park" in q): mult *= 1.15
        if rec_code == "31603" and not (_ATTRACTIONS_CUES.search(q) or "park" in q): mult *= 0.40

    if "engineer" in q and (_CONSTRUCTION_LABOUR_CUES.search(q) or _DRAFTER_CUES.search(q) or "civil" in q or "structural" in q):
        if rec_code in {"21421","21422"}: mult *= 1.20
        if rec_code == "21497": mult *= 0.30

    if "project manager" in q:
        if rec_code == "13299": mult *= 1.15
        if rec_code == "14310" and not _SPORTS_CUES.search(q): mult *= 0.20

    if "manager" in q and _CONSTRUCTION_LABOUR_CUES.search(q):
        if rec_code == "13299": mult *= 1.12

    if ("site supervisor" in q) or ("supervisor" in q and _CONSTRUCTION_LABOUR_CUES.search(q)):
        if rec_code == "83000": mult *= 1.20
        if "casino" in q or "gaming" in q:
            pass
        else:
            if re.match(r"\b51702\b", rec_code): mult *= 0.10

    if rec_code == "51702" and not (_GAMING_CUES.search(q) or "casino" in q):  mult *= 0.05
    if rec_code == "14310" and not _SPORTS_CUES.search(q):                      mult *= 0.10

    return mult

def _title_sector_conflict_penalty(title_text: str, rec_blob: str) -> float:
    ts = _sector_cues_from_text(title_text)
    cs = _sector_cues_from_text(rec_blob)
    if not ts or not cs: return 1.0
    if ts.isdisjoint(cs):
        heavy_title = {"arch_design","ict","engineering","managers"}
        sensitive_rec = {"fnb","hospitality","arts_media"}
        if (ts & heavy_title) and (cs & sensitive_rec):
            return 0.45
        return 0.70
    return 1.0

def _title_duty_coherence_conflict_penalty(title_text: str, duties_text: str, rec_blob: str) -> float:
    tt = set(_tokens(title_text)); dt = set(_tokens(duties_text))
    if not tt or not dt: return 1.0
    inter = len(tt & dt) / float(min(len(tt), len(dt)))
    if inter < 0.06:
        return 0.85
    return 1.0

# --------- Precomputed scoring (fast path) ----------
def _score_vs_record_precomputed(
    q_title: str, q_duties: str,
    q_norm: str, q_toks: Set[str], q_bis: Set[str],
    rec: Dict[str, str], edu_text_for_row: str,
    group_hint: Optional[str] = None,
    company_industry: str = ""
) -> Tuple[float, str, int]:
    t = rec.get("title","")
    b = rec.get("search_text","")
    
    # === A more intelligent base score calculation ===
    # 1. Calculate a powerful, dedicated score for the title match.
    title_score = _score_title_similarity(q_title, rec)
    
    # 2. Calculate the standard score for the description ("blob").
    b_norm = rec.get("blob_norm", _normalize(b))
    btoks  = rec.get("blob_tokens_set", set(_tokens(b_norm)))
    bbis   = rec.get("blob_bigrams", _bigrams_from_text(b_norm))

    s_diff_blob = _diff_ratio_normed(q_norm, b_norm)
    s_set_blob, s_jac_blob, acts_blob = _overlap_measure_sets(q_toks, btoks)
    s_bi_blob  = _bigram_overlap_sets(q_bis, bbis)
    
    base_blob  = 0.15*s_diff_blob + 0.40*s_set_blob + 0.20*s_jac_blob + 0.25*s_bi_blob
    if acts_blob > 0: base_blob *= (1.0 + min(0.45, 0.20 * acts_blob))

    # 3. Blend the two scores together. We give the title a strong weight (40%).
    base = (title_score * 0.40) + (base_blob * 0.60)
    
    manager_context_score = 0.0
    if "manager" in _tokens(q_title):
        manager_context_score = _score_manager_context(q_title + " " + q_duties, t + " " + b)
        base = (base * 0.5) + (manager_context_score * 0.5)

    engineer_context_score = 0.0
    if "engineer" in _tokens(q_title + " " + q_duties):
        engineer_context_score = _score_engineer_context(q_title + " " + q_duties, t + " " + b)
        base = (base * 0.5) + (engineer_context_score * 0.5)

    # --- All Multipliers and Penalties ---
    mult_sen   = _seniority_penalty(t, q_title + " " + q_duties, b)
    mult_dom   = _cross_domain_penalty(q_title + " " + q_duties, b)
    mult_guard = _sector_guard_penalty(q_title + " " + q_duties, t + " " + b)
    mult_coh   = _title_duty_coherence_penalty(q_title, q_duties, title_score > base_blob * 1.06)
    mult_role  = _role_anchor_boost(q_title + " " + q_duties, t + " " + b)
    ctx_mult   = _candidate_context_multiplier(rec.get("code",""), q_title, q_duties)
    ts_conf    = _title_sector_conflict_penalty(q_title, b)
    td_conf    = _title_duty_coherence_conflict_penalty(q_title, q_duties, b)
    grp_mult   = _group_hint_multiplier(rec.get("code",""), group_hint)
    industry_mult = _get_industry_multiplier(company_industry, rec)
    mult_title_seniority = _title_seniority_conflict_penalty(q_title, q_duties, t + " " + b)
    cluster_boost = _get_cluster_boost(q_title + " " + q_duties, t + " " + b)
    sub_penalty = _subordinate_context_penalty(rec.get("code", ""), q_title + " " + q_duties)
    machine_op_penalty = _machine_operator_context_penalty(rec.get("code", ""), q_title + " " + q_duties)
    discipline_penalty = _engineering_discipline_penalty(rec.get("code", ""), q_title + " " + q_duties)
    spec_man_penalty = _specific_manager_penalty(rec.get("code", ""), q_title + " " + q_duties)
    corp_mgr_penalty = _corporate_manager_penalty(rec.get("code", ""), q_title + " " + q_duties, q_title)
    drafter_handler = _drafter_discipline_handler(rec.get("code", ""), q_title + " " + q_duties)
    design_penalty = _design_discipline_penalty(rec.get("code", ""), q_title + " " + q_duties)
    safety_handler = _safety_discipline_handler(rec.get("code", ""), q_title + " " + q_duties)
    machinist_penalty = _machinist_drafter_penalty(rec.get("code", ""), q_title + " " + q_duties)
    marine_penalty = _marine_context_penalty(rec, q_title + " " + q_duties)
    
    score = (base * mult_sen * mult_dom * mult_guard * mult_coh * mult_role * ctx_mult * ts_conf * td_conf * grp_mult * industry_mult * mult_title_seniority * cluster_boost * sub_penalty * machine_op_penalty * discipline_penalty * spec_man_penalty * corp_mgr_penalty * drafter_handler * design_penalty * safety_handler * machinist_penalty * marine_penalty)
    
    action_hits = acts_blob

    explain = (
        f"ts_score={title_score:.2f} blob_score={base_blob:.2f} | "
        f"sen={mult_sen:.2f} dom={mult_dom:.2f} guard={mult_guard:.2f} coh={mult_coh:.2f} role={mult_role:.2f} "
        f"ctx={ctx_mult:.2f} tsec={ts_conf:.2f} td={td_conf:.2f} grp={grp_mult:.2f} ind={industry_mult:.2f} "
        f"tsen={mult_title_seniority:.2f} clust={cluster_boost:.2f} sub={sub_penalty:.2f} "
        f"mop={machine_op_penalty:.2f} edp={discipline_penalty:.2f} smp={spec_man_penalty:.2f} cmp={corp_mgr_penalty:.2f} ddh={drafter_handler:.2f} "
        f"ddp={design_penalty:.2f} sdh={safety_handler:.2f} mdp={machinist_penalty:.2f} mar={marine_penalty:.2f} "
        f"mgr_ctx={manager_context_score:.2f} eng_ctx={engineer_context_score:.2f} acts={action_hits}"
    )
    return score, explain, action_hits


# ---------- TF-IDF shortlist ----------
_TFIDF_VECT  = None
_TFIDF_MAT   = None
_TFIDF_TEXTS = None

def _tfidf_topk_indices(query_text: str, K: int = 150) -> Optional[List[int]]:
    if not _HAS_SK or _TFIDF_VECT is None or _TFIDF_MAT is None:
        return None
    q = _normalize(query_text)
    if not q:
        return None
    qv = _TFIDF_VECT.transform([q])
    sims = cosine_similarity(qv, _TFIDF_MAT, dense_output=False)
    row = sims.getrow(0)
    if row.nnz == 0:
        return None
    data = row.data
    idxs = row.indices
    if len(data) <= K:
        order = data.argsort()[::-1]
        return idxs[order].tolist()
    import numpy as np
    part = np.argpartition(data, -K)[-K:]
    order = part[np.argsort(data[part])[::-1]]
    return idxs[order].tolist()

# ---------- forced assignment helpers ----------
def _find_best_4_digit_parent(top_5_candidates: List[Dict], all_defs: List[Dict], scorer: Callable) -> Optional[Tuple[float, Dict, str]]:
    """
    Finds and scores the unique 4-digit parents of the top 5 candidates.
    Returns the best-scoring 4-digit parent if one is found.
    """
    if not top_5_candidates:
        return None

    # Find the unique 4-digit parent codes from the top 5 list
    parent_codes = {rec.get("code", "")[:4] for rec in top_5_candidates if rec.get("code")}
    
    # Get the full definition records for these parent codes
    parent_records = [rec for rec in all_defs if rec.get("code") in parent_codes]

    if not parent_records:
        return None

    # Score each of the 4-digit parent candidates
    best_4d_score, best_4d_rec, best_4d_explain = -1.0, None, ""
    for rec in parent_records:
        s, e, _ = scorer(rec)
        if s > best_4d_score:
            best_4d_score, best_4d_rec, best_4d_explain = s, rec, e
            
    if best_4d_rec:
        return best_4d_score, best_4d_rec, best_4d_explain
    return None

def _best_5digit_from(defs: List[Dict[str,str]], scorer, cand_indices=None):
    it = ((i, defs[i]) for i in cand_indices) if cand_indices is not None else enumerate(defs)
    best_s, best_r, best_e = -1.0, None, ""
    for _, r in it:
        code = r.get("code","")
        if not (code.isdigit() and len(code)==5):
            continue
        s, e, _ = scorer(r)
        if s > best_s:
            best_s, best_r, best_e = s, r, e
    if best_r is None:
        return None
    return best_r.get("code",""), best_r.get("title",""), best_s, best_e

def _pull_down_to_5(prefix_code: str, defs: List[Dict[str, str]], scorer: Callable[[Dict[str,str]], Tuple[float,str,int]]):
    prefix = (prefix_code or "").strip()
    if not prefix: return None
    kids = [r for r in defs if r.get("code","") and r.get("code","").isdigit() and len(r.get("code",""))==5 and r.get("code","").startswith(prefix)]
    if not kids: return None
    best_s, best_r, best_e = -1.0, None, ""
    for r in kids:
        s, e, _ = scorer(r)
        if s > best_s:
            best_s, best_r, best_e = s, r, e
    return (best_s, best_r, best_e)


# ---------- baked-in rule engine ----------
_BAKED_RULES: List[Tuple[re.Pattern, str, str, str, Optional[re.Pattern], bool]] = [
    # Cleaning Roles Section (Order is important: most specific to most general)
    (re.compile(r"(^\s*cl[ea]?[ae]?n[ae]?rs?$|of+ice\s+cl[ea][ea]ner)"), "91131", "Office/Commercial/Industrial establishment indoor cleaner", "rule_cleaner_exact", None, True),
    (re.compile(r"\bhousek[e]+p[ae]?rs?\b"), "51501", "Housekeeper (Private households, hotels and offices)", "rule_housekeeper", None, True),
    (re.compile(r"\bjanitors?\b"), "91299", "Other cleaning worker n.e.c.", "rule_janitor", None, True),
    (re.compile(r"\bgeneral\s+clea?n[ae]?rs?\b"), "91131", "Office/Commercial/Industrial establishment indoor cleaner", "rule_general_cleaner", None, True),
    
    # Human Resources
    (re.compile(r"\b(hr|human\s+resou?rce|human\s+capital)\s+(directors?|heads?|man[ae]?gers?)\b"), "12121", "Human resource manager", "rule_hr_manager", None, True),
    (re.compile(r"\b(recru[ia]ters?|recru[ia]tment|talent\s+acquisition)\b"), "24231", "Recruiter/Talent acquisition specialist", "rule_recruiter", None, True),
    (re.compile(r"\bpayroll\s+(specialists?|of+icers?|exec(utive)?s?)\b"), "41102", "Payroll officer", "rule_payroll_officer", None, True),
    (re.compile(r"\bpayroll\s+cl[ei]rks?\b"), "41102", "Payroll clerk", "rule_payroll_clerk", None, True),
    (re.compile(r"\b(hr|human\s+resou?rce)\s+(generalists?|business\s+partn[ae]?rs?|bp)\b"), "41102", "Human resource Clerk", "rule_hr_generalist_bp", None, True),
    (re.compile(r"\b(hr|human\s+resou?rce|human\s+capital)\s+(as+is+t[ae]nts?|exec(utive)?s?|co?ordinators?)\b"), "12121", "Human resource manager", "rule_hr_assistant_exec", None, True),
    (re.compile(r"^\s*(hr|human\s+resou?rce)\s*$"), "41102", "Human resource Clerk", "rule_hr_exact", None, True),

    # Education
    (re.compile(r"\b(chinese|mandarin|english|malay|tamil|japanese|korean|french|german|spanish)\s+(teach[ae]?rs?|tutors?|instructors?)\b"), "36201", "Language instructor (extracurriculum)", "rule_specific_language_teacher", None, True),
    (re.compile(r"\b(teach[ae]?rs?|tutors?)\b.*\blan?gu[ae]ge\b|\blan?gu[ae]ge\b.*\b(teach[ae]?rs?|tutors?)\b"), "36201", "Language instructor (extracurriculum)", "rule_language_teacher", None, True),
    
    # Retail / sales / customer service
    (re.compile(r"\b(reta?il|shops?|sto?res?)\s+man[ae]?gers?\b"), "14201", "Retail/Shop manager", "rule_retail_manager", None, True),
    (re.compile(r"\bcash[i][e]rs?\b"), "52302", "Cashier (general)", "rule_cashier", None, True),
    (re.compile(r"\b(sales?\s*(p[ae]rsons?|cl[ei]rks?|as+is+t[ae]nts?|executive|exec)|retail\s+sales?|shops?\s+as+is+t[ae]nts?)\b"), "52202", "Shop sales assistant", "rule_sales_person_clerk_asst", None, True),
    (re.compile(r"\b(sales?\s+sup[eia]rvisors?|flo?o?r\s+man[ae]?gers?)\b"), "52201", "Sales supervisor", "rule_retail_sales_sup", None, True),
    (re.compile(r"\b(custom[ae]?rs?\s*(s[eia]rvices?|sup+o?rt)|client\s+relations)\b"), "42245", "Customer service representative", "rule_customer_service_support", None, True),
    (re.compile(r"\bflyer\b"), "96291", "Leaflet/Newspaper distributor/deliverer", "rule_flyer_distributor", None, True),
    (re.compile(r"\bmerchandi?[sz]er\b"), "33225", "Merchandising/Category executive", "rule_mercahndiser", None, True),
    
    # tailor
    (re.compile(r"\b(tailors?|dres+mak[ae]?rs?)\b"), "75310", "Tailor/Dressmaker", "rule_tailor_dressmaker", None, True),

    # Logistics / store / warehouse
    (re.compile(r"\b(stores?\s*?k[e]+p[ae]?r|sto?re\s*man|sto?re\s*men)\b"), "43212", "Storekeeper", "rule_storekeeper", None, True),
    (re.compile(r"\b(warehouse\s+(as+is+t[ae]nts?|op[ae]?rators?|pick[ae]?rs?|pack[ae]?rs?))\b"), "93201", "Hand packer", "rule_warehouse_asst", None, True),
    (re.compile(r"\blogistics?\s+co?ordinators?\b"), "33461", "Logistics/Production planner", "rule_logistics_coord", None, True),
    (re.compile(r"\bpack[ae]?rs?\b"), "93201", "Hand packer", "rule_packer", None, True),
    (re.compile(r"\bgeneral\s+work[ae]?rs?\b"), "96293", "Odd job person", "rule_general_worker", None, True),
    (re.compile(r"\bfork\s*lifts?\b"), "83441", "Fork lift truck operator", "rule_forklift", None, False),

    # Drivers
    (re.compile(r"\b(bus|coach)\s+driv[ae]?rs?\b"), "83311", "Bus driver", "rule_bus_driver", None, True),
    (re.compile(r"\b(lor+y|truck)\s*driv[ae]?rs?\b"), "83321", "Lorry/Truck driver", "rule_lorry_truck_driver", None, True),
    (re.compile(r"\bdeliv[ae]?ry\s*driv[ae]?rs?\b"), "83229", "Car/Taxi/Van/Light goods vehicle driver n.e.c.", "rule_delivery_driver", None, True),
    (re.compile(r"^\s*driv[ae]?rs?\s*$"), "8322", "Car/Taxi/Van/Light goods vehicle driver", "rule_driver_exact", None, True),
    
    # Construction site roles
    (re.compile(r"\bcons?tr?ucti?on\s+work[ae]?rs?\b"), "93100", "Civil engineering/Building construction labourer", "rule_construction_worker", None, True),
    (re.compile(r"\bsite\s+co?ordinators?\b"), "31124", "Resident technical officer", "rule_site_coord", re.compile(r"\b(work\s*site|constru[ck]tion|building|engineering)\b"), True),
    (re.compile(r"site\s+sup[eia]rvisors?"), "71000", "Supervisor/General foreman (building and related trades)", "rule_site_supervisor", None, True),
    (re.compile(r"\b(sup[eia]rvisors?\s*cum\s*general\s*fo?rem[ae]n|fo?rem[ae]n)\b"), "71000", "Supervisor/General foreman (building and related trades)", "rule_foreman", None, True),
    (re.compile(r"\bsaf[ae]?ty\s+(of+icers?|sup[eia]rvisors?)\b"), "21493", "Industrial safety engineer", "rule_safety", None, True),
    (re.compile(r"\bquan?tity\s*s[ue]rv[ae]yors?\b"), "21494", "Quantity surveyor", "rule_qs", None, True),
    (re.compile(r"\bexc[ae]v[ae]to?rs?\s+op[ae]?rat?ors?\b"), "83421", "Excavating/Trench digging machine operator", "rule_excavator_operator", None, True),
    (re.compile(r"\bpumps?\s+op[ae]?rat?ors?\b"), "31153", "Machining/Tooling technician", "rule_pump_operator", re.compile(r"\b(plant|machine|industrial|construction|site|water|chemical)\b"), False),
    (re.compile(r"(quality|qa|qc)"), "21414", "Quality control/assurance engineer", "rule_QA_engineer", None, True),
    (re.compile(r"\bsafety\b"), "21493", "Industrial safety engineer", "rule_safety_engineer", None, True),
    

    # Construction Trades
    (re.compile(r"\bcarp[ae]nt[ae]?rs?\b"), "71151", "Carpenter", "rule_carpenter", None, True),
    (re.compile(r"\bplumb[ae]?rs?\b"), "71261", "Plumber", "rule_plumber", None, True),
    (re.compile(r"\bplast[ae]?r[ae]?rs?\b"), "71230", "Plasterer", "rule_plasterer", None, True),
    (re.compile(r"\btil[ae]?rs?\b"), "71220", "Tiler", "rule_tiler", None, True),
    (re.compile(r"\bglaz[ie]rs?\b"), "71250", "Glazier", "rule_glazier", None, True),
    (re.compile(r"\bweld[ae]?rs?\b"), "72120", "Welder", "rule_welder", None, True),
    (re.compile(r"\b(scaf+old[ae]?rs?|scaf+old)\b"), "71191", "Scaffolder", "rule_scaffolder", None, False),
    (re.compile(r"\b(^architect$)\b"), "21610", "Building architect", "rule_architect", None, True),
    (re.compile(r"\blifti?ng\b"), "83431", "Crane/Hoist operator (excluding port)", "rule_lifting_crane_supervisor", None, True),
    
    # Generic Engineers
    (re.compile(r"\b(senior|principal|lead|exec(utive)?s?)\s+eng?in[e]+rs?\b"), "21499", "Engineering professional n.e.c.", "rule_senior_exec_engineer", None, True),
    (re.compile(r"\b(sales?\s+eng?in[e]+rs?|techn?i?c?al\s+sales?)\b"), "24331", "Technical sales professional", "rule_sales_engineer", None, True),
    (re.compile(r"\bmechanics?\b"), "72310", "Automotive mechanic", "rule_vehicle_mechanic", re.compile(r"\b(auto|car|vehicle|motor|automotive|tyre|battery)\b"), False),
    (re.compile(r"cnc"), "31153", "Machining/Tooling technician", "rule_cnc", None, True),
    
    # software
    (re.compile(r"\b(ui|ux)\b"), "25124", "Interaction designer", "rule_UI|UX_designer", None, True),
    
    # Electrical / technicians / engineers
    (re.compile(r"\bel[e]?ctr[ia]ci?ans?\b"), "74110", "Electrician", "rule_electrician", None, True),
    (re.compile(r"\bm[e]?ch[ae]n?ic?al\s+techn?i?ci?ans?\b"), "31151", "Mechanical engineering technician", "rule_mech_tech", None, True),
    (re.compile(r"\beng?in[e]+ring\s+techn?i?ci?ans?\b"), "31129", "Civil engineering technician n.e.c.", "rule_eng_tech", None, True),
    (re.compile(r"\b(el[e]?ctr[ia]c?al\s+eng?in[e]+rs?|swi?tchbo?a?rds?|po?w[ae]?r\s+s[iy]st[e]?ms?)\b"), "21511", "Electrical engineer", "rule_elec_engineer", None, True),
    (re.compile(r"\b(air-?con|air\s*conditioning)\b"), "71271", "Air-conditioning and refrigeration mechanic", "rule_aircon", None, False),
    (re.compile(r"\bm[e]?ch[ae]n?ica?l?\s+techn?i?ci?ans?\b"), "31151", "Mechanical engineering technician", "rule_mech_tech", None, True),
    (re.compile(r"^techn?i?ci?an$"), "31151", "Mechanical engineering technician", "rule_mech_tech", None, True),
    
    # Healthcare Roles
    (re.compile(r"\b(docto?rs?|ph[iy]sicians?)\b"), "22110", "General practitioner/Physician", "rule_doctor_gp", None, True),
    (re.compile(r"\b(general\s+practition[ae]?rs?)\b"), "22110", "General practitioner/Physician", "rule_doctor_gp", None, True),
    (re.compile(r"\b(registered|staff)\s+nu?rse?s?\b"), "22200", "Nursing professional", "rule_registered_staff_nurse", None, True),
    (re.compile(r"\b(enrolled|assistant)\s+nu?rse?s?\b"), "32200", "Enrolled/Assistant nurse", "rule_enrolled_nurse", None, True),
    (re.compile(r"\b(nursing\s+aides?|healthcare\s+as+is+t[ae]nts?)\b"), "53201", "Healthcare assistant", "rule_nursing_aide", None, True),
    (re.compile(r"(^\s*nurse?s?\s*$|clinic\s+nurse)"), "22200", "Registered nurse and related nursing professional (excluding enrolled nurse)", "rule_nurse_exact", None, True),
    (re.compile(r"clinic\s+(manager|ma+nger|mgr|mgt)"), "13420", "Health services manager", "rule_clinic_manager", None, True),
    (re.compile(r"clinic\s+(as+is+tant|as+t)"), "42243", "Medical/Dental receptionist", "rule_clinic_assistant", None, True),
    
    # Project/ops/coordination
    (re.compile(r"\bpro?j[e]?ct\s+of+icers?\b"), "24213", "Business/Financial project management professional", "rule_project_officer", None, True),
    (re.compile(r"\bpro?j[e]?ct\s+co?ordinators?\b"), "13299", "Other production/operations manager n.e.c.", "rule_project_coord", None, True),
    (re.compile(r"\bpro?j[e]?ct\s+man[ae]?gers?\b"), "13299", "Other production/operations manager n.e.c.", "rule_project_manager", None, True),
    (re.compile(r"\bop[e]?rat?ions?\s+man[ae]?gers?\b"), "13299", "Other production/operations manager n.e.c.", "rule_ops_manager", None, True),
    (re.compile(r"\b^(manager|ma+nger)$\b"), "12112", "Administration manager", "rule_ops_manager", None, True),

    # C-Suite / Director Level
    (re.compile(r"\b(chi[ei]f\s+finan[cs]ial\s+of+icers?|cfo)\b"), "12111", "Budgeting/Financial accounting manager", "rule_cfo", None, True),
    (re.compile(r"\b(chi[ei]f\s+techn?ology\s+of+icers?|cto)\b"), "13301", "Chief information officer/Chief technology officer/Chief information security officer", "rule_cto", None, True),
    (re.compile(r"^\bdirector\b$"), "11201", "Managing director/Chief executive officer", "rule_ceo_md", None, True),
    (re.compile(r"\b(man[ae]?ging\s+directors?|chi[ei]f\s+exec(utive)?s?\s+of+icers?|ceo)\b"), "11201", "Managing director/Chief executive officer", "rule_ceo_md", None, True),
    (re.compile(r"\b(chi[ei]f\s+op[e]?rat?ing\s+of+icers?|coo)\b"), "11203", "Chief operating officer/General manager", "rule_coo_director", None, True),
    (re.compile(r"\bgeneral\s+(ma+nger|man[ae]?gers?)\b"), "11203", "Chief operating officer/General manager", "rule_general_manager", None, True),
    
    # public
    (re.compile(r"\bpubl?ic\s+rea?l+a?ti?ons?\b"), "24320", "24320	Public relations/Corporate communications professional", "rule_public_relations", None, True),
    
    # IT / web
    (re.compile(r"\b(web\s*sites?\s*admin?is?tr[ae]tors?|webm[ae]st[ae]?rs?)\b"), "35140", "Website administrator/Webmaster", "rule_web_admin", None, True),
    (re.compile(r"\b(it\s+sup+o?rts?|(it|techn?i?c?al|comp?ut[ae]?rs?)\s+help\s*desks?|desktops?\s+sup+o?rts?)\b"), "35123", "IT support technician", "rule_it_support", None, True),
    
    # Finance / accounts
    (re.compile(r"\b(exec(utive)?s?\s+ac+ou?nt[ae]nts?|financial\s+exec(utive)?s?)\b"), "24111", "Accountant (General)", "rule_exec_accountant", None, True),
    (re.compile(r"\bac+ou?nts?.*\bexec(utive)?s?\b"), "24111", "Accountant (General)", "rule_accounts_and_other_exec", None, True),
    (re.compile(r"^\s*ac+ou?nt[ae]nts?\s*$"), "24111", "Accountant (General)", "rule_accountant_exact", None, True),
    (re.compile(r"\b(as+is+t[ae]nts?\s*ac+ou?nt[ae]nts?|ac+ou?nts?\s*as+is+t[ae]nts?)\s+exec(utive)?s?\b"), "33131", "Assistant accountant", "rule_asst_accountant_exec", None, True),
    (re.compile(r"\b(as+is+t[ae]nts?\s*ac+ou?nt[ae]nts?|ac+ou?nts?\s*as+is+t[ae]nts?)\b"), "33131", "Assistant accountant", "rule_asst_accountant", None, True),
    (re.compile(r"\bfinan[cs]e\s+man[ae]?gers?\b"), "12111", "Budgeting/Financial accounting manager", "rule_finance_manager", None, True),
    (re.compile(r"\b(ac+ou?nts?|ac+t)\s+man[ae]?gers?\b"), "12211", "Sales manager", "rule_account_manager", None, True),
    (re.compile(r"\b(ac+ou?nti?ng)\s+(man[ae]?gers?|maanger)\b"), "12111", "Budgeting/Financial accounting manager", "rule_accounting_manager", None, True),
    
    # F&B / basic services
    (re.compile(r"\bco+k\b"), "51201", "Cook", "rule_cook", None, True),
    (re.compile(r"\bchef\b"), "34341", "Chef (excluding pastry chef)", "rule_chef", None, True),
    (re.compile(r"\bcof+e+\s*mak[ae]?rs?\b"), "94102", "Food/Drink stall assistant", "rule_coffeemaker_stall", None, True),
    (re.compile(r"\bbaristas?\b"), "51321", "Barista", "rule_barista", None, True),
    (re.compile(r"\b(wait[ae]?rs?|waitres+es?|s[eia]rvice\s+crew)\b"), "51312", "Waiter/Waitress", "rule_waiter", None, True),
    (re.compile(r"\bsec[ue]rity\s+(gua?rds?|of+icers?)\b"), "54144", "Security officer", "rule_security", None, True),
    (re.compile(r"\b(dish\s+washer|dishwasher)\b"), "91153", "Dishwasher", "rule_dishwasher", None, True),
    (re.compile(r"food\s+(proc+es+ing|proc+[eo]s+[eo]r)"), "75190", "Food processing and related trades worker n.e.c.", "rule_food_processor", None, True),
    
    # beauty
    (re.compile(r"\bmas+eur|mas+euse?|mas+age\b"), "32551", "Massage therapist", "rule_massage_therapist", None, True),
    
    # Religious
    (re.compile(r"\b(past[oe]rs?|pr[ie]{2}sts?|rev(er[ae]nd)?|im[ae]ms?|monks?|clerg[ey])\b"), "26369", "Religious professional", "rule_religious_professional", None, True),
    
    # Admin / office (This section is now last)
    (re.compile(r"\b(of+ice|admin|administ(r?)[ai]tive)\s+man[ae]?gers?\b"), "12112", "Administration/Office manager", "rule_admin_office_manager", None, True),
    (re.compile(r"\b(admin|administ(r?)[ai]tive)\s+(exec(utive)?s?|of+icers?)\b"),"41101", "General office clerk", "rule_admin_exec", None, True),
    (re.compile(r"\b(admin|administ(r?)[ai]tive)\s+(as+is+t[ae]nts?)\b"), "41101", "General office clerk", "rule_admin_asst", None, True),
    (re.compile(r"\b(admin|administ(r?)[ai]tive)\s+cl[ei]rks?\b"), "41101", "General office clerk", "rule_admin_clerk", None, True),
    (re.compile(r"\b(administ(r?)[ai]tive|admin)\s+work[ae]?rs?\b"), "41101", "General office clerk", "rule_admin_worker", None, True),
    (re.compile(r"^\s*(administ(r?)[ai]tive|admin)\s+work\s*$"), "41101", "General office clerk", "rule_admin_work_exact", None, True),
    (re.compile(r"^\s*(admin|admin?i?stration|admin?i?strator|admi?ni?strati?on)\s*$"), "41101", "General office clerk", "rule_admin_exact", None, True),
    (re.compile(r"\brec[ei]ptionists?\b"), "42261", "Receptionist", "rule_receptionist", None, True),
    (re.compile(r"\b(admin|administ(r?)[ai]tive)\b"), "41101", "General office clerk", "rule_admin_general_catch", None, True),
    (re.compile(r"^\s*cl[ei]rks?\s*$"), "41101", "General office clerk", "rule_clerk_exact", None, True),
    (re.compile(r"\b(p[ae]?rson[ae]ls?\s+as+is+t[ae]nts?|pa)\b"), "33494", "Executive secretary", "rule_personal_assistant_pa", None, True),
    (re.compile(r"\b(ope?rati?ons|ops)\s+(exe?cu?ti?ve?|exec.?)\b"), "33492", "Operations officer (administrative)", "rule_ops_exec", None, True),
]

def _route_drafter_code(title_text: str, duties_text: str) -> Tuple[str, str]:
    """
    Routes generic drafter roles to the correct subclass using a context-aware check
    to prevent misclassification from ambiguous keywords like "civil".
    """
    q = _normalize(f"{title_text} {duties_text}")

    # === NEW: Context-aware check to prevent ambiguity ===
    # We now check for the primary keyword PLUS a general construction/engineering keyword.
    
    # Check for Civil/Structural context
    has_civil_keyword = "civil" in q or "structural" in q
    has_construction_context = "construction" in q or _CONSTRUCTION_LABOUR_CUES.search(q) or "building" in q or "engineering" in q
    if has_civil_keyword and has_construction_context:
        return "31183", "Civil/Structural engineering draughtsperson"

    # Check for Electrical context
    has_electrical_keyword = "electrical" in q or "power" in q or "switchboard" in q
    if has_electrical_keyword and ("engineering" in q or _ELECTRICIAN_CUES.search(q)):
        return "31182", "Electrical/Electronics draughtsperson"

    # Check for Mechanical context
    has_mechanical_keyword = "mechanical" in q or "piping" in q or "hvac" in q or "acmv" in q
    if has_mechanical_keyword and ("engineering" in q or "workshop" in q):
        return "31181", "Mechanical draughtsperson"

    # Check for Architectural context
    has_architectural_keyword = "architectural" in q or "floor plan" in q or "elevation" in q or "interior" in q or "architect" in q
    if has_architectural_keyword: # This one is less ambiguous and can stand alone
        return "31184", "Architectural draughtsperson"
        
    # If no specific discipline with context is found, default to the general code.
    return "31189", "Draughtsperson n.e.c."

def _apply_baked_rules(title_text: str, duties_text: str) -> Optional[Tuple[str, str, str]]:
    """
    Returns (code, title, reason) if any baked rule matches.
    Can check title-only or combined text based on the rule's flag.
    """
    # === CORRECTED TYPO HERE ===
    t_norm = _normalize(title_text or "")
    # ===========================
    d_norm = _normalize(duties_text or "")
    both_norm = f"{t_norm} {d_norm}".strip()

    # The last item in the tuple is our new title_only_check flag
    for rx, code, occ_title, reason, cue, title_only_check in _BAKED_RULES:
        
        text_to_search = t_norm if title_only_check else both_norm
        
        if rx.search(text_to_search):
            if cue is not None and not cue.search(both_norm):
                continue
            
            if "rule_admin_exec" in reason and ("executive" not in both_norm and "exec" not in both_norm):
                continue
                
            return code, occ_title, reason

    # === MODIFIED: Conditional permutations now check the TITLE for the main role keyword ===
    # This prevents misclassifying junior roles that mention a senior title in their duties.

    if ("supervisor" in t_norm) and _CONSTRUCTION_LABOUR_CUES.search(both_norm):
        return "83000", "Mobile machinery supervisor/general foreman", "rule_supervisor_construction"

    if ("manager" in t_norm) and _CONSTRUCTION_LABOUR_CUES.search(both_norm):
        return "13299", "Other production/operations manager n.e.c.", "rule_manager_construction"

    if (("engineer" in t_norm) and 
        (_DRAFTER_CUES.search(both_norm) or "layout" in both_norm or "technical drawing" in both_norm) and 
        (_CONSTRUCTION_LABOUR_CUES.search(both_norm) or "civil" in both_norm or "structural" in both_norm or "construction" in both_norm)):
        return "21421", "Civil engineer", "rule_engineer_drawings_construction"

    #if _DRAFTER_CUES.search(both_norm) and "engineer" not in t_norm:
        #code, occ_title = _route_drafter_code(title_text, duties_text)
        #return code, occ_title, "rule_drafter_routed"

    #if ("supervisor" in t_norm) and not _GAMING_CUES.search(both_norm):
    #   return "13299", "Other production/operations manager n.e.c.", "rule_supervisor_generic"
    
    if "painter" in both_norm:
        # First, check for any vehicle-related keywords in the combined title and duties.
        if _VEHICLE_CUES.search(both_norm):
            return "71322", "Motor vehicle spray painter", "rule_painter_vehicle_context"
        # If no vehicle cues are found, it's safe to default to a building painter.
        else:
            return "71311", "House/Building painter", "rule_painter_building_default"
    
    return None

# ---------- main matcher ----------
FORCE_ASSIGN = True  # Always avoid X1000 by default

def best_match_duties_priority(title_text: str, duties_text: str, defs: List[Dict[str, str]],
                               title_map: Dict[str, Dict], expert_map: Dict[str, Tuple[str, str]], 
                               min_score_0_to_1: float, edu_text_for_row: str,
                               occ_group_hint_raw: Optional[str] = None,
                               company_industry: str = ""):
    duties_text = (duties_text or "").strip()
    title_text  = (title_text or "").strip()
    norm_title = _normalize(title_text)
    group_hint = _parse_occ_group_hint(occ_group_hint_raw or "")

    # === NEW ORDER OF OPERATIONS ===

    # 1. HIGHEST PRIORITY (As requested): Your hand-crafted Baked-in Rules.
    br = _apply_baked_rules(title_text, duties_text)
    if br:
        code, occ_title, reason = br
        return code, occ_title, 0.66, f"{reason}", [], "Baked-in Rule"

    # 2. SECOND PRIORITY (As requested): The validated Expert Map.
    expert_match = _validate_expert_match(title_text, duties_text, expert_map, defs)
    if expert_match:
        code, title = expert_match
        return code, title, 1.0, "Expert Map Match (Validated)", [], "Expert Map Match"

    # 3. THIRD PRIORITY (As requested): Automated Exact Title Match.
    if norm_title in title_map:
        rec = title_map[norm_title]
        if rec.get("is_5d"):
            return rec["code"], rec["title"], 1.0, "Exact Title Match", [], "Exact Title Match"
    # =================================

    # Handle X-Codes and empty inputs
    xhit = _xcode_checker(duties_text, title_text)
    if xhit is not None and xhit[0] in ("X3000","X4000","X5000"):
        return xhit[0], xhit[1], 1.0, "xcode-direct", [], "X-Code"
    if not duties_text and not title_text:
        return "X2000", X_TITLE_MAP["X2000"], 1.0, "xcode-noinfo", [], "No Input"

    # 4. FINAL FALLBACK: The full scoring engine
    q_text = f"{title_text} {duties_text}".strip()
    search_type_label = "Title + Duties (Combined)"
    q_norm = _normalize(q_text)
    q_toks = set(_tokens(q_norm))
    q_bis = _bigrams_from_text(q_norm)
        
    def scorer(rec: Dict[str,str]):
        return _score_vs_record_precomputed(title_text, duties_text, q_norm, q_toks, q_bis, rec, edu_text_for_row, group_hint, company_industry)

    five_digit_candidates = [r for r in defs if r.get("is_5d")]
    cand_indices = _tfidf_topk_indices(q_text, K=150)
    cand_iter_5d = [defs[i] for i in cand_indices if defs[i].get("is_5d")] if cand_indices is not None else five_digit_candidates

    all_candidates = []
    best_s, best_r, best_e = -1.0, None, ""
    for r in cand_iter_5d:
        s, e, _ = scorer(r)
        all_candidates.append({"score": s, "code": r.get("code", ""), "title": r.get("title", ""), "explain": e})
        if s > best_s:
            best_s, best_r, best_e = s, r, e

    all_candidates.sort(key=lambda x: x["score"], reverse=True)
    top_5 = all_candidates[:5]
    
    if best_r is None:
        return "X1000", X_TITLE_MAP["X1000"], 0.0, "xcode-no-match-found", top_5, search_type_label
    
    final_code, final_title, final_score, final_explain, final_search_type = best_r["code"], best_r["title"], best_s, best_e, search_type_label

    accept_bar = min_score_0_to_1
    if best_s < accept_bar:
        parent_result = _find_best_4_digit_parent(top_5, defs, scorer)
        if parent_result:
            best_4d_score, best_4d_rec, best_4d_explain = parent_result
            if best_4d_score >= accept_bar:
                final_code, final_title, final_score, final_explain, final_search_type = best_4d_rec["code"], best_4d_rec["title"], best_4d_score, best_4d_explain, "4-Digit Fallback"

    if final_score < min_score_0_to_1 and not final_code.startswith("X"):
        final_code = "X1000"
        final_title = X_TITLE_MAP["X1000"]
    
    return final_code, final_title, final_score, final_explain, top_5, final_search_type
# ---------- write back ----------
from openpyxl import load_workbook

def _write_back_copy(jobs_path: str, sheet, header_row: int,
                     code_col_name: str, title_col_name: str,
                     row_indices: List[int], codes: List[str], titles: List[str],
                     out_path: Optional[str]) -> str:
    save_path = out_path
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    shutil.copyfile(jobs_path, save_path)

    wb = load_workbook(save_path)
    ws = wb[wb.sheetnames[0]] if sheet is None else (wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet])

    # === NEW: Directly target columns AQ (43) and AR (44) ===
    hdr_row = header_row + 1
    code_col_idx = 43  # Column AQ
    title_col_idx = 44 # Column AR

    # For clarity, let's ensure the headers are written correctly in these specific columns.
    # This will overwrite any existing header in AQ5/AR5 or create it if it's missing.
    ws.cell(row=hdr_row, column=code_col_idx, value="SSOC Code (AQ)")
    ws.cell(row=hdr_row, column=title_col_idx, value="SSOC Title (AR)")
    # ========================================================

    for r0, c, t in zip(row_indices, codes, titles):
        r = r0 + 1
        # Write data to the specific column index
        ws.cell(row=r, column=code_col_idx,  value=(c if c else None))
        ws.cell(row=r, column=title_col_idx, value=(t if t else None))

    wb.save(save_path)
    return save_path

def _write_detailed_report(report_data: List[Dict], uen: str, original_stem: str, out_dir: str, timestamp: str):
    """Writes the detailed top-5 scoring report to a 'detailed_reports' subfolder."""
    if not report_data:
        print("[INFO] No data for detailed report, skipping.")
        return
    
    report_folder = os.path.join(out_dir, "detailed_reports")
    os.makedirs(report_folder, exist_ok=True)

    # Use our new helper to generate the path
    report_path = _generate_output_path(report_folder, uen, original_stem, "scoring_report", timestamp)
    
    try:
        df = pd.DataFrame(report_data)
        df.to_excel(report_path, index=False)
        print(f"      Detailed report -> detailed_reports\\{os.path.basename(report_path)}")
    except Exception as e:
        print(f"[WARN] Failed to write detailed scoring report: {e}", file=sys.stderr)

# ---------- batch helpers ----------
def _timestamp() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")

def _build_out_dir(base_dir: str, user_out_dir: Optional[str]) -> str:
    if user_out_dir:
        os.makedirs(user_out_dir, exist_ok=True)
        return user_out_dir
    
    parent_dir = os.path.dirname(base_dir)
    out_dir = os.path.join(parent_dir, f"ssoc_outputs_{_timestamp()}")
    os.makedirs(out_dir, exist_ok=True)
    return out_dir

def _generate_output_path(out_dir: str, uen: str, original_stem: str, suffix: str, timestamp: str) -> str:
    """Generates a clean output path using the UEN if available, otherwise falls back to the original filename."""
    # Use the UEN for the filename if it's a valid string, otherwise use the original filename
    filename_base = uen if uen and len(uen) > 5 else original_stem
    
    # Sanitize the base name to remove characters invalid for filenames
    filename_base = re.sub(r'[\\/*?:"<>|]', '_', filename_base)
    
    file_name = f"{filename_base}_{suffix}_{timestamp}.xlsx"
    return os.path.join(out_dir, file_name)

def _list_jobs_files(jobs_dir: str, recursive: bool, defs_path: Optional[str]) -> List[str]:
    patterns = ["*.xlsx", "*.xls"]  # add "*.xlsm" if needed
    files = []
    for pat in patterns:
        pattern = os.path.join(jobs_dir, "**", pat) if recursive else os.path.join(jobs_dir, pat)
        files.extend(glob.glob(pattern, recursive=recursive))
    defs_real = os.path.realpath(defs_path) if defs_path else None
    cleaned = []
    for f in files:
        if os.path.basename(f).startswith("~$"):
            continue
        if defs_real and os.path.realpath(f) == defs_real:
            continue
        cleaned.append(f)
    return sorted(set(cleaned))

# ---- skip unreadable probe (optional; default ON via CLI) ----
def _is_readable_excel(path: str) -> bool:
    ext = os.path.splitext(path.lower())[1]
    try:
        if ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            wb.close()
            return True
        if ext == ".xls":
            import xlrd
            xlrd.open_workbook(path, on_demand=True)
            return True
        return False
    except Exception:
        return False

# ---------- NEW: allow DEFAULT_JOBS_FILE (or --jobs) to be a folder ----------
def _resolve_jobs_input(args):
    """
    If --jobs (or DEFAULT_JOBS_FILE) is a directory, treat it as --jobs-dir.
    Otherwise keep it as a single file.
    """
    if args.jobs and os.path.isdir(args.jobs):
        args.jobs_dir, args.jobs = args.jobs, None

    if not args.jobs and not args.jobs_dir:
        df = DEFAULT_JOBS_FILE
        if os.path.isdir(df):
            args.jobs_dir = df
        else:
            args.jobs = df

# ---------- main (single-file processor to reuse in batch) ----------
def process_single_file(jobs_path: str, defs: List[Dict[str, str]], title_map: Dict[str, Dict],
                        expert_map: Dict[str, Tuple[str, str]], uen_to_ssic_map: Dict[str, str], 
                        ssic_definitions: Dict[str, str], args, out_dir: str) -> Tuple[str, Optional[str], int, int]:
    try:
        row_idxs, titles, duties, edus, groups, uen_for_file = load_jobs_separate(
            jobs_path, args.jobs_sheet, args.jobs_header_row,
            args.title_col_name, args.edu_col_name,
            args.title_col_index, args.edu_col_index,
            debug=args.debug
        )
    except Exception as e:
        print(f"[ERROR] {os.path.basename(jobs_path)}: Error loading jobs: {e}", file=sys.stderr)
        return "", None, 0, 0
        
    audit_path = None

    if not (titles or duties):
        print(f"[WARN] {os.path.basename(jobs_path)}: No job rows found; check sheet/header/columns.", file=sys.stderr)
        return "", None, 0, 0
    
    ts = _timestamp() # Keep timestamp for detailed report if needed
    original_stem = os.path.splitext(os.path.basename(jobs_path))[0]
    
    min_s = max(0.0, min(1.0, args.min_score/100.0))
    
    company_ssic = uen_to_ssic_map.get(uen_for_file, "")
    company_industry_description = ssic_definitions.get(company_ssic, "")

    if args.debug and uen_for_file:
        print(f"[INFO] UEN {uen_for_file} -> SSIC {company_ssic}")
        if company_industry_description:
            print(f"[INFO] Using 5-digit only industry context: '{company_industry_description[:100]}...'")

    def _score_row(i: int, t_text: str, d_text: str, edu_text: str, grp_hint_raw: str):
        code, occ_title, score, explain, top_5, search_type = best_match_duties_priority(
            t_text, d_text, defs, title_map, expert_map, min_s, edu_text, grp_hint_raw, company_industry_description
        )
        return i, code, occ_title, score, explain, top_5, search_type

    results = []
    if args.threads and args.threads > 1:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=args.threads) as ex:
            futs = [ex.submit(_score_row, i, t, d, e, g) for i, (t, d, e, g) in enumerate(zip(titles, duties, edus, groups))]
            for fut in as_completed(futs): results.append(fut.result())
        results.sort(key=lambda x: x[0])
    else:
        for i, (t, d, e, g) in enumerate(zip(titles, duties, edus, groups)):
            results.append(_score_row(i, t, d, e, g))

    out_codes, out_titles, detailed_report_data = [], [], []
    
    for i, code, occ_title, score, explain, top_5, search_type in results:
        final_code = code
        final_title = occ_title
        
        if score < min_s and not final_code.startswith("X"):
            final_code = "X1000"
            final_title = X_TITLE_MAP["X1000"]
        
        out_codes.append(final_code)
        out_titles.append(final_title)

        if args.detailed_report:
            base_info = {
                "Row": row_idxs[i] + 1, "Input Title": titles[i], "Input Duties": duties[i],
                "Search Type": search_type, "Industry Context": company_industry_description
            }
            if top_5:
                for rank, candidate in enumerate(top_5, 1):
                    report_row = base_info.copy()
                    report_row.update({"Rank": rank, "Candidate SSOC": candidate.get("code", ""), "Candidate Title": candidate.get("title", ""), "Score": round(candidate.get("score", 0) * 100, 2), "Explain": candidate.get("explain", "")})
                    detailed_report_data.append(report_row)
            else:
                report_row = base_info.copy()
                report_row.update({"Rank": 1, "Candidate SSOC": code, "Candidate Title": occ_title, "Score": round(score * 100, 2), "Explain": explain})
                detailed_report_data.append(report_row)

    # === MODIFIED: Create the output path using the original filename ===
    input_filename = os.path.basename(jobs_path)
    out_file_path = os.path.join(out_dir, input_filename)
    # ====================================================================
    
    try:
        saved = _write_back_copy(jobs_path, args.jobs_sheet, args.jobs_header_row, "SSOC 2015", "SSOC 2020", row_idxs, out_codes, out_titles, out_path=out_file_path)
    except Exception as e:
        print(f"[ERROR] {os.path.basename(jobs_path)}: Error writing workbook: {e}", file=sys.stderr)
        return "", None, 0, 0

    if args.detailed_report and detailed_report_data:
        _write_detailed_report(detailed_report_data, uen_for_file, original_stem, out_dir, ts)
        
    updated = sum(1 for _c, _t in zip(out_codes, out_titles) if _c or _t)
    total = len(out_codes)
    print(f"[OK] Processed '{os.path.basename(jobs_path)}' -> Saved as '{os.path.basename(saved)}' in output folder | rows updated: {updated}/{total}")
    if audit_path: print(f"      audit: {os.path.basename(audit_path)}")
    return saved, audit_path, updated, total
# ---------- main ----------
def main():
    parser = argparse.ArgumentParser(description="SSOC duties-first matcher (5-digit; SG sector taxonomy; guardrails + baked rules + title-only fallback + education weighting).")
    parser.add_argument("--defs", default=DEFAULT_DEFINITIONS_FILE, help="Path to SSOC 2024 definitions Excel")
    parser.add_argument("--jobs", default=None, help="Path to a single jobs Excel (optional if --jobs-dir is used)")
    parser.add_argument("--jobs-dir", default=None, help="Folder containing jobs Excel files (*.xlsx, *.xls)")
    parser.add_argument("--recursive", action="store_true", help="Recurse into subfolders of --jobs-dir")
    parser.add_argument("--out-dir", default=None, help="Folder to write outputs (default: create a timestamped folder)")

    parser.add_argument("--def-sheet", default=DEFAULT_DEF_SHEET, help="Definitions sheet name or index", type=str)
    parser.add_argument("--def-skip-rows", default=DEFAULT_DEF_SKIP_ROWS, type=int)

    parser.add_argument("--jobs-sheet", default=DEFAULT_JOBS_SHEET, help="Jobs sheet name or index", type=str)
    parser.add_argument("--jobs-header-row", default=DEFAULT_JOBS_HEADER_ROW, type=int, help="0-based header row index (A6 -> 5)")
    parser.add_argument("--title-col-name", default=DEFAULT_TITLE_COL_NAME)
    parser.add_argument("--edu-col-name", default=DEFAULT_EDU_COL_NAME)

    parser.add_argument("--title-col-index", default=DEFAULT_TITLE_COL_INDEX, type=lambda x: None if x in ("", "None") else int(x))
    parser.add_argument("--edu-col-index", default=DEFAULT_EDU_COL_INDEX, type=lambda x: None if x in ("", "None") else int(x))

    parser.add_argument("--min-score", default=DEFAULT_MIN_SCORE, type=float, help="0�100 threshold")
    parser.add_argument("--detailed-report", action="store_true", help="Generate a detailed Excel report with top 5 candidates for each job.")
    parser.add_argument("--debug", action="store_true", default=DEFAULT_DEBUG)
    parser.add_argument("--threads", type=int, default=1, help="Threads per file for row scoring (optional)")
    parser.add_argument("--file-threads", type=int, default=1, help="Parallelism across files for batch mode")

    parser.add_argument("--skip-unreadable", action="store_true", default=True,
                        help="Skip unreadable/corrupted Excel files and continue (default ON)")

    args = parser.parse_args()
    
    # comment away if do not want detailed reports
    # args.detailed_report = True

    def _sheet_arg(v):
        if v in (None, "", "None"): return None
        return int(v) if str(v).isdigit() else v

    def_sheet  = _sheet_arg(args.def_sheet)
    jobs_sheet = _sheet_arg(args.jobs_sheet)

    try:
        # === CORRECTED: Unpack the returned tuple into two variables ===
        defs, title_map = load_definitions(args.defs, def_sheet=def_sheet, def_skip_rows=args.def_skip_rows, debug=args.debug)
        # ===============================================================
    except Exception as e:
        print("Error loading definitions:", e, file=sys.stderr); sys.exit(1)
        
        
    expert_map = load_expert_map(DEFAULT_EXPERT_MAP_FILE, debug=args.debug)
    uen_to_ssic_map = load_uen_to_ssic_map(DEFAULT_SSIC_LIST_FILE, debug=args.debug)
    ssic_definitions = load_ssic_definitions(DEFAULT_SSIC_DEFS_FILE, debug=args.debug)

    global _TFIDF_VECT, _TFIDF_MAT, _TFIDF_TEXTS
    if _HAS_SK:
        # This line will now work correctly because `defs` is the list of records
        _TFIDF_TEXTS = [ (r.get("title_norm","") + " " + r.get("blob_norm","")).strip() for r in defs ]
        _TFIDF_VECT  = TfidfVectorizer(min_df=2, ngram_range=(1,2))
        _TFIDF_MAT   = _TFIDF_VECT.fit_transform(_TFIDF_TEXTS)

    _resolve_jobs_input(args)

    if args.jobs_dir:
        base_dir = os.path.abspath(args.jobs_dir)
        files = _list_jobs_files(base_dir, args.recursive, args.defs)
        if not files:
            print(f"No Excel files found in: {base_dir}", file=sys.stderr)
            sys.exit(1)

        if args.skip_unreadable:
            good = [f for f in files if _is_readable_excel(f)]
            bad_count = len(files) - len(good)
            if bad_count > 0:
                print(f"Skipping {bad_count} unreadable file(s).", file=sys.stderr)
            files = good

        out_dir = _build_out_dir(base_dir, args.out_dir)
        print(f"Found {len(files)} file(s). Outputs -> {out_dir}")

        results = []
        if args.file_threads and args.file_threads > 1:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            with ThreadPoolExecutor(max_workers=args.file_threads) as ex:
                # Pass the new title_map to the processing function
                futs = {ex.submit(process_single_file, f, defs, title_map, uen_to_ssic_map, ssic_definitions, args, out_dir): f for f in files}
                for fut in as_completed(futs):
                    results.append(fut.result())
        else:
            for f in files:
                # Pass the new title_map to the processing function
                results.append(process_single_file(f, defs, title_map, expert_map, uen_to_ssic_map, ssic_definitions, args, out_dir))

        total_files = len(files)
        ok_files = sum(1 for r in results if r[0])
        total_rows = sum(r[3] for r in results)
        total_updated = sum(r[2] for r in results)
        print(f"\nBatch done: {ok_files}/{total_files} files processed.")
        print(f"Rows updated (sum): {total_updated}/{total_rows}")
        print(f"Outputs in: {out_dir}")

    else:
        jobs_path = args.jobs or DEFAULT_JOBS_FILE
        if not os.path.exists(jobs_path):
            print(f"Jobs file not found: {jobs_path}", file=sys.stderr)
            sys.exit(1)
        base_dir = os.path.dirname(os.path.abspath(jobs_path)) or "."
        out_dir = _build_out_dir(base_dir, args.out_dir)
        # Pass the new title_map to the processing function
        process_single_file(f, defs, title_map, expert_map, uen_to_ssic_map, ssic_definitions, args, out_dir)
        
if __name__ == "__main__":
    main()

